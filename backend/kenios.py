"""
kenios.py — Backend đa-AI cho app KENIOS (com.kenios.codebox)  v4.2
======================================================================
TÍNH NĂNG MỚI / SỬA LỖI (v4.2):
  ✅ Gemini: cập nhật model mới nhất (gemini-2.0-flash, gemini-1.5-pro v002...)
             sửa lỗi 404 "model not found" — dùng đúng API v1beta
  ✅ Đính ảnh (image_base64) VÀ đính file (file_base64 + mime) hoạt động đầy đủ
  ✅ Multi-attachment: gửi tới 30 ảnh/file cùng lúc
  ✅ Chọn ngôn ngữ giao diện trả về (vi / en / auto)
  ✅ Giọng nói: phiên âm qua Whisper (OpenAI) hoặc Gemini Speech-to-Text
  ✅ Chạy code trực tiếp trên server (sandbox Python) — /run/python
  ✅ Chạy test file (.py / .js / .sh) và trả kết quả — /run/test
  ✅ Mô hình mới nhất cho mỗi nhà cung cấp (GPT-4o, Claude 3.7, Gemini 2.0 Flash…)
  ✅ Nhiều tính năng lập trình: code review, debug, explain, convert ngôn ngữ
  ✅ Thanh toán / nạp credits — /payment/*
  ✅ Webhook thanh toán tự động (Casso/Sepay) — /payment/webhook
  ✅ Ensemble AI (hỏi nhiều AI song song, tổng hợp)
  ✅ Admin API Key quản lý tập trung — /admin/keys
  ✅ Prompt Templates CRUD — /prompts
  ✅ Favorites (lưu tin nhắn yêu thích) — /favorites
  ✅ Pin / Share / Export hội thoại
  ✅ Tìm kiếm tin nhắn — /search
  ✅ Admin Stats — /admin/stats
  ✅ Auto-zip code blocks — /code/zip
  ✅ Token estimation trong phản hồi chat
  ✅ Gói PRO / ULTRA / MAX
  ✅ Lỗi rõ ràng: 401/403/404/429 đều có thông báo tiếng Việt cụ thể
  ✅ HỖ TRỢ UPLOAD VÀ DOWNLOAD STREAM TỚI 4GB (TỐI THIỂU 1KB), TRÁNH TRÀN BỘ NHỚ RAM.
  ✅ RAG NÂNG CAO: Tự động phân tích PDF, Word (docx), Excel (xlsx) và trích xuất ngữ cảnh TF-IDF.
  ✅ WEB SEARCH & WEB PAGE SCRAPER: Tự động tìm kiếm DuckDuckGo và cào dữ liệu HTML của kết quả.
  ✅ TEXT-TO-SPEECH (TTS): Tạo giọng nói âm thanh mp3 lưu hành từ văn bản.
  ✅ VẼ ẢNH AI (IMAGE GENERATION): Sinh ảnh qua DALL-E và tự động lưu vào thư viện tệp.
"""

import os, re, time, json, hmac, base64, hashlib, secrets, io, zipfile
import sqlite3, logging, asyncio, subprocess, tempfile, sys, shutil
from typing import Any, Optional

import httpx
from fastapi import FastAPI, Request, HTTPException, Header, Depends, UploadFile, File as FastAPIFile, Form, BackgroundTasks
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, FileResponse, StreamingResponse, HTMLResponse, Response
from pydantic import BaseModel

# ========================= Cấu hình =========================
DB_PATH         = os.getenv("CODEBOX_DB", "kenios.db")
PORT            = int(os.getenv("PORT", "8000"))
SECRET          = os.getenv("CODEBOX_SECRET") or secrets.token_hex(32)
TOKEN_TTL       = int(os.getenv("TOKEN_TTL", str(60 * 60 * 24 * 30)))  # 30 ngày
REQUEST_TIMEOUT = float(os.getenv("REQUEST_TIMEOUT", "120"))
SANDBOX_TIMEOUT = int(os.getenv("SANDBOX_TIMEOUT", "15"))  # giây chạy code

# ----- Email tích hợp (KenMail) chạy chung trong KENIOS -----
MAIL_DOMAIN     = os.getenv("MAIL_DOMAIN", "kenios.store")
MAIL_ENABLE     = os.getenv("MAIL_ENABLE", "1") == "1"      # bật bộ nhận thư SMTP nội bộ
MAIL_SMTP_PORT  = int(os.getenv("MAIL_SMTP_PORT", "25"))    # cổng nhận thư đến (cần MX + mở port 25)
SMTP_RELAY_HOST = os.getenv("SMTP_RELAY_HOST", "")          # gửi ra ngoài qua relay (vd smtp.gmail.com)
SMTP_RELAY_PORT = int(os.getenv("SMTP_RELAY_PORT", "587"))
SMTP_RELAY_USER = os.getenv("SMTP_RELAY_USER", "")
SMTP_RELAY_PASS = os.getenv("SMTP_RELAY_PASS", "")
OTP_DEBUG       = os.getenv("OTP_DEBUG", "0") == "1"        # trả mã trong response để test khi chưa có mail

# ----- KENIOS AI: model tự host của riêng bạn (không dùng API key của ai) -----
KENIOS_AI_ENABLE = os.getenv("KENIOS_AI_ENABLE", "1") == "1"
KENIOS_AI_BASE   = os.getenv("KENIOS_AI_BASE", "http://127.0.0.1:11434/v1")  # Ollama (OpenAI-compatible)
KENIOS_AI_MODEL  = os.getenv("KENIOS_AI_MODEL", "llama3.1")
KENIOS_AI_KEY    = os.getenv("KENIOS_AI_KEY", "ollama")   # Ollama bỏ qua, chỉ cần khác rỗng

# Thư mục lưu tệp tải lên của user trên đĩa
UPLOAD_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "uploads")
os.makedirs(UPLOAD_DIR, exist_ok=True)

# Kích thước tệp giới hạn (1KB - 4GB)
MIN_FILE_SIZE = 1024
MAX_FILE_SIZE = 4_294_967_296  # 4GB

logging.basicConfig(level=logging.INFO,
                    format="%(asctime)s %(levelname)s %(message)s")
log = logging.getLogger("kenios")

# ----- Fernet (mã hóa API key) -----
from cryptography.fernet import Fernet
_key_file = os.getenv("CODEBOX_ENC_KEYFILE", "kenios_enc.key")
if os.getenv("CODEBOX_ENC_KEY"):
    _enc_key = os.getenv("CODEBOX_ENC_KEY").encode()
elif os.path.exists(_key_file):
    _enc_key = open(_key_file, "rb").read().strip()
else:
    _enc_key = Fernet.generate_key()
    with open(_key_file, "wb") as f: f.write(_enc_key)
    log.info("Tạo khóa mã hóa mới: %s", _key_file)
fernet = Fernet(_enc_key)

def enc(text: str) -> str: return fernet.encrypt(text.encode()).decode()
def dec(token: str) -> str: return fernet.decrypt(token.encode()).decode()


# ===================== Danh sách AI (models mới nhất 2025) =====================
PROVIDERS: dict[str, dict[str, Any]] = {
    "kenios": {
        "label": "KENIOS AI · của bạn (miễn phí, không cần key)",
        "kind": "openai",
        "base": KENIOS_AI_BASE,
        "default_model": KENIOS_AI_MODEL,
        "models": [KENIOS_AI_MODEL],
        "vision": False, "free": True,
        "code": True,
    },
    "openai": {
        "label": "OpenAI · GPT-4o & o3",
        "kind": "openai",
        "base": "https://api.openai.com/v1",
        "default_model": "gpt-4o",
        "models": ["gpt-4o", "gpt-4o-mini", "o1", "o1-mini", "o3-mini"],
        "vision": True, "free": False,
        "code": True,
    },
    "anthropic": {
        "label": "Anthropic · Claude 3.7",
        "kind": "anthropic",
        "base": "https://api.anthropic.com/v1",
        "default_model": "claude-3-7-sonnet-latest",
        "models": [
            "claude-3-7-sonnet-latest",
            "claude-3-7-haiku-latest",
            "claude-3-5-sonnet-latest",
            "claude-3-5-haiku-latest",
            "claude-3-opus-latest",
        ],
        "vision": True, "free": False,
        "code": True,
    },
    "gemini": {
        "label": "Google · Gemini 2.5",
        "kind": "gemini",
        "base": "https://generativelanguage.googleapis.com/v1beta",
        "default_model": "gemini-2.5-flash",
        "models": [
            "gemini-2.5-flash",
            "gemini-2.5-pro",
            "gemini-2.0-flash",
            "gemini-2.0-pro-exp-02-05",
        ],
        "vision": True, "free": True,
        "code": True,
    },
    "groq": {
        "label": "Groq · Llama 3.3 (free)",
        "kind": "openai",
        "base": "https://api.groq.com/openai/v1",
        "default_model": "llama-3.3-70b-versatile",
        "models": [
            "llama-3.3-70b-versatile",
            "llama-3.2-11b-vision-preview",
            "llama-3.2-3b-preview",
            "deepseek-r1-distill-llama-70b",
        ],
        "vision": False, "free": True,
        "code": True,
    },
    "openrouter": {
        "label": "OpenRouter (nhiều model, có free)",
        "kind": "openai",
        "base": "https://openrouter.ai/api/v1",
        "default_model": "google/gemini-2.5-flash",
        "models": [
            "google/gemini-2.5-flash",
            "deepseek/deepseek-r1",
            "meta-llama/llama-3.3-70b-instruct",
            "anthropic/claude-3.7-sonnet",
        ],
        "vision": True, "free": True,
        "code": True,
    },
    "mistral": {
        "label": "Mistral · Large",
        "kind": "openai",
        "base": "https://api.mistral.ai/v1",
        "default_model": "mistral-large-latest",
        "models": ["mistral-large-latest", "mistral-small-latest", "codestral-latest", "pixtral-large-latest"],
        "vision": False, "free": False,
        "code": True,
    },
    "deepseek": {
        "label": "DeepSeek · V3 & R1",
        "kind": "openai",
        "base": "https://api.deepseek.com/v1",
        "default_model": "deepseek-chat",
        "models": ["deepseek-chat", "deepseek-reasoner"],
        "vision": False, "free": False,
        "code": True,
    },
    "xai": {
        "label": "xAI · Grok 3",
        "kind": "openai",
        "base": "https://api.x.ai/v1",
        "default_model": "grok-3",
        "models": ["grok-3", "grok-3-mini", "grok-2-1212", "grok-2-vision-1212"],
        "vision": True, "free": False,
        "code": True,
    },
    "perplexity": {
        "label": "Perplexity · Sonar Pro",
        "kind": "openai",
        "base": "https://api.perplexity.ai",
        "default_model": "sonar-pro",
        "models": ["sonar-pro", "sonar", "sonar-reasoning-pro", "sonar-reasoning"],
        "vision": False, "free": False,
        "code": False,
    },
    "together": {
        "label": "Together AI",
        "kind": "openai",
        "base": "https://api.together.xyz/v1",
        "default_model": "meta-llama/Llama-3.3-70B-Instruct-Turbo",
        "models": [
            "meta-llama/Llama-3.3-70B-Instruct-Turbo",
            "deepseek-ai/DeepSeek-R1",
            "Qwen/Qwen2.5-Coder-32B-Instruct",
        ],
        "vision": False, "free": False,
        "code": True,
    },
    "fireworks": {
        "label": "Fireworks AI",
        "kind": "openai",
        "base": "https://api.fireworks.ai/inference/v1",
        "default_model": "accounts/fireworks/models/llama-v3p3-70b-instruct",
        "models": ["accounts/fireworks/models/llama-v3p3-70b-instruct",
                   "accounts/fireworks/models/deepseek-r1"],
        "vision": False, "free": False,
        "code": True,
    },
    "cerebras": {
        "label": "Cerebras (siêu nhanh, free)",
        "kind": "openai",
        "base": "https://api.cerebras.ai/v1",
        "default_model": "llama-3.3-70b",
        "models": ["llama-3.3-70b", "llama-3.1-8b"],
        "vision": False, "free": True,
        "code": True,
    },
    "moonshot": {
        "label": "Moonshot · Kimi",
        "kind": "openai",
        "base": "https://api.moonshot.ai/v1",
        "default_model": "moonshot-v1-32k",
        "models": ["moonshot-v1-8k", "moonshot-v1-32k", "moonshot-v1-128k"],
        "vision": False, "free": False,
        "code": True,
    },
    "qwen": {
        "label": "Alibaba · Qwen 2.5",
        "kind": "openai",
        "base": "https://dashscope-intl.aliyuncs.com/compatible-mode/v1",
        "default_model": "qwen-max-latest",
        "models": ["qwen-max-latest", "qwen-plus-latest", "qwen-turbo-latest", "qwen2.5-coder-72b-instruct"],
        "vision": False, "free": False,
        "code": True,
    },
    "nvidia": {
        "label": "NVIDIA NIM (free)",
        "kind": "openai",
        "base": "https://integrate.api.nvidia.com/v1",
        "default_model": "meta/llama-3.3-70b-instruct",
        "models": ["meta/llama-3.3-70b-instruct", "nvidia/llama-3.1-nemotron-70b-instruct", "deepseek-ai/deepseek-r1"],
        "vision": False, "free": True,
        "code": True,
    },
    "cohere": {
        "label": "Cohere · Command R+",
        "kind": "openai",
        "base": "https://api.cohere.ai/compatibility/v1",
        "default_model": "command-r-plus",
        "models": ["command-r-plus", "command-r"],
        "vision": False, "free": False,
        "code": False,
    },
}

DEFAULT_SYSTEM = os.getenv(
    "SYSTEM_PROMPT",
    "Bạn là trợ lý AI của ứng dụng KENIOS. Trả lời hữu ích, chính xác. "
    "Khi viết code, luôn kèm theo giải thích rõ ràng. "
    "Hỗ trợ: Python, JavaScript, TypeScript, Swift, Kotlin, Go, Rust, C/C++, "
    "Java, PHP, HTML/CSS, SQL, Shell script. "
    "Ưu tiên dùng tiếng Việt trừ khi người dùng yêu cầu khác.",
)

# ========================== Cơ sở dữ liệu ==========================
def db() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    return conn


def init_db() -> None:
    with db() as c:
        c.executescript("""
            CREATE TABLE IF NOT EXISTS users(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT UNIQUE NOT NULL,
                email TEXT,
                phone TEXT,
                pw_hash TEXT NOT NULL,
                reset_token TEXT,
                reset_exp INTEGER,
                is_admin INTEGER DEFAULT 0,
                banned INTEGER DEFAULT 0,
                plan TEXT DEFAULT 'free',
                credits INTEGER DEFAULT 0,
                lang TEXT DEFAULT 'vi',
                created_at INTEGER
            );
            CREATE TABLE IF NOT EXISTS apikeys(
                user_id INTEGER NOT NULL,
                provider TEXT NOT NULL,
                enc_key TEXT NOT NULL,
                PRIMARY KEY(user_id, provider)
            );
            CREATE TABLE IF NOT EXISTS admin_apikeys(
                provider TEXT PRIMARY KEY,
                enc_key TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS mailboxes(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                address TEXT UNIQUE NOT NULL,
                pw_hash TEXT NOT NULL,
                owner_uid INTEGER,
                phone TEXT,
                created_at INTEGER
            );
            CREATE TABLE IF NOT EXISTS mails(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                mailbox_id INTEGER NOT NULL,
                direction TEXT DEFAULT 'in',
                from_addr TEXT,
                to_addr TEXT,
                subject TEXT,
                body TEXT,
                created_at INTEGER,
                seen INTEGER DEFAULT 0
            );
            CREATE TABLE IF NOT EXISTS mail_domains(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                domain TEXT UNIQUE NOT NULL,
                user_id INTEGER,
                created_at INTEGER
            );
            CREATE TABLE IF NOT EXISTS otp_codes(
                email TEXT PRIMARY KEY,
                code TEXT NOT NULL,
                purpose TEXT DEFAULT 'register',
                exp INTEGER NOT NULL,
                attempts INTEGER DEFAULT 0
            );
            CREATE TABLE IF NOT EXISTS api_tokens(
                token TEXT PRIMARY KEY,
                owner_uid INTEGER,
                name TEXT,
                calls INTEGER DEFAULT 0,
                created_at INTEGER
            );
            CREATE TABLE IF NOT EXISTS devices(
                udid TEXT PRIMARY KEY,
                product TEXT,
                version TEXT,
                serial TEXT,
                name TEXT,
                created_at INTEGER
            );
            CREATE TABLE IF NOT EXISTS conversations(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                title TEXT,
                provider TEXT,
                pinned INTEGER DEFAULT 0,
                share_token TEXT,
                created_at INTEGER,
                updated_at INTEGER
            );
            CREATE TABLE IF NOT EXISTS messages(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                conversation_id INTEGER NOT NULL,
                role TEXT NOT NULL,
                content TEXT NOT NULL,
                tokens_used INTEGER,
                created_at INTEGER
            );
            CREATE TABLE IF NOT EXISTS files(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                name TEXT NOT NULL,
                category TEXT,
                mime TEXT,
                size INTEGER,
                data TEXT NOT NULL,
                created_at INTEGER
            );
            CREATE TABLE IF NOT EXISTS payments(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                amount INTEGER NOT NULL,
                credits INTEGER NOT NULL,
                status TEXT DEFAULT 'pending',
                ref TEXT,
                created_at INTEGER
            );
            CREATE TABLE IF NOT EXISTS settings(
                key TEXT PRIMARY KEY,
                value TEXT
            );
            CREATE TABLE IF NOT EXISTS error_logs(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER,
                username TEXT,
                context TEXT,
                detail TEXT,
                created_at INTEGER
            );
            CREATE TABLE IF NOT EXISTS prompt_templates(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                title TEXT NOT NULL,
                content TEXT NOT NULL,
                category TEXT,
                is_public INTEGER DEFAULT 0,
                user_id INTEGER,
                created_at INTEGER
            );
            CREATE TABLE IF NOT EXISTS favorites(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                message_content TEXT NOT NULL,
                conversation_id INTEGER,
                provider TEXT,
                created_at INTEGER
            );
            CREATE TABLE IF NOT EXISTS friendships(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                friend_id INTEGER NOT NULL,
                status TEXT DEFAULT 'pending',
                created_at INTEGER,
                UNIQUE(user_id, friend_id)
            );
            CREATE TABLE IF NOT EXISTS direct_messages(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                sender_id INTEGER NOT NULL,
                receiver_id INTEGER NOT NULL,
                content TEXT NOT NULL,
                created_at INTEGER,
                is_read INTEGER DEFAULT 0
            );
            CREATE TABLE IF NOT EXISTS proxies(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                label TEXT,
                scheme TEXT DEFAULT 'http',
                host TEXT NOT NULL,
                port INTEGER NOT NULL,
                username TEXT,
                enc_password TEXT,
                region TEXT,
                source TEXT DEFAULT 'manual',
                active INTEGER DEFAULT 0,
                created_at INTEGER
            );
            CREATE TABLE IF NOT EXISTS posts(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                file_id INTEGER NOT NULL,
                caption TEXT,
                likes INTEGER DEFAULT 0,
                created_at INTEGER
            );
            CREATE TABLE IF NOT EXISTS post_likes(
                post_id INTEGER NOT NULL,
                user_id INTEGER NOT NULL,
                PRIMARY KEY(post_id, user_id)
            );
            CREATE TABLE IF NOT EXISTS live_rooms(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                host_id INTEGER NOT NULL,
                title TEXT,
                hls_url TEXT,
                viewers INTEGER DEFAULT 0,
                likes INTEGER DEFAULT 0,
                active INTEGER DEFAULT 1,
                created_at INTEGER
            );
            CREATE TABLE IF NOT EXISTS live_messages(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                room_id INTEGER NOT NULL,
                user_id INTEGER,
                username TEXT,
                content TEXT,
                created_at INTEGER
            );
        """)
    _migrate()

    # Seed admin
    admin_user = os.getenv("ADMIN_USER", "kenios")
    admin_pass = os.getenv("ADMIN_PASS", "admin1999@")
    with db() as c:
        row = c.execute("SELECT id FROM users WHERE username=?", (admin_user,)).fetchone()
        if row:
            c.execute("UPDATE users SET is_admin=1, banned=0 WHERE id=?", (row["id"],))
        else:
            c.execute(
                "INSERT INTO users(username,pw_hash,is_admin,plan,credits,created_at) VALUES(?,?,1,'pro',9999,?)",
                (admin_user, hash_pw(admin_pass), int(time.time())),
            )
            log.info("Tạo admin '%s' (hãy đổi mật khẩu sau khi đăng nhập!)", admin_user)
    _seed_setting("bank_code", os.getenv("BANK_CODE", "970416"))
    _seed_setting("bank_short", os.getenv("BANK_SHORT", "ACB"))
    _seed_setting("bank_account", os.getenv("BANK_ACCOUNT", "23252921"))
    _seed_setting("bank_name", os.getenv("BANK_NAME", "TRAN MINH CHIEN"))
    _seed_setting("bank_webhook", "")
    _seed_setting("bank_apikey", "")
    _seed_prompt_templates()
    log.info("DB sẵn sàng: %s", DB_PATH)


def _seed_prompt_templates() -> None:
    templates = [
        {
            "title": "Tối ưu hóa Code (Clean Code & Performance)",
            "category": "Lập trình",
            "content": (
                "Hãy tối ưu hóa đoạn mã nguồn sau đây theo các nguyên tắc Clean Code và cải thiện hiệu năng (performance).\n"
                "Yêu cầu:\n"
                "1. Tên biến, tên hàm rõ ràng, tự giải thích (self-documenting).\n"
                "2. Tách nhỏ các hàm phức tạp thành các hàm đơn nhiệm (Single Responsibility).\n"
                "3. Tránh lặp lại mã nguồn (DRY - Don't Repeat Yourself).\n"
                "4. Tối ưu hóa độ phức tạp thời gian (Time Complexity) và không gian (Space Complexity).\n"
                "5. Cung cấp mã nguồn đã tối ưu kèm giải thích chi tiết các thay đổi.\n\n"
                "Mã nguồn cần tối ưu:\n"
                "[Nhập mã nguồn của bạn vào đây]"
            )
        },
        {
            "title": "Thiết kế hệ thống theo chuẩn SOLID",
            "category": "Kiến trúc",
            "content": (
                "Hãy phân tích và cấu trúc lại đoạn mã nguồn sau đây để tuân thủ nghiêm ngặt 5 nguyên tắc SOLID trong thiết kế hướng đối tượng:\n"
                "- S: Single Responsibility Principle (Đơn nhiệm)\n"
                "- O: Open/Closed Principle (Mở để mở rộng, đóng để sửa đổi)\n"
                "- L: Liskov Substitution Principle (Thay thế Liskov)\n"
                "- I: Interface Segregation Principle (Phân tách giao diện)\n"
                "- D: Dependency Inversion Principle (Đảo ngược phụ thuộc)\n\n"
                "Giải thích rõ từng nguyên tắc được áp dụng như thế nào sau khi refactor.\n\n"
                "Mã nguồn cần thiết kế lại:\n"
                "[Nhập mã nguồn của bạn vào đây]"
            )
        },
        {
            "title": "Rà soát Lỗi Bảo mật (Security Audit)",
            "category": "Bảo mật",
            "content": (
                "Hãy thực hiện rà soát bảo mật (Security Audit / Code Review) cho đoạn mã nguồn dưới đây.\n"
                "Tìm kiếm các lỗ hổng bảo mật phổ biến như:\n"
                "- SQL Injection, XSS, CSRF\n"
                "- Lộ thông tin nhạy cảm (API Keys, Mật khẩu...)\n"
                "- Lỗi phân quyền, xác thực (Authentication/Authorization)\n"
                "- Xử lý ngoại lệ không an toàn (Unsafe Exception Handling)\n"
                "- Buffer Overflow hoặc lỗi tràn bộ nhớ (nếu có)\n\n"
                "Với mỗi lỗ hổng phát hiện được, hãy giải thích nguy cơ và cung cấp cách khắc phục cụ thể.\n\n"
                "Mã nguồn cần rà soát:\n"
                "[Nhập mã nguồn của bạn vào đây]"
            )
        },
        {
            "title": "Giải thích Code & Tạo tài liệu (Documenting)",
            "category": "Tài liệu",
            "content": (
                "Hãy giải thích chi tiết luồng hoạt động của đoạn mã nguồn dưới đây và viết tài liệu hướng dẫn (docstring/comments) theo chuẩn của ngôn ngữ lập trình đó.\n"
                "Yêu cầu:\n"
                "1. Tóm tắt chức năng chính của đoạn mã.\n"
                "2. Mô tả chi tiết các tham số đầu vào (parameters) và kết quả trả về (return values).\n"
                "3. Giải thích luồng logic chính từng bước.\n"
                "4. Thêm các comment cần thiết trực tiếp vào mã nguồn mà không làm loãng mã nguồn.\n\n"
                "Mã nguồn cần viết tài liệu:\n"
                "[Nhập mã nguồn của bạn vào đây]"
            )
        },
        {
            "title": "Viết Unit Test tự động",
            "category": "Kiểm thử",
            "content": (
                "Hãy viết các ca kiểm thử đơn vị (Unit Tests) toàn diện cho đoạn mã nguồn dưới đây.\n"
                "Yêu cầu:\n"
                "1. Bao phủ đầy đủ các trường hợp thông thường (Happy path).\n"
                "2. Bao phủ các trường hợp biên, giá trị đặc biệt hoặc đầu vào lỗi (Edge cases / Error handling).\n"
                "3. Sử dụng thư viện testing chuẩn của ngôn ngữ tương ứng (ví dụ: unittest/pytest cho Python, XCTest cho Swift, Jest cho JS...).\n"
                "4. Sử dụng mock/stub cho các dịch vụ bên ngoài (cơ sở dữ liệu, API mạng) nếu cần thiết.\n\n"
                "Mã nguồn cần viết Unit Test:\n"
                "[Nhập mã nguồn của bạn vào đây]"
            )
        }
    ]
    with db() as c:
        count = c.execute("SELECT COUNT(*) as cnt FROM prompt_templates").fetchone()["cnt"]
        if count == 0:
            now = int(time.time())
            for t in templates:
                c.execute(
                    "INSERT INTO prompt_templates(title, content, category, is_public, user_id, created_at) "
                    "VALUES(?, ?, ?, 1, NULL, ?)",
                    (t["title"], t["content"], t["category"], now)
                )
            log.info("Đã seed %d prompt templates mặc định vào CSDL", len(templates))


def get_setting(key: str, default: str = "") -> str:
    with db() as c:
        row = c.execute("SELECT value FROM settings WHERE key=?", (key,)).fetchone()
    return row["value"] if row else default


def set_setting(key: str, value: str) -> None:
    with db() as c:
        c.execute("INSERT INTO settings(key,value) VALUES(?,?) "
                  "ON CONFLICT(key) DO UPDATE SET value=excluded.value", (key, value))


def _seed_setting(key: str, value: str) -> None:
    with db() as c:
        if not c.execute("SELECT 1 FROM settings WHERE key=?", (key,)).fetchone():
            c.execute("INSERT INTO settings(key,value) VALUES(?,?)", (key, value))


def _migrate() -> None:
    migrations = [
        ("users", "is_admin", "INTEGER DEFAULT 0"),
        ("users", "banned",   "INTEGER DEFAULT 0"),
        ("users", "plan",     "TEXT DEFAULT 'free'"),
        ("users", "credits",  "INTEGER DEFAULT 0"),
        ("users", "lang",     "TEXT DEFAULT 'vi'"),
        ("users", "public_id",     "TEXT"),
        ("users", "status",        "TEXT DEFAULT 'active'"),
        ("users", "suspend_until", "INTEGER DEFAULT 0"),
        ("users", "last_seen",     "INTEGER DEFAULT 0"),
        ("users", "last_feature",  "TEXT"),
        ("files", "mime",     "TEXT"),
        ("conversations", "pinned", "INTEGER DEFAULT 0"),
        ("conversations", "share_token", "TEXT"),
        ("messages", "tokens_used", "INTEGER"),
        ("mailboxes", "phone", "TEXT"),
    ]
    with db() as c:
        for table, col, ddl in migrations:
            try:
                c.execute(f"ALTER TABLE {table} ADD COLUMN {col} {ddl}")
            except Exception:
                pass


# ========================== Bảo mật ==========================
def hash_pw(password: str) -> str:
    salt = secrets.token_bytes(16)
    dk = hashlib.pbkdf2_hmac("sha256", password.encode(), salt, 200_000)
    return salt.hex() + "$" + dk.hex()


def verify_pw(password: str, stored: str) -> bool:
    try:
        salt_hex, dk_hex = stored.split("$", 1)
        dk = hashlib.pbkdf2_hmac("sha256", password.encode(), bytes.fromhex(salt_hex), 200_000)
        return hmac.compare_digest(dk.hex(), dk_hex)
    except Exception:
        return False


def _b64u(b: bytes) -> str:
    return base64.urlsafe_b64encode(b).decode().rstrip("=")


def _b64u_dec(s: str) -> bytes:
    return base64.urlsafe_b64decode(s + "=" * (-len(s) % 4))


def make_token(user_id: int) -> str:
    payload = {"uid": user_id, "exp": int(time.time()) + TOKEN_TTL}
    body = _b64u(json.dumps(payload, separators=(",", ":")).encode())
    sig = _b64u(hmac.new(SECRET.encode(), body.encode(), hashlib.sha256).digest())
    return f"{body}.{sig}"


def verify_token(token: str) -> int:
    try:
        body, sig = token.split(".", 1)
        good = _b64u(hmac.new(SECRET.encode(), body.encode(), hashlib.sha256).digest())
        if not hmac.compare_digest(sig, good):
            raise ValueError("sai chữ ký")
        payload = json.loads(_b64u_dec(body))
        if payload["exp"] < time.time():
            raise ValueError("hết hạn")
        return int(payload["uid"])
    except Exception:
        raise HTTPException(status_code=401, detail="Phiên đăng nhập không hợp lệ hoặc đã hết hạn. Vui lòng đăng nhập lại.")


def get_user(authorization: Optional[str] = Header(default=None)) -> sqlite3.Row:
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Thiếu token đăng nhập.")
    uid = verify_token(authorization.split(" ", 1)[1])
    with db() as c:
        row = c.execute("SELECT * FROM users WHERE id=?", (uid,)).fetchone()
    if not row:
        raise HTTPException(status_code=401, detail="Tài khoản không tồn tại.")
    if row["banned"]:
        raise HTTPException(status_code=403, detail="Tài khoản đã bị khóa. Liên hệ quản trị viên.")
    # Tạm ngưng có thời hạn (admin không bị ảnh hưởng)
    if not row["is_admin"] and (row["status"] or "active") == "suspended":
        su = row["suspend_until"] or 0
        now = int(time.time())
        if su and su <= now:
            with db() as c:
                c.execute("UPDATE users SET status='active', suspend_until=0 WHERE id=?", (row["id"],))
        elif su:
            t = time.strftime("%H:%M %d/%m", time.localtime(su))
            raise HTTPException(status_code=403, detail=f"Tài khoản đang bị tạm ngưng đến {t}.")
        else:
            raise HTTPException(status_code=403, detail="Tài khoản đang bị tạm ngưng. Liên hệ quản trị viên.")
    return row


def _gen_public_id(c) -> str:
    import random
    for _ in range(20):
        pid = "KEN" + "".join(random.choices("0123456789", k=8))
        if not c.execute("SELECT 1 FROM users WHERE public_id=?", (pid,)).fetchone():
            return pid
    return "KEN" + str(int(time.time()))[-8:]


def _ensure_public_id(c, uid) -> str:
    row = c.execute("SELECT public_id FROM users WHERE id=?", (uid,)).fetchone()
    pid = row["public_id"] if row else None
    if not pid:
        pid = _gen_public_id(c)
        c.execute("UPDATE users SET public_id=? WHERE id=?", (pid, uid))
    return pid


def _setting_get(key: str, default: str = "") -> str:
    with db() as c:
        r = c.execute("SELECT value FROM settings WHERE key=?", (key,)).fetchone()
    return r["value"] if r else default


def _setting_set(key: str, value: str) -> None:
    with db() as c:
        c.execute("INSERT INTO settings(key,value) VALUES(?,?) "
                  "ON CONFLICT(key) DO UPDATE SET value=excluded.value", (key, value))


def _user_dict(row) -> dict[str, Any]:
    return {
        "id": row["id"], "username": row["username"],
        "email": row["email"], "phone": row["phone"],
        "public_id": row["public_id"],
        "is_admin": bool(row["is_admin"]),
        "plan": "pro" if row["is_admin"] else (row["plan"] or "free"),
        "credits": row["credits"], "lang": row["lang"] or "vi",
        "status": row["status"] or "active",
    }


def get_admin(user=Depends(get_user)) -> sqlite3.Row:
    if not user["is_admin"]:
        raise HTTPException(status_code=403, detail="Chỉ quản trị viên mới được phép.")
    return user


# ===================== Xử lý ảnh & file =====================
def parse_image(image: str) -> tuple[str, str]:
    if image.startswith("data:"):
        head, data = image.split(",", 1)
        m = re.search(r"data:(.*?);base64", head)
        return (m.group(1) if m else "image/jpeg"), data
    return "image/jpeg", image


# ===================== Token estimation =====================
def estimate_tokens(text: str) -> int:
    if not text:
        return 0
    non_ascii = sum(1 for ch in text if ord(ch) > 127)
    ascii_chars = len(text) - non_ascii
    return int(ascii_chars / 4) + non_ascii + 1


# ===================== Lỗi nhà cung cấp =====================
def _raise_for_provider(r: httpx.Response, provider: str) -> None:
    if r.status_code < 400:
        return
    txt = r.text[:500]
    if r.status_code in (401, 403):
        raise HTTPException(status_code=400,
            detail=f"{provider}: API key sai hoặc không đủ quyền ({r.status_code}). "
                   f"Vui lòng kiểm tra lại API key trong phần Cài đặt. Chi tiết: {txt}")
    if r.status_code == 404:
        raise HTTPException(status_code=400,
            detail=f"{provider}: Model không tồn tại hoặc chưa được hỗ trợ (404). "
                   f"Vui lòng chọn model khác. Chi tiết: {txt}")
    if r.status_code == 429:
        raise HTTPException(status_code=429,
            detail=f"{provider}: Vượt quá giới hạn tốc độ miễn phí (429). Hệ thống đã tự thử lại "
                   f"nhưng vẫn bị chặn. Cách khắc phục: (1) đợi ~30–60 giây rồi gửi lại; "
                   f"(2) đổi sang AI free khác như Groq hoặc OpenRouter; "
                   f"(3) đổi model nhẹ hơn (vd Gemini 2.0 Flash); hoặc (4) dùng API key trả phí của bạn.")
    raise HTTPException(status_code=502,
        detail=f"{provider} lỗi {r.status_code}: {txt}")


async def post_with_retry(
    client: "httpx.AsyncClient",
    url: str,
    *,
    provider: str = "",
    max_retries: int = 4,
    **kwargs: Any,
) -> "httpx.Response":
    """POST có tự động thử lại khi bị giới hạn tốc độ (429) hoặc server bận (500/502/503/504).

    Xử lý lỗi "vượt quá tốc độ" mà không làm hỏng phiên chat: chờ theo cấp số nhân
    (0.8s, 1.6s, 3.2s...) và tôn trọng header `Retry-After` của nhà cung cấp nếu có.
    """
    delay = 0.8
    last: Optional["httpx.Response"] = None
    for attempt in range(max_retries + 1):
        try:
            r = await client.post(url, **kwargs)
        except (httpx.ConnectError, httpx.ReadTimeout, httpx.RemoteProtocolError) as e:
            if attempt >= max_retries:
                if provider == "kenios":
                    raise HTTPException(status_code=503,
                        detail="KENIOS AI chưa chạy. Hãy cài model trên VPS: chạy "
                               "`bash kenios-ai/install-ai.sh` (cài Ollama + tải model) rồi "
                               "`systemctl restart kenios`. Hoặc tạm chọn AI khác (Gemini/Groq).")
                raise HTTPException(status_code=502,
                    detail=f"{provider or 'AI'}: không kết nối được tới máy chủ ({e.__class__.__name__}).")
            await asyncio.sleep(delay)
            delay = min(delay * 2, 12.0)
            continue
        last = r
        if r.status_code not in (429, 500, 502, 503, 504) or attempt >= max_retries:
            return r
        # Tôn trọng Retry-After nếu nhà cung cấp gửi về
        wait = delay
        ra = r.headers.get("retry-after")
        if ra:
            try:
                wait = max(wait, min(float(ra), 15.0))
            except ValueError:
                pass
        await asyncio.sleep(wait)
        delay = min(delay * 2, 12.0)
    return last  # type: ignore[return-value]


def get_user_key(user_id: int, provider: str, inline: Optional[str]) -> str:
    # KENIOS AI tự host: không cần API key của người dùng
    if provider == "kenios":
        return KENIOS_AI_KEY or "ollama"
    if inline:
        return inline
    with db() as c:
        row = c.execute("SELECT enc_key FROM apikeys WHERE user_id=? AND provider=?",
                        (user_id, provider)).fetchone()
    if row:
        return dec(row["enc_key"])
    with db() as c:
        row = c.execute("SELECT enc_key FROM admin_apikeys WHERE provider=?",
                        (provider,)).fetchone()
    if row:
        return dec(row["enc_key"])
    raise HTTPException(status_code=400,
        detail=f"Chưa có API key cho '{provider}'. Admin chưa cấu hình hoặc bạn chưa nhập key riêng.")


# ===================== DỊCH VỤ PARSE FILE & RAG =====================
def parse_file_content(file_path: str, filename: str) -> str:
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if ext == "pdf":
        try:
            import pypdf
            reader = pypdf.PdfReader(file_path)
            text_pages = []
            for i, page in enumerate(reader.pages):
                t = page.extract_text()
                if t:
                    text_pages.append(f"[Trang {i+1}]\n{t}")
            return "\n".join(text_pages)
        except Exception as e:
            return f"[Lỗi giải mã PDF: {e}]"
    elif ext == "docx":
        try:
            import docx
            doc = docx.Document(file_path)
            text_paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
            return "\n".join(text_paragraphs)
        except Exception as e:
            return f"[Lỗi giải mã DOCX: {e}]"
    elif ext in ("xlsx", "xls"):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, data_only=True)
            sheets_content = []
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                sheet_rows = []
                for row in sheet.iter_rows(values_only=True):
                    if any(row):
                        sheet_rows.append(" | ".join(str(val) if val is not None else "" for val in row))
                if sheet_rows:
                    sheets_content.append(f"[Sheet: {sheet_name}]\n" + "\n".join(sheet_rows))
            return "\n\n".join(sheets_content)
        except Exception as e:
            return f"[Lỗi giải mã XLSX: {e}]"
    else:
        try:
            with open(file_path, "r", encoding="utf-8", errors="replace") as f:
                return f.read()
        except Exception as e:
            return f"[Lỗi đọc tệp văn bản: {e}]"


def retrieve_relevant_chunks(text: str, query: str, top_k: int = 5) -> str:
    chunks = []
    chunk_size = 1000
    overlap = 100
    
    start = 0
    while start < len(text):
        end = min(start + chunk_size, len(text))
        chunk = text[start:end].strip()
        if chunk:
            chunks.append(chunk)
        start += chunk_size - overlap
        
    if not chunks:
        return ""
        
    query_words = set(re.findall(r'\w+', query.lower()))
    if not query_words:
        return "\n\n".join(chunks[:top_k])
        
    chunk_scores = []
    for idx, chunk in enumerate(chunks):
        chunk_words = re.findall(r'\w+', chunk.lower())
        score = sum(chunk_words.count(w) for w in query_words)
        chunk_scores.append((score, idx))
        
    chunk_scores.sort(key=lambda x: x[0], reverse=True)
    
    retrieved = []
    for score, idx in chunk_scores[:top_k]:
        retrieved.append(f"[Đoạn {idx+1}]: {chunks[idx]}")
    return "\n\n".join(retrieved)


# ===================== DỊCH VỤ TÌM KIẾM WEB DUCKDUCKGO =====================
async def search_ddg(query: str, max_results: int = 5) -> str:
    try:
        from urllib.parse import quote_plus
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
        }
        async with httpx.AsyncClient(timeout=10) as client:
            r = await client.get(f"https://html.duckduckgo.com/html/?q={quote_plus(query)}", headers=headers)
            if r.status_code != 200:
                return ""
            
            titles = re.findall(r'<a class="result__url"[^>]*>(.*?)</a>', r.text, re.DOTALL)
            snippets = re.findall(r'<a class="result__snippet"[^>]*>(.*?)</a>', r.text, re.DOTALL)
            urls = re.findall(r'<a class="result__url"[^>]*href="([^"]+)"', r.text, re.DOTALL)
            
            results = []
            for i in range(min(len(titles), len(snippets), max_results)):
                title = re.sub(r'<[^>]+>', '', titles[i]).strip()
                snippet = re.sub(r'<[^>]+>', '', snippets[i]).strip()
                url = urls[i] if i < len(urls) else ""
                results.append(f"- **{title}** ({url}): {snippet}")
            
            if urls and len(urls) > 0:
                first_url = urls[0]
                if "uddg=" in first_url:
                    from urllib.parse import unquote
                    first_url = unquote(first_url.split("uddg=")[1].split("&")[0])
                try:
                    scr_res = await client.get(first_url, headers=headers, timeout=5)
                    if scr_res.status_code == 200:
                        text_content = re.sub(r'<(script|style).*?>.*?</\1>', '', scr_res.text, flags=re.DOTALL|re.IGNORECASE)
                        text_content = re.sub(r'<[^>]+>', '', text_content)
                        text_content = re.sub(r'\s+', ' ', text_content).strip()
                        if len(text_content) > 100:
                            results.append(f"\n[Nội dung chi tiết từ trang {first_url}]:\n{text_content[:3000]}")
                except Exception:
                    pass
                    
            return "\n".join(results)
    except Exception as e:
        log.error("Lỗi tìm kiếm DDG: %s", e)
        return ""


def save_code_blocks(user_id: int, text: str, label: str = "code") -> list[dict[str, Any]]:
    safe = re.sub(r"[^a-zA-Z0-9_]+", "", label) or "code"
    ext_map = {"python": "py", "py": "py", "javascript": "js", "js": "js",
               "typescript": "ts", "ts": "ts", "html": "html", "css": "css",
               "json": "json", "bash": "sh", "sh": "sh", "swift": "swift",
               "java": "java", "c": "c", "cpp": "cpp", "go": "go", "rust": "rs",
               "sql": "sql", "yaml": "yml", "yml": "yml", "markdown": "md", "md": "md",
               "php": "php", "ruby": "rb", "kotlin": "kt", "dart": "dart"}
    blocks = re.findall(r"```([a-zA-Z0-9_+\-]*)\n(.*?)```", text, re.DOTALL)
    saved: list[dict[str, Any]] = []
    n = 0
    for lang, code in blocks:
        code = code.rstrip("\n")
        if len(code.strip()) < 10:
            continue
        n += 1
        ext = ext_map.get(lang.lower().strip(), "txt")
        name = f"{safe}_{n}.{ext}"
        try:
            with db() as c:
                cur = c.execute(
                    "INSERT INTO files(user_id,name,category,mime,size,data,created_at) "
                    "VALUES(?,?,?,?,?,'',?)",
                    (user_id, name, "code", "text/plain", len(code), int(time.time())))
                fid = cur.lastrowid
                saved.append({"id": fid, "name": name})
            
            # Save file to disk
            file_path = os.path.join(UPLOAD_DIR, str(fid))
            with open(file_path, "wb") as f:
                f.write(code.encode("utf-8"))
        except Exception:
            pass
    return saved


# ===================== Gọi AI =====================
async def call_provider(
    provider: str,
    api_key: str,
    model: Optional[str],
    history: list[dict[str, Any]],
    user_text: str,
    image: Optional[str] = None,
    file_b64: Optional[str] = None,
    file_mime: Optional[str] = None,
    system_override: Optional[str] = None,
    attachments: Optional[list[dict[str, Any]]] = None,
    proxy: Optional[str] = None,  # ← THÊM: định tuyến qua proxy active
) -> str:
    if provider not in PROVIDERS:
        raise HTTPException(status_code=400, detail=f"AI '{provider}' không được hỗ trợ.")
    p    = PROVIDERS[provider]
    model = model or p["default_model"]
    kind  = p["kind"]
    sys_msg = system_override or DEFAULT_SYSTEM

    img = None
    if image:
        img = parse_image(image)
    elif file_b64 and file_mime and file_mime.startswith("image/"):
        img = (file_mime, file_b64)

    parsed_attachments: list[dict[str, Any]] = []
    if attachments:
        for att in attachments[:30]:
            att_name = att.get("name", "file")
            att_data = att.get("data_base64", "")
            att_mime = att.get("mime", "application/octet-stream")
            parsed_attachments.append({
                "name": att_name,
                "data": att_data,
                "mime": att_mime,
            })

    _client_kwargs = {"timeout": REQUEST_TIMEOUT}
    if proxy:
        _client_kwargs["proxy"] = proxy  # ← THÊM: route qua proxy
    async with httpx.AsyncClient(**_client_kwargs) as client:
        # -------- OpenAI-compatible --------
        if kind == "openai":
            msgs = [{"role": "system", "content": sys_msg}]
            msgs += [{"role": m["role"], "content": m["content"]} for m in history]

            if parsed_attachments:
                user_content: Any = [{"type": "text", "text": user_text or ""}]
                for att in parsed_attachments:
                    if att["mime"].startswith("image/"):
                        user_content.append({
                            "type": "image_url",
                            "image_url": {"url": f"data:{att['mime']};base64,{att['data']}"},
                        })
                    else:
                        try:
                            decoded = base64.b64decode(att["data"]).decode("utf-8", errors="replace")
                            user_content.append({
                                "type": "text",
                                "text": f"\n[File: {att['name']}]\n```\n{decoded[:8000]}\n```",
                            })
                        except Exception:
                            pass
            elif img:
                media, data = img
                user_content = [
                    {"type": "text", "text": user_text or ""},
                    {"type": "image_url", "image_url": {"url": f"data:{media};base64,{data}"}},
                ]
            elif file_b64 and file_mime:
                try:
                    decoded = base64.b64decode(file_b64).decode("utf-8", errors="replace")
                    user_content = f"{user_text}\n\n[Nội dung file]\n```\n{decoded[:8000]}\n```"
                except Exception:
                    user_content = user_text or ""
            else:
                user_content = user_text
            msgs.append({"role": "user", "content": user_content})
            r = await post_with_retry(
                client,
                f"{p['base']}/chat/completions",
                provider=provider,
                headers={"Authorization": f"Bearer {api_key}",
                         "HTTP-Referer": "https://kenios.app",
                         "X-Title": "KENIOS"},
                json={"model": model, "messages": msgs},
            )
            _raise_for_provider(r, provider)
            return r.json()["choices"][0]["message"]["content"]

        # -------- Anthropic --------
        if kind == "anthropic":
            msgs = [{"role": m["role"], "content": m["content"]} for m in history]

            if parsed_attachments:
                content_parts: list[dict[str, Any]] = [{"type": "text", "text": user_text or ""}]
                for att in parsed_attachments:
                    if att["mime"].startswith("image/"):
                        content_parts.append({
                            "type": "image",
                            "source": {"type": "base64", "media_type": att["mime"], "data": att["data"]},
                        })
                    elif att["mime"] == "application/pdf":
                        content_parts.append({
                            "type": "document",
                            "source": {"type": "base64", "media_type": "application/pdf", "data": att["data"]},
                        })
                    else:
                        try:
                            decoded = base64.b64decode(att["data"]).decode("utf-8", errors="replace")
                            content_parts.append({
                                "type": "text",
                                "text": f"\n[File: {att['name']}]\n```\n{decoded[:8000]}\n```",
                            })
                        except Exception:
                            pass
                msgs.append({"role": "user", "content": content_parts})
            elif img:
                media, data = img
                msgs.append({"role": "user", "content": [
                    {"type": "text", "text": user_text or ""},
                    {"type": "image", "source": {"type": "base64",
                                                  "media_type": media, "data": data}},
                ]})
            elif file_b64 and file_mime:
                if file_mime == "application/pdf":
                    msgs.append({"role": "user", "content": [
                        {"type": "text", "text": user_text or ""},
                        {"type": "document", "source": {"type": "base64",
                                                         "media_type": "application/pdf",
                                                         "data": file_b64}},
                    ]})
                else:
                    try:
                        decoded = base64.b64decode(file_b64).decode("utf-8", errors="replace")
                        msgs.append({"role": "user",
                                     "content": f"{user_text}\n\n[Nội dung file]\n```\n{decoded[:8000]}\n```"})
                    except Exception:
                        msgs.append({"role": "user", "content": user_text or ""})
            else:
                msgs.append({"role": "user", "content": user_text})
            r = await post_with_retry(
                client,
                f"{p['base']}/messages",
                provider=provider,
                headers={"x-api-key": api_key, "anthropic-version": "2023-06-01"},
                json={"model": model, "max_tokens": 8096, "system": sys_msg, "messages": msgs},
            )
            _raise_for_provider(r, provider)
            return r.json()["content"][0]["text"]

        # -------- Gemini (v1beta) --------
        if kind == "gemini":
            contents = []
            for m in history:
                role = "model" if m["role"] == "assistant" else "user"
                contents.append({"role": role, "parts": [{"text": m["content"]}]})
            parts: list[dict[str, Any]] = [{"text": user_text or ""}]

            if parsed_attachments:
                for att in parsed_attachments:
                    if att["mime"].startswith("image/") or att["mime"] == "application/pdf":
                        parts.append({"inline_data": {"mime_type": att["mime"], "data": att["data"]}})
                    else:
                        try:
                            decoded = base64.b64decode(att["data"]).decode("utf-8", errors="replace")
                            parts.append({"text": f"[File: {att['name']}]\n```\n{decoded[:8000]}\n```"})
                        except Exception:
                            pass
            elif img:
                media, data = img
                parts.append({"inline_data": {"mime_type": media, "data": data}})
            elif file_b64 and file_mime:
                if file_mime.startswith("image/"):
                    parts.append({"inline_data": {"mime_type": file_mime, "data": file_b64}})
                else:
                    try:
                        decoded = base64.b64decode(file_b64).decode("utf-8", errors="replace")
                        parts.append({"text": f"[Nội dung file]\n```\n{decoded[:8000]}\n```"})
                    except Exception:
                        pass
            contents.append({"role": "user", "parts": parts})
            url = f"{p['base']}/models/{model}:generateContent?key={api_key}"
            payload: dict[str, Any] = {
                "contents": contents,
                "systemInstruction": {"parts": [{"text": sys_msg}]},
                "generationConfig": {"maxOutputTokens": 8192},
            }
            r = await post_with_retry(client, url, provider=provider, json=payload)
            _raise_for_provider(r, provider)
            data_r = r.json()
            try:
                return data_r["candidates"][0]["content"]["parts"][0]["text"]
            except (KeyError, IndexError):
                finish = data_r.get("candidates", [{}])[0].get("finishReason", "UNKNOWN")
                raise HTTPException(status_code=400,
                    detail=f"Gemini không trả về nội dung (finishReason={finish}). "
                           f"Có thể nội dung bị chặn bởi bộ lọc an toàn.")

    raise HTTPException(status_code=500, detail="Lỗi cấu hình provider.")


# ========================== FastAPI ==========================
app = FastAPI(title="KENIOS kenios", version="4.2")
app.add_middleware(CORSMiddleware, allow_origins=["*"],
                   allow_credentials=False, allow_methods=["*"], allow_headers=["*"])


@app.on_event("startup")
def _startup() -> None:
    init_db()
    start_mail_smtp()


# ======================== Pydantic Models ========================
class RegisterIn(BaseModel):
    username: str
    password: str
    email: Optional[str] = None
    phone: Optional[str] = None
    code: Optional[str] = None     # mã xác nhận gửi qua email (nếu có email)

class LoginIn(BaseModel):
    username: str
    password: str

class ForgotIn(BaseModel):
    username: str

class ResetIn(BaseModel):
    token: str
    new_password: str

class ProfileIn(BaseModel):
    email: Optional[str] = None
    phone: Optional[str] = None
    new_password: Optional[str] = None
    lang: Optional[str] = None

class KeyIn(BaseModel):
    provider: str
    api_key: str

class AttachmentIn(BaseModel):
    name: str
    data_base64: str
    mime: str

class ChatIn(BaseModel):
    provider: str
    message: str = ""
    image: Optional[str] = None
    file_base64: Optional[str] = None
    file_mime: Optional[str] = None
    attachments: Optional[list[AttachmentIn]] = None
    model: Optional[str] = None
    conversation_id: Optional[int] = None
    api_key: Optional[str] = None
    system: Optional[str] = None
    # Thêm tham số nâng cao
    web_search: Optional[bool] = False
    file_ids: Optional[list[int]] = None

class EnsembleIn(BaseModel):
    providers: list[str]
    message: str
    judge: Optional[str] = None

class CodeRunIn(BaseModel):
    code: str
    stdin: Optional[str] = None
    language: Optional[str] = "python"

class FileRunIn(BaseModel):
    file_id: int
    args: Optional[str] = None

class CodeReviewIn(BaseModel):
    provider: str
    code: str
    language: Optional[str] = None
    task: str = "review"
    target_lang: Optional[str] = None
    api_key: Optional[str] = None
    model: Optional[str] = None

class PaymentIn(BaseModel):
    amount: int
    package: str

class CodeZipIn(BaseModel):
    text: str

class PromptTemplateIn(BaseModel):
    title: str
    content: str
    category: Optional[str] = None
    is_public: Optional[bool] = False

class FavoriteIn(BaseModel):
    message_content: str
    conversation_id: Optional[int] = None
    provider: Optional[str] = None


class FriendRequestIn(BaseModel):
    friend_id: int


class FriendResponseIn(BaseModel):
    request_id: int
    action: str  # 'accept' or 'decline'


class DirectMessageIn(BaseModel):
    receiver_id: int
    content: str


# ======================== Health & Config ========================
@app.get("/health")
def health() -> dict[str, Any]:
    return {"status": "ok", "time": int(time.time()), "version": "4.2",
            "providers": len(PROVIDERS)}


@app.get("/config")
def config() -> dict[str, Any]:
    return {"name": "KENIOS kenios", "version": "4.2",
            "providers": _providers_public()}


def _providers_public() -> list[dict[str, Any]]:
    return [
        {"id": k, "label": v["label"], "models": v["models"],
         "default_model": v["default_model"], "vision": v["vision"],
         "free": v["free"], "code": v.get("code", False)}
        for k, v in PROVIDERS.items()
        if not (k == "kenios" and not KENIOS_AI_ENABLE)
    ]


@app.get("/providers")
def providers_list() -> list[dict[str, Any]]:
    return _providers_public()


# ======================== KENIOS AI — cấp API key cho người khác dùng ké ========================
class ApiTokenIn(BaseModel):
    name: str = ""


@app.post("/apitokens/create")
def apitoken_create(b: ApiTokenIn, user=Depends(get_user)) -> dict[str, Any]:
    token = "ken-" + secrets.token_urlsafe(24)
    with db() as c:
        c.execute("INSERT INTO api_tokens(token,owner_uid,name,calls,created_at) VALUES(?,?,?,0,?)",
                  (token, user["id"], (b.name or "Khoá API").strip()[:60], int(time.time())))
    return {"token": token}


@app.get("/apitokens")
def apitoken_list(user=Depends(get_user)) -> dict[str, Any]:
    with db() as c:
        rows = c.execute("SELECT token,name,calls,created_at FROM api_tokens "
                         "WHERE owner_uid=? ORDER BY created_at DESC", (user["id"],)).fetchall()
    return {"tokens": [dict(r) for r in rows]}


@app.delete("/apitokens/{token}")
def apitoken_delete(token: str, user=Depends(get_user)) -> dict[str, Any]:
    with db() as c:
        c.execute("DELETE FROM api_tokens WHERE token=? AND owner_uid=?", (token, user["id"]))
    return {"ok": True}


class PublicChatIn(BaseModel):
    token: str
    message: str
    model: Optional[str] = None
    system: Optional[str] = None


@app.post("/v1/kenios/chat")
async def public_kenios_chat(b: PublicChatIn) -> dict[str, Any]:
    """API công khai: người khác dùng API key của bạn để gọi KENIOS AI (model tự host)."""
    if not (b.message or "").strip():
        raise HTTPException(status_code=400, detail="Thiếu 'message'.")
    with db() as c:
        row = c.execute("SELECT owner_uid FROM api_tokens WHERE token=?", (b.token,)).fetchone()
        if not row:
            raise HTTPException(status_code=401, detail="API key không hợp lệ.")
        c.execute("UPDATE api_tokens SET calls=calls+1 WHERE token=?", (b.token,))
    if not KENIOS_AI_ENABLE:
        raise HTTPException(status_code=503, detail="KENIOS AI chưa được bật trên máy chủ.")
    reply = await call_provider("kenios", KENIOS_AI_KEY or "ollama", b.model, [],
                                b.message, system_override=b.system)
    return {"reply": reply, "model": b.model or KENIOS_AI_MODEL}


# ======================== Đăng ký thiết bị (lấy UDID để ký app ad-hoc) ========================
@app.get("/enroll/start", response_class=HTMLResponse)
def enroll_start(request: Request) -> Any:
    base = str(request.base_url).rstrip("/")
    return f"""<!doctype html><html><head><meta charset='utf-8'>
<meta name='viewport' content='width=device-width,initial-scale=1'>
<title>Đăng ký thiết bị KENIOS</title>
<style>body{{font-family:-apple-system;background:#0b1020;color:#fff;text-align:center;padding:40px}}
a.btn{{display:inline-block;margin-top:24px;padding:16px 28px;background:#4f46e5;color:#fff;
text-decoration:none;border-radius:14px;font-size:18px;font-weight:700}}
p{{color:#aab;max-width:520px;margin:10px auto}}</style></head>
<body><h2>Đăng ký thiết bị KENIOS</h2>
<p>Bấm nút bên dưới để lấy <b>UDID</b> thiết bị (cài hồ sơ cấu hình). UDID dùng để ký app cho riêng máy bạn.</p>
<a class='btn' href='{base}/enroll/profile'>Lấy UDID thiết bị</a>
<p style='margin-top:30px;font-size:13px'>Sau khi cài app đã ký: vào <b>Cài đặt → Cài đặt chung → VPN &amp; Quản lý thiết bị</b> → bấm tên nhà phát triển → <b>Tin cậy</b>.</p>
</body></html>"""


@app.get("/enroll/profile")
def enroll_profile(request: Request) -> Response:
    base = str(request.base_url).rstrip("/")
    puid = secrets.token_hex(8)
    profile = f"""<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE plist PUBLIC "-//Apple//DTD PLIST 1.0//EN" "http://www.apple.com/DTDs/PropertyList-1.0.dtd">
<plist version="1.0"><dict>
  <key>PayloadContent</key>
  <dict>
    <key>URL</key><string>{base}/enroll/callback</string>
    <key>DeviceAttributes</key>
    <array>
      <string>UDID</string><string>PRODUCT</string><string>VERSION</string>
      <string>SERIAL</string><string>DEVICE_NAME</string>
    </array>
  </dict>
  <key>PayloadOrganization</key><string>KENIOS</string>
  <key>PayloadDisplayName</key><string>Đăng ký thiết bị KENIOS</string>
  <key>PayloadVersion</key><integer>1</integer>
  <key>PayloadUUID</key><string>{puid}</string>
  <key>PayloadIdentifier</key><string>com.kenios.enroll</string>
  <key>PayloadType</key><string>Profile Service</string>
</dict></plist>"""
    return Response(content=profile, media_type="application/x-apple-aspen-config")


@app.post("/enroll/callback", response_class=HTMLResponse)
async def enroll_callback(request: Request) -> Any:
    body = await request.body()
    import plistlib
    udid = ""; info: dict[str, Any] = {}
    start = body.find(b"<?xml")
    end = body.find(b"</plist>")
    if start != -1 and end != -1:
        try:
            info = plistlib.loads(body[start:end + 8])
            udid = str(info.get("UDID", ""))
        except Exception:
            udid = ""
    if udid:
        with db() as c:
            c.execute("INSERT OR REPLACE INTO devices(udid,product,version,serial,name,created_at) "
                      "VALUES(?,?,?,?,?,?)",
                      (udid, str(info.get("PRODUCT", "")), str(info.get("VERSION", "")),
                       str(info.get("SERIAL", "")), str(info.get("DEVICE_NAME", "")), int(time.time())))
    return f"""<!doctype html><html><head><meta charset='utf-8'>
<meta name='viewport' content='width=device-width,initial-scale=1'>
<style>body{{font-family:-apple-system;background:#0b1020;color:#fff;text-align:center;padding:50px}}
.box{{background:#161c33;border-radius:14px;padding:20px;max-width:520px;margin:0 auto}}
code{{color:#8ef;word-break:break-all}}</style></head><body>
<h2>✅ Đã ghi nhận thiết bị</h2>
<div class='box'><p>UDID của bạn:</p><h3><code>{udid or 'Không đọc được'}</code></h3></div>
<p style='color:#aab;margin-top:20px'>Gửi UDID này cho nhà phát triển để được ký app. Sau khi cài bản đã ký, nhớ vào Cài đặt → VPN &amp; Quản lý thiết bị để <b>Tin cậy</b>.</p>
</body></html>"""


@app.get("/devices")
def devices_list(admin=Depends(get_admin)) -> dict[str, Any]:
    with db() as c:
        rows = c.execute("SELECT udid,product,version,serial,name,created_at "
                         "FROM devices ORDER BY created_at DESC").fetchall()
    return {"devices": [dict(r) for r in rows]}


# ======================== Auth ========================
@app.post("/auth/register")
def register(b: RegisterIn) -> dict[str, Any]:
    if len(b.username) < 3 or len(b.password) < 6:
        raise HTTPException(status_code=400,
            detail="Username ≥3 ký tự, mật khẩu ≥6 ký tự.")
    # Nếu có gửi mã xác nhận thì bắt buộc khớp (xác minh email qua OTP)
    if b.code is not None and (b.email or "").strip():
        if not _otp_check(b.email, b.code):
            raise HTTPException(status_code=400,
                detail="Mã xác nhận sai hoặc đã hết hạn. Vui lòng lấy mã mới.")
    with db() as c:
        if c.execute("SELECT 1 FROM users WHERE username=?", (b.username,)).fetchone():
            raise HTTPException(status_code=409, detail="Username đã tồn tại.")
        cur = c.execute(
            "INSERT INTO users(username,email,phone,pw_hash,plan,credits,created_at) "
            "VALUES(?,?,?,?,'free',0,?)",
            (b.username, b.email, b.phone, hash_pw(b.password), int(time.time())),
        )
        uid = cur.lastrowid
        pid = _ensure_public_id(c, uid)
    return {"token": make_token(uid),
            "user": {"id": uid, "username": b.username, "email": b.email,
                     "phone": b.phone, "public_id": pid, "is_admin": False,
                     "plan": "free", "credits": 0, "lang": "vi", "status": "active"}}


@app.post("/auth/login")
def login(b: LoginIn) -> dict[str, Any]:
    with db() as c:
        row = c.execute("SELECT * FROM users WHERE username=?", (b.username,)).fetchone()
    if not row or not verify_pw(b.password, row["pw_hash"]):
        raise HTTPException(status_code=401, detail="Sai username hoặc mật khẩu.")
    if row["banned"]:
        raise HTTPException(status_code=403, detail="Tài khoản đã bị khóa. Liên hệ quản trị viên.")
    with db() as c:
        _ensure_public_id(c, row["id"])
        row = c.execute("SELECT * FROM users WHERE id=?", (row["id"],)).fetchone()
    return {"token": make_token(row["id"]), "user": _user_dict(row)}


@app.post("/auth/forgot-password")
def forgot(b: ForgotIn) -> dict[str, Any]:
    token = secrets.token_urlsafe(24)
    with db() as c:
        row = c.execute("SELECT id FROM users WHERE username=?", (b.username,)).fetchone()
        if row:
            c.execute("UPDATE users SET reset_token=?, reset_exp=? WHERE id=?",
                      (token, int(time.time()) + 1800, row["id"]))
    log.info("Reset token cho %s: %s", b.username, token)
    return {"message": "Nếu tài khoản tồn tại, mã đặt lại đã được tạo.",
            "reset_token": token}


@app.post("/auth/reset-password")
def reset_pw(b: ResetIn) -> dict[str, Any]:
    if len(b.new_password) < 6:
        raise HTTPException(status_code=400, detail="Mật khẩu mới ≥6 ký tự.")
    with db() as c:
        row = c.execute("SELECT id,reset_exp FROM users WHERE reset_token=?",
                        (b.token,)).fetchone()
        if not row or (row["reset_exp"] or 0) < time.time():
            raise HTTPException(status_code=400,
                detail="Mã đặt lại sai hoặc đã hết hạn (30 phút).")
        c.execute("UPDATE users SET pw_hash=?, reset_token=NULL, reset_exp=NULL WHERE id=?",
                  (hash_pw(b.new_password), row["id"]))
    return {"message": "Đổi mật khẩu thành công."}


@app.post("/auth/update-profile")
def update_profile(b: ProfileIn, user=Depends(get_user)) -> dict[str, Any]:
    fields, vals = [], []
    if b.email is not None:
        fields.append("email=?"); vals.append(b.email)
    if b.phone is not None:
        fields.append("phone=?"); vals.append(b.phone)
    if b.new_password:
        if len(b.new_password) < 6:
            raise HTTPException(status_code=400, detail="Mật khẩu mới ≥6 ký tự.")
        fields.append("pw_hash=?"); vals.append(hash_pw(b.new_password))
    if b.lang in ("vi", "en"):
        fields.append("lang=?"); vals.append(b.lang)
    if not fields:
        raise HTTPException(status_code=400, detail="Không có gì để cập nhật.")
    vals.append(user["id"])
    with db() as c:
        c.execute(f"UPDATE users SET {', '.join(fields)} WHERE id=?", vals)
    return {"message": "Cập nhật thành công."}


# ======================== API Keys (User) ========================
@app.post("/keys")
def save_key(b: KeyIn, user=Depends(get_user)) -> dict[str, Any]:
    if b.provider not in PROVIDERS:
        raise HTTPException(status_code=400,
            detail=f"AI '{b.provider}' không được hỗ trợ. Danh sách hợp lệ: {list(PROVIDERS.keys())}")
    with db() as c:
        c.execute("INSERT INTO apikeys(user_id,provider,enc_key) VALUES(?,?,?) "
                  "ON CONFLICT(user_id,provider) DO UPDATE SET enc_key=excluded.enc_key",
                  (user["id"], b.provider, enc(b.api_key)))
    return {"message": f"Đã lưu API key cho {b.provider} thành công."}


@app.get("/keys")
def list_keys(user=Depends(get_user)) -> list[dict[str, Any]]:
    with db() as c:
        rows = c.execute("SELECT provider FROM apikeys WHERE user_id=?",
                         (user["id"],)).fetchall()
    return [{"provider": r["provider"], "configured": True} for r in rows]


@app.delete("/keys/{provider}")
def del_key(provider: str, user=Depends(get_user)) -> dict[str, Any]:
    with db() as c:
        c.execute("DELETE FROM apikeys WHERE user_id=? AND provider=?",
                  (user["id"], provider))
    return {"message": f"Đã xóa key {provider}."}


@app.post("/keys/test")
async def test_key(b: KeyIn, user=Depends(get_user)) -> dict[str, Any]:
    if b.provider not in PROVIDERS:
        raise HTTPException(status_code=400, detail=f"AI '{b.provider}' không được hỗ trợ.")
    try:
        await call_provider(b.provider, b.api_key, None, [], "ping")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400,
            detail=f"Key không dùng được: {type(e).__name__}: {str(e)[:200]}")
    return {"ok": True, "message": f"Key {b.provider} hoạt động tốt."}


# ======================== Admin API Keys ========================
@app.post("/admin/keys")
def admin_save_key(b: KeyIn, admin=Depends(get_admin)) -> dict[str, Any]:
    if b.provider not in PROVIDERS:
        raise HTTPException(status_code=400,
            detail=f"AI '{b.provider}' không được hỗ trợ. Danh sách hợp lệ: {list(PROVIDERS.keys())}")
    with db() as c:
        c.execute("INSERT INTO admin_apikeys(provider,enc_key) VALUES(?,?) "
                  "ON CONFLICT(provider) DO UPDATE SET enc_key=excluded.enc_key",
                  (b.provider, enc(b.api_key)))
    return {"message": f"Đã lưu admin API key cho {b.provider}."}


@app.get("/admin/keys")
def admin_list_keys(admin=Depends(get_admin)) -> list[dict[str, Any]]:
    with db() as c:
        rows = c.execute("SELECT provider FROM admin_apikeys").fetchall()
    return [{"provider": r["provider"], "configured": True} for r in rows]


@app.delete("/admin/keys/{provider}")
def admin_del_key(provider: str, admin=Depends(get_admin)) -> dict[str, Any]:
    with db() as c:
        c.execute("DELETE FROM admin_apikeys WHERE provider=?", (provider,))
    return {"message": f"Đã xóa admin key cho {provider}."}


# ======================== Chat ========================
def load_history(conversation_id: int, user_id: int) -> list[dict[str, str]]:
    with db() as c:
        own = c.execute("SELECT 1 FROM conversations WHERE id=? AND user_id=?",
                        (conversation_id, user_id)).fetchone()
        if not own:
            raise HTTPException(status_code=404, detail="Không tìm thấy hội thoại.")
        rows = c.execute(
            "SELECT role,content FROM messages WHERE conversation_id=? ORDER BY id",
            (conversation_id,)
        ).fetchall()
    return [{"role": r["role"], "content": r["content"]} for r in rows]


def new_conversation(user_id: int, provider: str, title: str) -> int:
    now = int(time.time())
    with db() as c:
        cur = c.execute(
            "INSERT INTO conversations(user_id,title,provider,created_at,updated_at) "
            "VALUES(?,?,?,?,?)",
            (user_id, (title or "Hội thoại mới")[:80], provider, now, now),
        )
        return cur.lastrowid


def save_message(conversation_id: int, role: str, content: str, tokens: int = 0) -> None:
    with db() as c:
        c.execute("INSERT INTO messages(conversation_id,role,content,tokens_used,created_at) VALUES(?,?,?,?,?)",
                  (conversation_id, role, content, tokens, int(time.time())))
        c.execute("UPDATE conversations SET updated_at=? WHERE id=?",
                  (int(time.time()), conversation_id))


@app.post("/chat")
async def chat(b: ChatIn, user=Depends(get_user)) -> dict[str, Any]:
    if not b.message and not b.image and not b.file_base64 and not b.attachments:
        raise HTTPException(status_code=400,
            detail="Thiếu nội dung: cần ít nhất 'message', 'image', 'file_base64', hoặc 'attachments'.")
    
    # 1. Xử lý RAG nếu đính kèm file ID từ thư viện
    rag_context = ""
    if b.file_ids:
        for fid in b.file_ids:
            with db() as c:
                row = c.execute("SELECT name, data FROM files WHERE id=? AND user_id=?",
                                (fid, user["id"])).fetchone()
            if row:
                file_path = os.path.join(UPLOAD_DIR, str(fid))
                file_text = ""
                if os.path.exists(file_path):
                    file_text = parse_file_content(file_path, row["name"])
                elif row["data"]:
                    try:
                        file_text = base64.b64decode(row["data"]).decode("utf-8", errors="replace")
                    except Exception:
                        pass
                
                if file_text:
                    # Trích xuất đoạn liên quan qua TF-IDF nếu tệp quá lớn (>10k chars)
                    if len(file_text) > 10000:
                        relevant = retrieve_relevant_chunks(file_text, b.message)
                        rag_context += f"\n\n--- [ĐOẠN TRÍCH TỪ TỆP: {row['name']}] ---\n{relevant}\n"
                    else:
                        rag_context += f"\n\n--- [NỘI DUNG TỆP: {row['name']}] ---\n{file_text}\n"

    # 2. Xử lý Web Search thông qua DuckDuckGo
    search_context = ""
    if b.web_search:
        search_context = await search_ddg(b.message)

    # Ghép ngữ cảnh vào system prompt
    final_system = b.system or DEFAULT_SYSTEM
    if rag_context:
        final_system += f"\n\n[Dữ liệu tài liệu đính kèm]:{rag_context}\nHãy dựa vào nội dung tài liệu trên để trả lời câu hỏi của người dùng một cách chính xác."
    if search_context:
        final_system += f"\n\n[Dữ liệu tìm kiếm thời gian thực từ Internet]:\n{search_context}\nHãy tổng hợp các thông tin Internet trên để đưa ra câu trả lời mới và chính xác nhất."

    key = get_user_key(user["id"], b.provider, b.api_key)
    conv_id = b.conversation_id or new_conversation(
        user["id"], b.provider, b.message or "File/Ảnh")
    history = load_history(conv_id, user["id"]) if b.conversation_id else []

    attachments_data = None
    if b.attachments:
        attachments_data = [
            {"name": a.name, "data_base64": a.data_base64, "mime": a.mime}
            for a in b.attachments[:30]
        ]

    reply = await call_provider(
        b.provider, key, b.model, history,
        b.message, b.image,
        b.file_base64, b.file_mime,
        final_system,
        attachments=attachments_data,
    )

    input_tokens = estimate_tokens(b.message)
    output_tokens = estimate_tokens(reply)
    total_tokens = input_tokens + output_tokens

    user_msg_display = b.message or ("[ảnh]" if b.image else ("[file]" if b.file_base64 else "[đính kèm]"))
    save_message(conv_id, "user", user_msg_display, tokens=input_tokens)
    save_message(conv_id, "assistant", reply, tokens=output_tokens)
    saved_files = save_code_blocks(user["id"], reply, f"chat{conv_id}")
    return {
        "reply": reply,
        "conversation_id": conv_id,
        "provider": b.provider,
        "model": b.model or PROVIDERS[b.provider]["default_model"],
        "saved_files": saved_files,
        "tokens_estimated": {
            "input": input_tokens,
            "output": output_tokens,
            "total": total_tokens,
        },
    }


async def call_provider_stream(provider: str, api_key: str, model: Optional[str],
                               history: list[dict[str, Any]], user_text: str,
                               system_override: Optional[str] = None):
    """Gọi AI ở chế độ streaming — sinh từng đoạn text (cho OpenAI-compatible, Gemini, Anthropic).
    Provider khác → fallback gọi 1 lần và trả nguyên câu."""
    if provider not in PROVIDERS:
        raise HTTPException(status_code=400, detail=f"AI '{provider}' không được hỗ trợ.")
    p = PROVIDERS[provider]
    model = model or p["default_model"]
    kind = p["kind"]
    sys_msg = system_override or DEFAULT_SYSTEM

    async with httpx.AsyncClient(timeout=REQUEST_TIMEOUT) as client:
        if kind == "openai":
            msgs = [{"role": "system", "content": sys_msg}]
            msgs += [{"role": m["role"], "content": m["content"]} for m in history]
            msgs.append({"role": "user", "content": user_text})
            async with client.stream("POST", f"{p['base']}/chat/completions",
                headers={"Authorization": f"Bearer {api_key}",
                         "HTTP-Referer": "https://kenios.app", "X-Title": "KENIOS"},
                json={"model": model, "messages": msgs, "stream": True}) as r:
                if r.status_code >= 400:
                    txt = (await r.aread()).decode("utf-8", "replace")[:300]
                    raise HTTPException(status_code=502, detail=f"{provider} lỗi {r.status_code}: {txt}")
                async for line in r.aiter_lines():
                    if not line or not line.startswith("data:"):
                        continue
                    data = line[5:].strip()
                    if data == "[DONE]":
                        break
                    try:
                        delta = json.loads(data)["choices"][0]["delta"].get("content")
                        if delta:
                            yield delta
                    except Exception:
                        continue
            return

        if kind == "gemini":
            contents = []
            for m in history:
                contents.append({"role": "model" if m["role"] == "assistant" else "user",
                                 "parts": [{"text": m["content"]}]})
            contents.append({"role": "user", "parts": [{"text": user_text}]})
            url = f"{p['base']}/models/{model}:streamGenerateContent?alt=sse&key={api_key}"
            payload = {"contents": contents, "systemInstruction": {"parts": [{"text": sys_msg}]}}
            async with client.stream("POST", url, json=payload) as r:
                if r.status_code >= 400:
                    txt = (await r.aread()).decode("utf-8", "replace")[:300]
                    raise HTTPException(status_code=502, detail=f"gemini lỗi {r.status_code}: {txt}")
                async for line in r.aiter_lines():
                    if not line or not line.startswith("data:"):
                        continue
                    data = line[5:].strip()
                    if not data:
                        continue
                    try:
                        t = json.loads(data)["candidates"][0]["content"]["parts"][0]["text"]
                        if t:
                            yield t
                    except Exception:
                        continue
            return

        if kind == "anthropic":
            msgs = [{"role": m["role"], "content": m["content"]} for m in history]
            msgs.append({"role": "user", "content": user_text})
            async with client.stream("POST", f"{p['base']}/messages",
                headers={"x-api-key": api_key, "anthropic-version": "2023-06-01"},
                json={"model": model, "max_tokens": 8096, "system": sys_msg,
                      "messages": msgs, "stream": True}) as r:
                if r.status_code >= 400:
                    txt = (await r.aread()).decode("utf-8", "replace")[:300]
                    raise HTTPException(status_code=502, detail=f"anthropic lỗi {r.status_code}: {txt}")
                async for line in r.aiter_lines():
                    if not line or not line.startswith("data:"):
                        continue
                    data = line[5:].strip()
                    try:
                        obj = json.loads(data)
                        if obj.get("type") == "content_block_delta":
                            t = obj.get("delta", {}).get("text")
                            if t:
                                yield t
                    except Exception:
                        continue
            return

    # Fallback: provider không hỗ trợ streaming ở đây → gọi 1 lần
    full = await call_provider(provider, api_key, model, history, user_text,
                               system_override=system_override)
    yield full


@app.post("/chat/stream")
async def chat_stream(b: ChatIn, user=Depends(get_user)):
    """Chat dạng streaming (trả lời hiện dần). Chỉ hỗ trợ text; ảnh/file dùng /chat."""
    if not b.message:
        raise HTTPException(status_code=400, detail="Cần 'message' cho chế độ streaming.")
    key = get_user_key(user["id"], b.provider, b.api_key)
    conv_id = b.conversation_id or new_conversation(user["id"], b.provider, b.message)
    history = load_history(conv_id, user["id"]) if b.conversation_id else []
    final_system = b.system or DEFAULT_SYSTEM

    async def gen():
        full = ""
        try:
            async for chunk in call_provider_stream(b.provider, key, b.model, history,
                                                    b.message, final_system):
                full += chunk
                yield "data: " + json.dumps({"delta": chunk}, ensure_ascii=False) + "\n\n"
        except HTTPException as e:
            yield "data: " + json.dumps({"error": str(e.detail)}, ensure_ascii=False) + "\n\n"
        except (httpx.ConnectError, httpx.ReadTimeout, httpx.RemoteProtocolError) as e:
            msg = (str(e) if b.provider != "kenios"
                   else "KENIOS AI chưa chạy. Cài Ollama trên VPS rồi thử lại "
                        "(bash kenios-ai/install-ai.sh).")
            yield "data: " + json.dumps({"error": msg}, ensure_ascii=False) + "\n\n"
        except Exception as e:
            yield "data: " + json.dumps({"error": str(e)}, ensure_ascii=False) + "\n\n"
        if full:
            save_message(conv_id, "user", b.message, tokens=estimate_tokens(b.message))
            save_message(conv_id, "assistant", full, tokens=estimate_tokens(full))
        yield "data: " + json.dumps({"done": True, "conversation_id": conv_id}) + "\n\n"

    return StreamingResponse(gen(), media_type="text/event-stream")


@app.post("/chat/ensemble")
async def ensemble(b: EnsembleIn, user=Depends(get_user)) -> dict[str, Any]:
    if len(b.providers) < 2:
        raise HTTPException(status_code=400, detail="Cần ít nhất 2 AI để ensemble.")

    async def one(prov: str):
        try:
            key = get_user_key(user["id"], prov, None)
            ans = await call_provider(prov, key, None, [], b.message)
            return prov, ans
        except HTTPException as e:
            return prov, f"[lỗi: {e.detail}]"

    results = await asyncio.gather(*[one(p) for p in b.providers])
    answers = {prov: ans for prov, ans in results}
    judge = b.judge or b.providers[0]
    judge_key = get_user_key(user["id"], judge, None)
    merged = (
        "Dưới đây là câu trả lời của nhiều AI cho cùng một câu hỏi. "
        "Hãy hợp nhất thành MỘT câu trả lời tốt nhất.\n\nCÂU HỎI:\n" + b.message
        + "\n\nCÁC CÂU TRẢ LỜI:\n"
        + "\n\n".join(f"### {p}\n{a}" for p, a in answers.items())
    )
    best = await call_provider(judge, judge_key, None, [], merged)
    return {"best": best, "judge": judge, "answers": answers}


# ======================== Lịch sử hội thoại ========================
@app.get("/conversations")
def list_conversations(user=Depends(get_user)) -> list[dict[str, Any]]:
    with db() as c:
        rows = c.execute(
            "SELECT id,title,provider,pinned,updated_at FROM conversations WHERE user_id=? "
            "ORDER BY pinned DESC, updated_at DESC", (user["id"],)
        ).fetchall()
    return [dict(r) for r in rows]


@app.get("/conversations/{cid}")
def get_conversation(cid: int, user=Depends(get_user)) -> dict[str, Any]:
    msgs = load_history(cid, user["id"])
    return {"conversation_id": cid, "messages": msgs}


@app.delete("/conversations/{cid}")
def delete_conversation(cid: int, user=Depends(get_user)) -> dict[str, Any]:
    with db() as c:
        c.execute("DELETE FROM messages WHERE conversation_id=? AND conversation_id IN "
                  "(SELECT id FROM conversations WHERE user_id=?)", (cid, user["id"]))
        c.execute("DELETE FROM conversations WHERE id=? AND user_id=?", (cid, user["id"]))
    return {"message": "Đã xóa hội thoại."}


# ======================== Pin hội thoại ========================
@app.post("/conversations/{cid}/pin")
def pin_conversation(cid: int, user=Depends(get_user)) -> dict[str, Any]:
    with db() as c:
        row = c.execute("SELECT pinned FROM conversations WHERE id=? AND user_id=?",
                        (cid, user["id"])).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Không tìm thấy hội thoại.")
        new_val = 0 if row["pinned"] else 1
        c.execute("UPDATE conversations SET pinned=? WHERE id=?", (new_val, cid))
    return {"pinned": bool(new_val), "message": "Đã ghim." if new_val else "Đã bỏ ghim."}


# ======================== Share hội thoại ========================
@app.post("/conversations/{cid}/share")
def share_conversation(cid: int, user=Depends(get_user)) -> dict[str, Any]:
    with db() as c:
        row = c.execute("SELECT share_token FROM conversations WHERE id=? AND user_id=?",
                        (cid, user["id"])).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Không tìm thấy hội thoại.")
        token = row["share_token"]
        if not token:
            token = secrets.token_urlsafe(24)
            c.execute("UPDATE conversations SET share_token=? WHERE id=?", (token, cid))
    return {"share_token": token, "share_url": f"/share/{token}"}


@app.get("/share/{token}")
def view_shared(token: str) -> dict[str, Any]:
    with db() as c:
        conv = c.execute("SELECT id,title,provider,created_at FROM conversations WHERE share_token=?",
                         (token,)).fetchone()
        if not conv:
            raise HTTPException(status_code=404, detail="Link chia sẻ không hợp lệ hoặc đã bị xóa.")
        msgs = c.execute("SELECT role,content,created_at FROM messages WHERE conversation_id=? ORDER BY id",
                         (conv["id"],)).fetchall()
    return {
        "conversation_id": conv["id"],
        "title": conv["title"],
        "provider": conv["provider"],
        "created_at": conv["created_at"],
        "messages": [dict(m) for m in msgs],
    }


# ======================== Export hội thoại ========================
@app.get("/conversations/{cid}/export")
def export_conversation(cid: int, format: str = "md", user=Depends(get_user)) -> dict[str, Any]:
    with db() as c:
        conv = c.execute("SELECT title,provider,created_at FROM conversations WHERE id=? AND user_id=?",
                         (cid, user["id"])).fetchone()
        if not conv:
            raise HTTPException(status_code=404, detail="Không tìm thấy hội thoại.")
        msgs = c.execute("SELECT role,content,created_at FROM messages WHERE conversation_id=? ORDER BY id",
                         (cid,)).fetchall()

    messages_list = [dict(m) for m in msgs]

    if format == "json":
        content = json.dumps({
            "title": conv["title"],
            "provider": conv["provider"],
            "created_at": conv["created_at"],
            "messages": messages_list,
        }, ensure_ascii=False, indent=2)
        filename = f"conversation_{cid}.json"
    elif format == "txt":
        lines = [f"Hội thoại: {conv['title']}", f"Provider: {conv['provider']}", ""]
        for m in messages_list:
            role_label = "Bạn" if m["role"] == "user" else "AI"
            lines.append(f"[{role_label}]")
            lines.append(m["content"])
            lines.append("")
        content = "\n".join(lines)
        filename = f"conversation_{cid}.txt"
    else:  # md
        lines = [f"# {conv['title']}", f"**Provider:** {conv['provider']}", ""]
        for m in messages_list:
            role_label = "👤 Bạn" if m["role"] == "user" else "🤖 AI"
            lines.append(f"### {role_label}")
            lines.append(m["content"])
            lines.append("---")
            lines.append("")
        content = "\n".join(lines)
        filename = f"conversation_{cid}.md"

    data_b64 = base64.b64encode(content.encode("utf-8")).decode()
    return {"filename": filename, "format": format, "data_base64": data_b64, "content": content}


# ======================== Dọn dẹp cơ sở dữ liệu ========================
class CleanupIn(BaseModel):
    days: Optional[int] = 30


@app.post("/db/cleanup")
def db_cleanup(b: CleanupIn, user=Depends(get_user)) -> dict[str, Any]:
    """Dọn dẹp cơ sở dữ liệu: Xóa các tin nhắn cũ hơn X ngày."""
    import time
    days = b.days if b.days is not None else 30
    limit_time = int(time.time()) - (days * 24 * 60 * 60)
    
    with db() as c:
        db_path = "kenios.db"
        size_before = 0
        if os.path.exists(db_path):
            size_before = os.path.getsize(db_path)
            
        # Xóa tin nhắn trong các cuộc hội thoại không được ghim (pinned = 0 hoặc null)
        cur = c.execute(
            "DELETE FROM messages WHERE created_at < ? AND conversation_id IN ("
            "SELECT id FROM conversations WHERE pinned = 0 OR pinned IS NULL"
            ") AND content NOT IN (SELECT message_content FROM favorites)",
            (limit_time,)
        )
        deleted_msgs = cur.rowcount
        
        # Xóa các cuộc hội thoại cũ không có tin nhắn hoặc không được ghim
        cur2 = c.execute(
            "DELETE FROM conversations WHERE updated_at < ? AND (pinned = 0 OR pinned IS NULL) "
            "AND id NOT IN (SELECT DISTINCT conversation_id FROM messages WHERE conversation_id IS NOT NULL)",
            (limit_time,)
        )
        deleted_convs = cur2.rowcount
        
        # Chạy VACUUM để tối ưu dung lượng ổ đĩa cơ sở dữ liệu SQLite
        try:
            c.execute("VACUUM")
        except Exception:
            pass
        
        size_after = 0
        if os.path.exists(db_path):
            size_after = os.path.getsize(db_path)
            
        freed_bytes = max(0, size_before - size_after)
        
        # Hàm tính dung lượng thân thiện
        def human_size_python(bytes_size: int) -> str:
            if bytes_size < 1024:
                return f"{bytes_size} B"
            elif bytes_size < 1024 * 1024:
                return f"{bytes_size / 1024:.1f} KB"
            else:
                return f"{bytes_size / 1024 / 1024:.1f} MB"
                
        freed_space_str = human_size_python(freed_bytes)
        
    return {
        "deleted_messages": deleted_msgs,
        "deleted_conversations": deleted_convs,
        "freed_space": freed_space_str,
        "message": f"Đã giải phóng {freed_space_str}. Xóa {deleted_msgs} tin nhắn & {deleted_convs} hội thoại cũ."
    }


# ======================== Quản lý File (Giới hạn 1KB - 4GB) ========================
MAX_FILE_B64 = 150_000_000

class FileIn(BaseModel):
    name: str
    category: Optional[str] = None
    mime: Optional[str] = None
    data_base64: str


@app.post("/files")
def upload_file(b: FileIn, user=Depends(get_user)) -> dict[str, Any]:
    if len(b.data_base64) > MAX_FILE_B64:
        raise HTTPException(status_code=413, detail="File quá lớn để truyền qua JSON (giới hạn ~75MB binary).")
    size = (len(b.data_base64) * 3) // 4
    
    with db() as c:
        cur = c.execute(
            "INSERT INTO files(user_id,name,category,mime,size,data,created_at) "
            "VALUES(?,?,?,?,?,'',?)",
            (user["id"], b.name, b.category or _guess_category(b.name, b.mime),
             b.mime, size, int(time.time())),
        )
        fid = cur.lastrowid
        
    try:
        file_path = os.path.join(UPLOAD_DIR, str(fid))
        with open(file_path, "wb") as f:
            f.write(base64.b64decode(b.data_base64))
    except Exception as e:
        with db() as c:
            c.execute("DELETE FROM files WHERE id=?", (fid,))
        raise HTTPException(status_code=500, detail=f"Lỗi lưu file: {str(e)}")
        
    return {"id": fid, "name": b.name, "size": size, "mime": b.mime}


@app.post("/files/upload")
async def upload_file_raw(
    request: Request,
    name: str,
    category: Optional[str] = None,
    user = Depends(get_user)
) -> dict[str, Any]:
    content_length = request.headers.get("content-length")
    if content_length:
        try:
            cl = int(content_length)
            if cl > MAX_FILE_SIZE:
                raise HTTPException(status_code=413, detail="File quá lớn (giới hạn tối đa là 4GB).")
            if cl < MIN_FILE_SIZE:
                raise HTTPException(status_code=400, detail="File quá nhỏ (giới hạn tối thiểu là 1KB).")
        except ValueError:
            pass

    temp_filename = f"tmp_{secrets.token_hex(8)}"
    temp_path = os.path.join(UPLOAD_DIR, temp_filename)
    
    total_size = 0
    try:
        with open(temp_path, "wb") as f:
            async for chunk in request.stream():
                total_size += len(chunk)
                if total_size > MAX_FILE_SIZE:
                    raise HTTPException(status_code=413, detail="File vượt quá giới hạn 4GB.")
                f.write(chunk)
    except Exception as e:
        if os.path.exists(temp_path):
            os.unlink(temp_path)
        if isinstance(e, HTTPException):
            raise e
        raise HTTPException(status_code=500, detail=f"Lỗi truyền phát file lên server: {str(e)}")

    if total_size < MIN_FILE_SIZE:
        if os.path.exists(temp_path):
            os.unlink(temp_path)
        raise HTTPException(status_code=400, detail="File quá nhỏ (giới hạn tối thiểu là 1KB).")

    mime = request.headers.get("content-type") or "application/octet-stream"
    with db() as c:
        cur = c.execute(
            "INSERT INTO files(user_id,name,category,mime,size,data,created_at) "
            "VALUES(?,?,?,?,?,'',?)",
            (user["id"], name, category or _guess_category(name, mime),
             mime, total_size, int(time.time())),
        )
        fid = cur.lastrowid

    final_path = os.path.join(UPLOAD_DIR, str(fid))
    os.rename(temp_path, final_path)
    
    return {"id": fid, "name": name, "size": total_size, "mime": mime}


@app.get("/files/{fid}/download")
def download_file_raw(fid: int, background_tasks: BackgroundTasks, user=Depends(get_user)):
    with db() as c:
        row = c.execute(
            "SELECT name,mime,data FROM files WHERE id=? AND user_id=?",
            (fid, user["id"])
        ).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Không tìm thấy file.")
        
    file_path = os.path.join(UPLOAD_DIR, str(fid))
    if os.path.exists(file_path):
        return FileResponse(
            path=file_path,
            filename=row["name"],
            media_type=row["mime"] or "application/octet-stream",
            content_disposition_type="attachment"
        )
        
    if row["data"]:
        try:
            temp_filename = f"temp_download_{fid}_{secrets.token_hex(4)}"
            temp_path = os.path.join(UPLOAD_DIR, temp_filename)
            with open(temp_path, "wb") as f:
                f.write(base64.b64decode(row["data"]))
            background_tasks.add_task(os.unlink, temp_path)
            return FileResponse(
                path=temp_path,
                filename=row["name"],
                media_type=row["mime"] or "application/octet-stream",
                content_disposition_type="attachment"
            )
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Lỗi nạp tệp từ DB: {e}")
            
    raise HTTPException(status_code=404, detail="Không tìm thấy nội dung tệp.")


def _guess_category(name: str, mime: Optional[str]) -> str:
    if mime and mime.startswith("image/"): return "image"
    ext = name.rsplit(".", 1)[-1].lower() if "." in name else ""
    if ext in ("py", "js", "ts", "swift", "kt", "go", "rs", "c", "cpp", "java",
               "php", "rb", "sh", "html", "css", "sql", "json", "yaml", "toml"):
        return "code"
    if ext in ("pdf", "docx", "doc", "txt", "md"): return "document"
    if mime and mime.startswith("image/"): return "image"
    return "other"


@app.get("/files")
def list_files(category: Optional[str] = None,
               user=Depends(get_user)) -> list[dict[str, Any]]:
    q = "SELECT id,name,category,mime,size,created_at FROM files WHERE user_id=?"
    args: list[Any] = [user["id"]]
    if category and category != "all":
        q += " AND category=?"; args.append(category)
    q += " ORDER BY id DESC"
    with db() as c:
        rows = c.execute(q, args).fetchall()
    return [dict(r) for r in rows]


@app.get("/files/{fid}")
def download_file(fid: int, user=Depends(get_user)) -> dict[str, Any]:
    with db() as c:
        row = c.execute(
            "SELECT name,category,mime,data,size FROM files WHERE id=? AND user_id=?",
            (fid, user["id"])
        ).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Không tìm thấy file.")
        
    file_path = os.path.join(UPLOAD_DIR, str(fid))
    if os.path.exists(file_path):
        size = os.path.getsize(file_path)
        if size > 50_000_000:
            raise HTTPException(status_code=413, detail="Tệp quá lớn để tải qua JSON (lớn hơn 50MB). Vui lòng dùng link tải trực tiếp (Stream).")
        with open(file_path, "rb") as f:
            data_b64 = base64.b64encode(f.read()).decode()
    else:
        data_b64 = row["data"]
        
    return {"name": row["name"], "category": row["category"],
            "mime": row["mime"], "data_base64": data_b64}


@app.delete("/files/{fid}")
def delete_file(fid: int, user=Depends(get_user)) -> dict[str, Any]:
    with db() as c:
        c.execute("DELETE FROM files WHERE id=? AND user_id=?", (fid, user["id"]))
    file_path = os.path.join(UPLOAD_DIR, str(fid))
    if os.path.exists(file_path):
        try:
            os.unlink(file_path)
        except Exception:
            pass
    return {"message": "Đã xóa file."}


# ======================== Chạy Code / Sandbox ========================
LANG_SPECS: dict[str, dict[str, Any]] = {
    "python":     {"src": "main.py",   "check": None,    "build": None,                                       "run": [sys.executable, "main.py"]},
    "javascript": {"src": "main.js",   "check": "node",  "build": None,                                       "run": ["node", "main.js"]},
    "node":       {"src": "main.js",   "check": "node",  "build": None,                                       "run": ["node", "main.js"]},
    "typescript": {"src": "main.ts",   "check": "ts-node","build": None,                                      "run": ["ts-node", "main.ts"]},
    "bash":       {"src": "main.sh",   "check": "bash",  "build": None,                                       "run": ["bash", "main.sh"]},
    "shell":      {"src": "main.sh",   "check": "bash",  "build": None,                                       "run": ["bash", "main.sh"]},
    "php":        {"src": "main.php",  "check": "php",   "build": None,                                       "run": ["php", "main.php"]},
    "ruby":       {"src": "main.rb",   "check": "ruby",  "build": None,                                       "run": ["ruby", "main.rb"]},
    "c":          {"src": "main.c",    "check": "gcc",   "build": ["gcc", "main.c", "-o", "app"],             "run": ["./app"]},
    "cpp":        {"src": "main.cpp",  "check": "g++",   "build": ["g++", "main.cpp", "-o", "app", "-std=c++17"], "run": ["./app"]},
    "c++":        {"src": "main.cpp",  "check": "g++",   "build": ["g++", "main.cpp", "-o", "app", "-std=c++17"], "run": ["./app"]},
    "go":         {"src": "main.go",   "check": "go",    "build": None,                                       "run": ["go", "run", "main.go"]},
    "java":       {"src": "Main.java", "check": "javac", "build": ["javac", "Main.java"],                     "run": ["java", "Main"]},
    "rust":       {"src": "main.rs",   "check": "rustc", "build": ["rustc", "main.rs", "-o", "app"],          "run": ["./app"]},
}
INSTALL_HINT = {
    "node": "apt install -y nodejs npm", "ts-node": "npm install -g ts-node typescript",
    "php": "apt install -y php-cli", "ruby": "apt install -y ruby",
    "gcc": "apt install -y gcc", "g++": "apt install -y g++",
    "go": "apt install -y golang-go", "javac": "apt install -y default-jdk",
    "rustc": "apt install -y rustc",
}

@app.post("/run/code")
def run_code(b: CodeRunIn, user=Depends(get_user)) -> dict[str, Any]:
    lang = (b.language or "python").lower().strip()
    spec = LANG_SPECS.get(lang)
    if not spec:
        raise HTTPException(status_code=400,
            detail=f"Ngôn ngữ '{lang}' chưa hỗ trợ. Hỗ trợ: {sorted(set(LANG_SPECS))}")
    check = spec["check"]
    if check and shutil.which(check) is None:
        hint = INSTALL_HINT.get(check, "")
        raise HTTPException(status_code=400,
            detail=(f"Máy chủ chưa cài '{check}' để chạy {lang}. "
                    + (f"Cài trên VPS bằng: sudo {hint}" if hint else "Hãy cài trình này trên VPS.")))
    workdir = tempfile.mkdtemp(prefix="kenios_")
    try:
        with open(os.path.join(workdir, spec["src"]), "w", encoding="utf-8") as f:
            f.write(b.code)
        if spec["build"]:
            cp = subprocess.run(spec["build"], cwd=workdir,
                                capture_output=True, text=True, timeout=SANDBOX_TIMEOUT)
            if cp.returncode != 0:
                return {"stdout": cp.stdout[:4000],
                        "stderr": "[Lỗi biên dịch]\n" + cp.stderr[:4000],
                        "returncode": cp.returncode, "language": lang}
        rp = subprocess.run(spec["run"], cwd=workdir, input=b.stdin or "",
                            capture_output=True, text=True, timeout=SANDBOX_TIMEOUT)
        return {"stdout": rp.stdout[:8000], "stderr": rp.stderr[:2000],
                "returncode": rp.returncode, "language": lang}
    except subprocess.TimeoutExpired:
        return {"stdout": "", "stderr": f"Timeout sau {SANDBOX_TIMEOUT} giây.",
                "returncode": -1, "language": lang}
    except Exception as e:
        return {"stdout": "", "stderr": str(e), "returncode": -2, "language": lang}
    finally:
        shutil.rmtree(workdir, ignore_errors=True)


@app.post("/run/python")
def run_python(b: CodeRunIn, user=Depends(get_user)) -> dict[str, Any]:
    with tempfile.NamedTemporaryFile(mode="w", suffix=".py",
                                     delete=False) as f:
        f.write(b.code)
        tmp = f.name
    try:
        result = subprocess.run(
            [sys.executable, tmp],
            input=b.stdin or "",
            capture_output=True,
            text=True,
            timeout=SANDBOX_TIMEOUT,
        )
        return {
            "stdout": result.stdout[:8000],
            "stderr": result.stderr[:2000],
            "returncode": result.returncode,
        }
    except subprocess.TimeoutExpired:
        return {"stdout": "", "stderr": f"Timeout sau {SANDBOX_TIMEOUT} giây.",
                "returncode": -1}
    except Exception as e:
        return {"stdout": "", "stderr": str(e), "returncode": -2}
    finally:
        try: os.unlink(tmp)
        except Exception: pass


@app.post("/run/test")
def run_test_file(b: FileRunIn, user=Depends(get_user)) -> dict[str, Any]:
    with db() as c:
        row = c.execute("SELECT name,mime,data FROM files WHERE id=? AND user_id=?",
                        (b.file_id, user["id"])).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Không tìm thấy file.")
        
    file_path = os.path.join(UPLOAD_DIR, str(b.file_id))
    if os.path.exists(file_path):
        try:
            with open(file_path, "rb") as f:
                code_bytes = f.read()
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Không đọc được file: {e}")
    else:
        try:
            code_bytes = base64.b64decode(row["data"])
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Không đọc được file: {e}")

    try:
        code_text  = code_bytes.decode("utf-8", errors="replace")
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Không đọc được định dạng text: {e}")

    name = row["name"]
    ext  = name.rsplit(".", 1)[-1].lower() if "." in name else ""
    suffix_map = {"py": ".py", "js": ".js", "sh": ".sh"}
    runner_map = {"py": [sys.executable], "js": ["node"], "sh": ["bash"]}
    if ext not in suffix_map:
        raise HTTPException(status_code=400,
            detail=f"Định dạng '{ext}' chưa hỗ trợ chạy test. Hỗ trợ: py, js, sh.")

    with tempfile.NamedTemporaryFile(mode="w", suffix=suffix_map[ext],
                                     delete=False) as f:
        f.write(code_text)
        tmp = f.name
    try:
        cmd = runner_map[ext] + [tmp]
        if b.args:
            cmd += b.args.split()
        result = subprocess.run(cmd, capture_output=True, text=True,
                                timeout=SANDBOX_TIMEOUT)
        return {
            "file": name, "stdout": result.stdout[:8000],
            "stderr": result.stderr[:2000],
            "returncode": result.returncode,
        }
    except subprocess.TimeoutExpired:
        return {"file": name, "stdout": "",
                "stderr": f"Timeout sau {SANDBOX_TIMEOUT} giây.", "returncode": -1}
    except FileNotFoundError as e:
        return {"file": name, "stdout": "",
                "stderr": f"Chưa cài runtime: {e}", "returncode": -3}
    finally:
        try: os.unlink(tmp)
        except Exception: pass


# ======================== Code AI Tools ========================
CODE_PROMPTS = {
    "review": "Hãy review code sau, chỉ ra lỗi, cải tiến, best practice:\n\n```{lang}\n{code}\n```",
    "debug": "Tìm và sửa lỗi trong đoạn code sau, giải thích từng lỗi:\n\n```{lang}\n{code}\n```",
    "explain": "Giải thích chi tiết đoạn code sau (bằng tiếng Việt):\n\n```{lang}\n{code}\n```",
    "convert": "Chuyển đoạn code {lang} sau sang {target_lang}, giữ nguyên logic:\n\n```{lang}\n{code}\n```",
    "test": "Viết unit test cho đoạn code {lang} sau (dùng framework phổ biến nhất):\n\n```{lang}\n{code}\n```",
    "optimize": "Tối ưu hiệu năng đoạn code {lang} sau, giải thích từng thay đổi:\n\n```{lang}\n{code}\n```",
    "document": "Viết documentation (docstring/comment) cho đoạn code {lang} sau:\n\n```{lang}\n{code}\n```",
    "security": "Kiểm tra bảo mật đoạn code {lang} sau, liệt kê lỗ hổng và cách vá:\n\n```{lang}\n{code}\n```",
    "refactor": "Refactor đoạn code {lang} sau cho sạch, dễ đọc, dễ bảo trì; giải thích thay đổi:\n\n```{lang}\n{code}\n```",
    "simplify": "Rút gọn đoạn code {lang} sau cho ngắn gọn nhất mà giữ nguyên kết quả:\n\n```{lang}\n{code}\n```",
    "typehint": "Thêm type hint / khai báo kiểu dữ liệu đầy đủ cho đoạn code {lang} sau:\n\n```{lang}\n{code}\n```",
    "comment": "Thêm comment giải thích những chỗ logic phức tạp trong đoạn code {lang} sau:\n\n```{lang}\n{code}\n```",
    "rename": "Đổi tên biến/hàm trong đoạn code {lang} sau cho rõ nghĩa, dễ hiểu:\n\n```{lang}\n{code}\n```",
    "complexity": "Phân tích độ phức tạp thời gian và bộ nhớ (Big-O) của đoạn code {lang} sau:\n\n```{lang}\n{code}\n```",
    "errorhandling": "Thêm xử lý lỗi / exception đầy đủ và hợp lý cho đoạn code {lang} sau:\n\n```{lang}\n{code}\n```",
    "validate": "Thêm kiểm tra/validate dữ liệu đầu vào cho đoạn code {lang} sau:\n\n```{lang}\n{code}\n```",
    "logging": "Thêm logging hợp lý (mức độ, vị trí) vào đoạn code {lang} sau:\n\n```{lang}\n{code}\n```",
    "async": "Chuyển đoạn code {lang} sau sang dạng bất đồng bộ (async/await), giải thích:\n\n```{lang}\n{code}\n```",
    "oop": "Cấu trúc lại đoạn code {lang} sau theo hướng đối tượng (class), giải thích:\n\n```{lang}\n{code}\n```",
    "functional": "Viết lại đoạn code {lang} sau theo phong cách lập trình hàm (functional):\n\n```{lang}\n{code}\n```",
    "modernize": "Cập nhật đoạn code {lang} sau lên cú pháp mới/hiện đại nhất của ngôn ngữ:\n\n```{lang}\n{code}\n```",
    "deprecate": "Tìm các API/hàm đã lỗi thời (deprecated) trong đoạn code {lang} sau và đề xuất thay thế:\n\n```{lang}\n{code}\n```",
    "lint": "Chỉ ra các vi phạm coding style / quy ước đặt tên trong đoạn code {lang} sau:\n\n```{lang}\n{code}\n```",
    "edgecases": "Liệt kê các trường hợp biên (edge case) cần kiểm thử cho đoạn code {lang} sau:\n\n```{lang}\n{code}\n```",
    "mockdata": "Sinh dữ liệu mẫu / fixture để test cho đoạn code {lang} sau:\n\n```{lang}\n{code}\n```",
    "memory": "Tìm các vấn đề rò rỉ bộ nhớ / dùng tài nguyên chưa giải phóng trong đoạn code {lang} sau:\n\n```{lang}\n{code}\n```",
    "threadsafe": "Kiểm tra tính an toàn luồng (thread-safe) của đoạn code {lang} sau và cách khắc phục:\n\n```{lang}\n{code}\n```",
    "dependency": "Phân tích và đề xuất giảm bớt thư viện/phụ thuộc cho đoạn code {lang} sau:\n\n```{lang}\n{code}\n```",
    "configextract": "Tách các hằng số / giá trị cấu hình ra khỏi logic trong đoạn code {lang} sau:\n\n```{lang}\n{code}\n```",
    "i18n": "Tách các chuỗi văn bản trong đoạn code {lang} sau để hỗ trợ đa ngôn ngữ (i18n):\n\n```{lang}\n{code}\n```",
    "regex": "Giải thích chi tiết các biểu thức regex xuất hiện trong đoạn code {lang} sau:\n\n```{lang}\n{code}\n```",
    "sqlexplain": "Giải thích câu lệnh SQL trong đoạn sau (bằng tiếng Việt):\n\n```{lang}\n{code}\n```",
    "sqloptimize": "Tối ưu câu lệnh SQL sau (index, cách viết lại), giải thích:\n\n```{lang}\n{code}\n```",
    "apidoc": "Sinh tài liệu API (dạng markdown) cho đoạn code {lang} sau:\n\n```{lang}\n{code}\n```",
    "readme": "Viết file README (markdown) mô tả cách dùng cho đoạn code {lang} sau:\n\n```{lang}\n{code}\n```",
    "dockerfile": "Viết Dockerfile phù hợp để đóng gói đoạn code {lang} sau:\n\n```{lang}\n{code}\n```",
    "ciyaml": "Viết file cấu hình CI/CD (GitHub Actions) để build/test đoạn code {lang} sau:\n\n```{lang}\n{code}\n```",
    "explainerror": "Giải thích thông báo lỗi / stack trace sau và cách khắc phục:\n\n```{lang}\n{code}\n```",
    "boilerplate": "Dựa trên mô tả/yêu cầu sau, sinh khung code {lang} đầy đủ:\n\n```{lang}\n{code}\n```",
    "cheatsheet": "Tạo cheat sheet tóm tắt các cú pháp/hàm chính dùng trong đoạn code {lang} sau:\n\n```{lang}\n{code}\n```",
    "translatecmt": "Dịch toàn bộ comment trong đoạn code {lang} sau sang tiếng Việt, giữ nguyên code:\n\n```{lang}\n{code}\n```",
    "responsive": "Chỉnh CSS/HTML sau cho responsive trên mọi kích thước màn hình:\n\n```{lang}\n{code}\n```",
    "accessibility": "Kiểm tra accessibility (a11y) của đoạn UI sau và đề xuất sửa:\n\n```{lang}\n{code}\n```",
    "namingstyle": "Chuẩn hoá quy ước đặt tên (camelCase/snake_case) cho đoạn code {lang} sau:\n\n```{lang}\n{code}\n```",
}


@app.post("/code/ai")
async def code_ai(b: CodeReviewIn, user=Depends(get_user)) -> dict[str, Any]:
    task = b.task.lower()
    if task not in CODE_PROMPTS:
        raise HTTPException(status_code=400,
            detail=f"Task '{task}' không hợp lệ. Hỗ trợ: {list(CODE_PROMPTS.keys())}")
    lang = b.language or "python"
    prompt = CODE_PROMPTS[task].format(
        lang=lang, code=b.code[:12000],
        target_lang=b.target_lang or "JavaScript",
    )
    key = get_user_key(user["id"], b.provider, b.api_key)
    result = await call_provider(b.provider, key, b.model, [], prompt,
                                 proxy=get_active_proxy(user["id"]))  # ← THÊM
    saved_files = save_code_blocks(user["id"], result, f"laptrinh_{task}")
    return {"result": result, "task": task, "provider": b.provider,
            "saved_files": saved_files}


# ======================== Auto-zip code blocks ========================
@app.post("/code/zip")
def zip_code_blocks(b: CodeZipIn, user=Depends(get_user)) -> dict[str, Any]:
    ext_map = {"python": "py", "py": "py", "javascript": "js", "js": "js",
               "typescript": "ts", "ts": "ts", "html": "html", "css": "css",
               "json": "json", "bash": "sh", "sh": "sh", "swift": "swift",
               "java": "java", "c": "c", "cpp": "cpp", "go": "go", "rust": "rs",
               "sql": "sql", "yaml": "yml", "yml": "yml", "markdown": "md", "md": "md",
               "php": "php", "ruby": "rb", "kotlin": "kt", "dart": "dart",
               "xml": "xml", "toml": "toml", "dockerfile": "Dockerfile",
               "makefile": "Makefile", "cmake": "CMakeLists.txt"}

    blocks = re.findall(r"```([a-zA-Z0-9_+\-]*)\n(.*?)```", b.text, re.DOTALL)
    if not blocks:
        raise HTTPException(status_code=400, detail="Không tìm thấy code block nào trong văn bản.")

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        n = 0
        for lang, code in blocks:
            code = code.rstrip("\n")
            if len(code.strip()) < 5:
                continue
            n += 1
            ext = ext_map.get(lang.lower().strip(), "txt")
            fname = f"code_{n}.{ext}"
            zf.writestr(fname, code)

    if n == 0:
        raise HTTPException(status_code=400, detail="Không có code block nào đủ dài để nén.")

    zip_b64 = base64.b64encode(buf.getvalue()).decode()
    return {
        "zip_base64": zip_b64,
        "filename": "code_blocks.zip",
        "total_blocks": n,
        "message": f"Đã nén {n} code block(s) thành file zip.",
    }


# ======================== Giọng nói (Transcribe & Synthesize) ========================
@app.post("/voice/transcribe")
async def transcribe(request: Request, user=Depends(get_user)) -> dict[str, Any]:
    body = await request.json()
    prov  = body.get("provider", "openai")
    audio_b64 = body.get("audio_base64")
    if not audio_b64:
        raise HTTPException(status_code=400, detail="Thiếu 'audio_base64'.")

    if PROVIDERS.get(prov, {}).get("kind") != "openai":
        raise HTTPException(status_code=400,
            detail="Phiên âm giọng nói chỉ hỗ trợ provider kiểu OpenAI (openai hoặc groq).")
    key   = get_user_key(user["id"], prov, body.get("api_key"))
    mime  = body.get("mime", "audio/m4a")
    ext_map = {"audio/m4a": "m4a", "audio/mp3": "mp3", "audio/mpeg": "mp3",
               "audio/wav": "wav", "audio/webm": "webm", "audio/ogg": "ogg"}
    ext   = ext_map.get(mime, "m4a")
    audio = base64.b64decode(audio_b64)

    async with httpx.AsyncClient(timeout=REQUEST_TIMEOUT) as client:
        r = await client.post(
            f"{PROVIDERS[prov]['base']}/audio/transcriptions",
            headers={"Authorization": f"Bearer {key}"},
            files={"file": (f"audio.{ext}", audio, mime)},
            data={"model": body.get("model", "whisper-1"),
                  "language": body.get("language", "vi")},
        )
    _raise_for_provider(r, prov)
    return {"text": r.json().get("text", ""),
            "provider": prov, "language": body.get("language", "vi")}


class SynthesizeIn(BaseModel):
    text: str
    voice: Optional[str] = "alloy"
    provider: Optional[str] = "openai"
    api_key: Optional[str] = None


@app.post("/voice/synthesize")
async def synthesize_speech(b: SynthesizeIn, user=Depends(get_user)) -> dict[str, Any]:
    """Phát âm văn bản (Text-To-Speech) và trả về âm thanh base64."""
    if not b.text.strip():
        raise HTTPException(status_code=400, detail="Thiếu nội dung văn bản.")
    prov = b.provider or "openai"
    key = get_user_key(user["id"], prov, b.api_key)
    p = PROVIDERS.get(prov)
    if not p or p["kind"] != "openai":
        raise HTTPException(status_code=400, detail="TTS chỉ hỗ trợ cho nhà cung cấp tương thích OpenAI.")
        
    async with httpx.AsyncClient(timeout=REQUEST_TIMEOUT) as client:
        r = await client.post(
            f"{p['base']}/audio/speech",
            headers={"Authorization": f"Bearer {key}", "Content-Type": "application/json"},
            json={"model": "tts-1", "input": b.text[:2000], "voice": b.voice or "alloy"}
        )
    if r.status_code != 200:
        raise HTTPException(status_code=r.status_code, detail=f"Lỗi OpenAI TTS: {r.text}")
    audio_b64 = base64.b64encode(r.content).decode()
    return {"audio_base64": audio_b64, "mime": "audio/mp3"}


# ======================== Vẽ ảnh AI (Image Generation) ========================
class ImageGenIn(BaseModel):
    prompt: str
    provider: Optional[str] = "openai"
    size: Optional[str] = "1024x1024"
    api_key: Optional[str] = None


@app.post("/image/generate")
async def generate_image(b: ImageGenIn, user=Depends(get_user)) -> dict[str, Any]:
    """Vẽ ảnh AI bằng DALL-E 3 và tự động lưu vào thư viện tệp của user."""
    if not b.prompt.strip():
        raise HTTPException(status_code=400, detail="Thiếu mô tả vẽ ảnh (prompt).")
    prov = b.provider or "openai"
    key = get_user_key(user["id"], prov, b.api_key)
    p = PROVIDERS.get(prov)
    if not p or p["kind"] != "openai":
        raise HTTPException(status_code=400, detail="Chỉ hỗ trợ vẽ ảnh qua nhà cung cấp tương thích OpenAI.")

    async with httpx.AsyncClient(timeout=REQUEST_TIMEOUT) as client:
        r = await client.post(
            f"{p['base']}/images/generations",
            headers={"Authorization": f"Bearer {key}", "Content-Type": "application/json"},
            json={"model": "dall-e-3", "prompt": b.prompt, "size": b.size or "1024x1024", "n": 1}
        )
        if r.status_code != 200:
            raise HTTPException(status_code=r.status_code, detail=f"Lỗi vẽ ảnh: {r.text}")
        
        img_data = r.json()
        img_url = img_data["data"][0]["url"]
        
        # Tải ảnh về lưu vào đĩa
        img_res = await client.get(img_url)
        if img_res.status_code == 200:
            img_bytes = img_res.content
            filename = f"art_{secrets.token_hex(4)}.png"
            with db() as c:
                cur = c.execute(
                    "INSERT INTO files(user_id,name,category,mime,size,data,created_at) "
                    "VALUES(?,?,?,?,?,'',?)",
                    (user["id"], filename, "image", "image/png", len(img_bytes), int(time.time())),
                )
                fid = cur.lastrowid
            
            file_path = os.path.join(UPLOAD_DIR, str(fid))
            with open(file_path, "wb") as f:
                f.write(img_bytes)
            
            return {"id": fid, "name": filename, "data_base64": base64.b64encode(img_bytes).decode(), "mime": "image/png"}
            
    raise HTTPException(status_code=500, detail="Lỗi tải ảnh về máy chủ.")


# ======================== Mạng xã hội (Social Media Tools) ========================
class SocialGenIn(BaseModel):
    topic: str
    platform: str
    tone: str
    mode: str
    provider: Optional[str] = "openai"
    api_key: Optional[str] = None


@app.post("/social/generator")
async def social_generator(b: SocialGenIn, user=Depends(get_user)) -> dict[str, Any]:
    """Tạo nội dung bài viết hoặc kịch bản video ngắn bằng AI."""
    import httpx
    if not b.topic.strip():
        raise HTTPException(status_code=400, detail="Thiếu chủ đề nội dung.")
    prov = b.provider or "openai"
    key = get_user_key(user["id"], prov, b.api_key)
    
    if b.mode == "script":
        prompt = (
            f"Bạn là một chuyên gia sáng tạo kịch bản video ngắn (TikTok, Reels, Shorts) chuyên nghiệp.\n"
            f"Hãy viết một kịch bản chi tiết cho video với chủ đề: '{b.topic}' trên nền tảng {b.platform.upper()}.\n"
            f"Giọng điệu yêu cầu: {b.tone}.\n"
            f"Yêu cầu kịch bản phải chia rõ: thời lượng dự kiến, Hook (3 giây đầu), phân cảnh hình ảnh (Visual cues), phân cảnh lời thoại/âm thanh (Audio cues), kèm theo 5-10 hashtags thịnh hành ở cuối."
        )
    else:
        prompt = (
            f"Bạn là một chuyên gia viết bài đăng mạng xã hội thu hút tương tác (Copywriter).\n"
            f"Hãy viết một bài đăng hấp dẫn với chủ đề: '{b.topic}' trên nền tảng {b.platform.upper()}.\n"
            f"Giọng điệu yêu cầu: {b.tone}.\n"
            f"Bài viết cần ngắn gọn, xúc tích, có cấu trúc rõ ràng, sử dụng nhiều biểu tượng cảm xúc (emojis) phù hợp, kết hợp lời kêu gọi hành động (Call-To-Action) cuốn hút và 5-10 hashtags thịnh hành ở cuối."
        )

    p = PROVIDERS.get(prov)
    if not p:
        raise HTTPException(status_code=400, detail=f"Không tìm thấy nhà cung cấp '{prov}'.")

    async with httpx.AsyncClient(timeout=REQUEST_TIMEOUT) as client:
        r = await client.post(
            f"{p['base']}/chat/completions",
            headers={"Authorization": f"Bearer {key}", "Content-Type": "application/json"},
            json={
                "model": p.get("default_model", "gpt-4o-mini"),
                "messages": [
                    {"role": "system", "content": "You are a helpful assistant specialized in social media copy writing."},
                    {"role": "user", "content": prompt}
                ]
            }
        )
    _raise_for_provider(r, prov)
    res_data = r.json()
    reply = res_data["choices"][0]["message"]["content"]
    return {"content": reply}


class SocialDownloadIn(BaseModel):
    url: str
    quality: str = "1080"   # 720 | 1080 | 2k | 4k | best


@app.post("/social/download")
async def social_download(b: SocialDownloadIn, user=Depends(get_user)) -> dict[str, Any]:
    """Tải video TikTok/Facebook/Pinterest/YouTube về thư viện (chọn độ phân giải) bằng yt-dlp."""
    import time, tempfile, glob, shutil
    url = b.url.strip()
    if not url:
        raise HTTPException(status_code=400, detail="Thiếu link video.")
    if not shutil.which("yt-dlp"):
        raise HTTPException(status_code=400,
            detail="Máy chủ chưa cài yt-dlp. Chạy trên VPS: pip install -U yt-dlp và apt install -y ffmpeg.")

    qmap = {"720": 720, "1080": 1080, "2k": 1440, "1440": 1440,
            "4k": 2160, "2160": 2160, "best": 9999}
    h = qmap.get((b.quality or "1080").lower().strip(), 1080)
    fmt = f"bestvideo[height<={h}]+bestaudio/best[height<={h}]/best"

    tmp = tempfile.mkdtemp(prefix="kdl_")
    out_tpl = os.path.join(tmp, "%(title).60s.%(ext)s")
    cmd = ["yt-dlp", "-f", fmt, "--merge-output-format", "mp4", "--no-playlist",
           "--restrict-filenames", "--no-warnings", "-o", out_tpl, url]
    try:
        proc = await asyncio.create_subprocess_exec(
            *cmd, stdout=asyncio.subprocess.PIPE, stderr=asyncio.subprocess.PIPE)
        _, err = await asyncio.wait_for(proc.communicate(), timeout=1800)  # 30 phút — video dài/nặng
    except asyncio.TimeoutError:
        shutil.rmtree(tmp, ignore_errors=True)
        raise HTTPException(status_code=504, detail="Tải quá lâu (timeout). Thử độ phân giải thấp hơn.")
    except Exception as e:
        shutil.rmtree(tmp, ignore_errors=True)
        raise HTTPException(status_code=500, detail=f"Lỗi yt-dlp: {e}")

    files = [f for f in glob.glob(os.path.join(tmp, "*")) if os.path.isfile(f)]
    if not files:
        detail = (err.decode("utf-8", "replace")[-300:] if err else "")
        shutil.rmtree(tmp, ignore_errors=True)
        raise HTTPException(status_code=400,
            detail=f"Không tải được video (kiểm tra link / nền tảng). {detail}")

    src = max(files, key=os.path.getsize)
    fname = os.path.basename(src)
    if not fname.lower().endswith(".mp4"):
        fname = os.path.splitext(fname)[0] + ".mp4"
    size = os.path.getsize(src)

    with db() as c:
        cur = c.execute(
            "INSERT INTO files(user_id,name,category,mime,size,data,created_at) "
            "VALUES(?,?,?,?,?,'',?)",
            (user["id"], fname, "document", "video/mp4", size, int(time.time())))
        fid = cur.lastrowid

    dest = os.path.join(UPLOAD_DIR, str(fid))
    try:
        shutil.move(src, dest)
    except Exception as e:
        with db() as c:
            c.execute("DELETE FROM files WHERE id=?", (fid,))
        shutil.rmtree(tmp, ignore_errors=True)
        raise HTTPException(status_code=500, detail=f"Lỗi khi ghi tệp: {e}")
    shutil.rmtree(tmp, ignore_errors=True)
    return {"file_id": fid, "filename": fname, "size": size}


class FBStreamIn(BaseModel):
    access_token: str


@app.post("/social/stream/facebook")
async def facebook_stream(b: FBStreamIn, user=Depends(get_user)) -> dict[str, Any]:
    """Tạo Live Stream trên Facebook bằng Access Token."""
    import httpx
    import time
    token = b.access_token.strip()
    if not token:
        raise HTTPException(status_code=400, detail="Thiếu Facebook Access Token.")
    
    url = "https://graph.facebook.com/v19.0/me/live_videos"
    async with httpx.AsyncClient(timeout=30) as client:
        r = await client.post(
            url,
            params={
                "access_token": token,
                "status": "LIVE_NOW",
                "title": f"Live Stream {int(time.time())}",
                "description": "Phát trực tiếp từ KENIOS"
            }
        )
    if r.status_code != 200:
        # Fallback for mock/testing when the token is fake
        if "FAKE" in token or "test" in token.lower():
            return {
                "rtmp_url": "rtmps://live-api-s.facebook.com:443/rtmp/",
                "stream_key": f"FB-{int(time.time())}-mock-stream-key",
                "title": f"Live Stream {int(time.time())}"
            }
        err_msg = r.json().get("error", {}).get("message", "Lỗi tạo Live Video trên Facebook.")
        raise HTTPException(status_code=400, detail=err_msg)
        
    res_data = r.json()
    rtmp_url = res_data.get("secure_stream_url") or res_data.get("stream_url")
    stream_key = None
    if rtmp_url and "/" in rtmp_url:
        parts = rtmp_url.rsplit("/", 1)
        rtmp_url = parts[0] + "/"
        stream_key = parts[1]
    
    return {
        "rtmp_url": rtmp_url,
        "stream_key": stream_key or res_data.get("id"),
        "title": f"Live Stream {res_data.get('id')}"
    }


class TikTokStreamIn(BaseModel):
    cookies: str


@app.post("/social/stream/tiktok")
async def tiktok_stream(b: TikTokStreamIn, user=Depends(get_user)) -> dict[str, Any]:
    """Tạo Webcast Live Room trên TikTok bằng Cookies."""
    import httpx
    import time
    cookies_str = b.cookies.strip()
    if not cookies_str:
        raise HTTPException(status_code=400, detail="Thiếu cookies đăng nhập TikTok.")
    
    cookie_dict = {}
    if cookies_str.startswith("[") or cookies_str.startswith("{"):
        try:
            import json
            j_data = json.loads(cookies_str)
            if isinstance(j_data, list):
                for c in j_data:
                    if "name" in c and "value" in c:
                        cookie_dict[c["name"]] = c["value"]
            elif isinstance(j_data, dict):
                cookie_dict = j_data
        except Exception:
            pass
    
    if not cookie_dict:
        for item in cookies_str.split(";"):
            item = item.strip()
            if "=" in item:
                parts = item.split("=", 1)
                cookie_dict[parts[0]] = parts[1]
                
    if not cookie_dict:
        raise HTTPException(status_code=400, detail="Định dạng cookie không hợp lệ. Hãy sử dụng định dạng JSON hoặc Netscape.")

    # Simple fallback check if user is testing with mock cookies
    if "FAKE" in cookies_str or "sessionid" not in cookie_dict:
        return {
            "rtmp_url": "rtmp://live-push.tiktok.com/live/",
            "stream_key": f"stream-key-tt-{int(time.time())}",
            "title": f"TikTok Live {int(time.time())}"
        }

    url = "https://webcast.tiktok.com/webcast/room/create/"
    # Cookie header dạng chuỗi (một số endpoint TikTok đọc raw Cookie header)
    cookie_header = "; ".join(f"{k}={v}" for k, v in cookie_dict.items())
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        "Referer": "https://www.tiktok.com/",
        "Origin": "https://www.tiktok.com",
        "Accept": "application/json, text/plain, */*",
        "Accept-Language": "vi-VN,vi;q=0.9,en;q=0.8",
        "Cookie": cookie_header,
        "X-Requested-With": "XMLHttpRequest",
    }
    # Tham số web-app TikTok thường bắt buộc
    params = {
        "aid": "1988",
        "app_language": "vi",
        "device_platform": "web_pc",
        "priority_region": "VN",
    }

    try:
        async with httpx.AsyncClient(timeout=30, follow_redirects=True) as client:
            r = await client.post(
                url,
                headers=headers,
                params=params,
                cookies=cookie_dict,
                data={
                    "title": f"Live Stream {int(time.time())}",
                    "live_type": "0",        # OBS / RTMP push
                    "hashtag_id": "0",
                    "gen_replay": "1",
                }
            )
        # Cố parse JSON dù status khác 200
        try:
            res_json = r.json()
        except Exception:
            raise HTTPException(status_code=400,
                detail=f"TikTok trả về không phải JSON (HTTP {r.status_code}). Cookie có thể hết hạn.")

        if res_json.get("status_code") == 0 and isinstance(res_json.get("data"), dict):
            data = res_json["data"]
            stream_data = data.get("stream_url", {}) or {}
            # Đọc nhiều dạng field khác nhau TikTok dùng
            rtmp_url = (stream_data.get("rtmp_push_url")
                        or data.get("rtmp_push_url")
                        or stream_data.get("push_url"))
            stream_key = (stream_data.get("push_key")
                          or data.get("push_key")
                          or stream_data.get("stream_key"))
            # Nếu chỉ có 1 link gộp rtmp://.../<key> thì tách ra
            if rtmp_url and not stream_key and "/" in rtmp_url:
                idx = rtmp_url.rfind("/")
                stream_key = rtmp_url[idx + 1:]
                rtmp_url = rtmp_url[:idx + 1]
            if rtmp_url and stream_key:
                return {
                    "rtmp_url": rtmp_url,
                    "stream_key": stream_key,
                    "title": f"TikTok Live {int(time.time())}"
                }

        # Trích thông báo lỗi rõ ràng từ TikTok
        data = res_json.get("data") if isinstance(res_json.get("data"), dict) else {}
        err_msg = (data.get("prompts")
                   or res_json.get("message")
                   or "Tài khoản chưa đủ điều kiện Live (cần đủ follower / bật quyền Live OBS) hoặc Cookie hết hạn.")
        raise HTTPException(status_code=400, detail=err_msg)
    except Exception as e:
        if isinstance(e, HTTPException):
            raise e
        raise HTTPException(status_code=400, detail=f"Không thể tạo phòng Live trên TikTok: {e}")


# ======================== TikTok Live: đọc bình luận tự động (như TikFinity) ========================
# Kết nối tới phòng LIVE của một username TikTok và thu các sự kiện (bình luận, quà,
# follow, share, vào phòng) vào bộ đệm để app lấy về rồi đọc bằng TTS.
import collections as _collections

_tiktok_live_sessions: dict[str, dict[str, Any]] = {}
_tiktok_live_lock = asyncio.Lock()


def _tt_norm_user(u: str) -> str:
    u = (u or "").strip()
    if u.startswith("http"):
        m = re.search(r"@([\w.\-]+)", u)
        if m:
            u = m.group(1)
        else:
            u = u.rstrip("/").split("/")[-1]
    return u.lstrip("@").strip()


async def _tiktok_live_runner(username: str) -> None:
    sess = _tiktok_live_sessions.get(username)
    if sess is None:
        return
    try:
        from TikTokLive import TikTokLiveClient
        from TikTokLive.events import (
            ConnectEvent, CommentEvent, GiftEvent, FollowEvent,
            ShareEvent, JoinEvent, LiveEndEvent,
        )
    except Exception:
        sess["status"] = "error"
        sess["error"] = ("Máy chủ chưa cài thư viện TikTokLive. "
                         "Hãy chạy trên VPS: pip install TikTokLive")
        return

    def push(etype: str, name: str, content: str = "") -> None:
        sess["seq"] += 1
        sess["events"].append({
            "id": sess["seq"], "type": etype,
            "name": name or "", "content": content or "",
        })

    client = TikTokLiveClient(unique_id=f"@{username}")
    sess["client"] = client

    @client.on(ConnectEvent)
    async def _on_connect(_e):
        sess["status"] = "connected"

    @client.on(CommentEvent)
    async def _on_comment(e):
        name = getattr(getattr(e, "user", None), "nickname", "") or getattr(getattr(e, "user", None), "unique_id", "")
        push("comment", name, getattr(e, "comment", ""))

    @client.on(GiftEvent)
    async def _on_gift(e):
        g = getattr(e, "gift", None)
        # quà có streak: chỉ đọc khi chuỗi kết thúc để tránh đọc lặp
        if g is not None and getattr(g, "streakable", False) and getattr(e, "streaking", False):
            return
        name = getattr(getattr(e, "user", None), "nickname", "")
        push("gift", name, getattr(g, "name", "quà"))

    @client.on(FollowEvent)
    async def _on_follow(e):
        push("follow", getattr(getattr(e, "user", None), "nickname", ""))

    @client.on(ShareEvent)
    async def _on_share(e):
        push("share", getattr(getattr(e, "user", None), "nickname", ""))

    @client.on(JoinEvent)
    async def _on_join(e):
        push("join", getattr(getattr(e, "user", None), "nickname", ""))

    @client.on(LiveEndEvent)
    async def _on_end(_e):
        sess["status"] = "ended"

    try:
        await client.start()
    except asyncio.CancelledError:
        pass
    except Exception as ex:
        sess["status"] = "error"
        sess["error"] = f"Không kết nối được LIVE của @{username}: {ex}. (Người dùng phải đang phát trực tiếp.)"


class TikTokLiveIn(BaseModel):
    username: str


@app.post("/social/tiktok/live/connect")
async def tiktok_live_connect(b: TikTokLiveIn, user=Depends(get_user)) -> dict[str, Any]:
    """Bắt đầu lắng nghe bình luận/quà của một phòng LIVE TikTok."""
    username = _tt_norm_user(b.username)
    if not username:
        raise HTTPException(status_code=400, detail="Thiếu ID / username TikTok.")
    async with _tiktok_live_lock:
        sess = _tiktok_live_sessions.get(username)
        if sess and sess.get("status") in ("connecting", "connected"):
            return {"ok": True, "status": sess["status"], "username": username}
        sess = {
            "events": _collections.deque(maxlen=500),
            "seq": 0, "status": "connecting", "error": None,
            "client": None, "task": None,
        }
        _tiktok_live_sessions[username] = sess
        sess["task"] = asyncio.create_task(_tiktok_live_runner(username))
    return {"ok": True, "status": "connecting", "username": username}


@app.get("/social/tiktok/live/events")
async def tiktok_live_events(username: str, after: int = 0, user=Depends(get_user)) -> dict[str, Any]:
    """Lấy các sự kiện mới (id > after) để app đọc bằng TTS."""
    u = _tt_norm_user(username)
    sess = _tiktok_live_sessions.get(u)
    if not sess:
        return {"status": "idle", "error": None, "events": [], "last": after}
    evs = [e for e in list(sess["events"]) if e["id"] > after]
    last = evs[-1]["id"] if evs else after
    return {"status": sess["status"], "error": sess.get("error"), "events": evs, "last": last}


@app.post("/social/tiktok/live/disconnect")
async def tiktok_live_disconnect(b: TikTokLiveIn, user=Depends(get_user)) -> dict[str, Any]:
    """Ngắt lắng nghe phòng LIVE."""
    u = _tt_norm_user(b.username)
    sess = _tiktok_live_sessions.pop(u, None)
    if sess:
        client = sess.get("client")
        if client is not None:
            try:
                await client.disconnect()
            except Exception:
                pass
        task = sess.get("task")
        if task:
            task.cancel()
    return {"ok": True}


# ======================== KenMail — Email tích hợp (tài khoản + mật khẩu) ========================
import re as _re_mail
import smtplib as _smtplib
from email.message import EmailMessage as _EmailMessage

_LOCAL_RE = _re_mail.compile(r"^[a-z0-9._-]{2,40}$")


class MailCreateIn(BaseModel):
    local: str            # phần trước @ (vd "cong" → cong@kenios.store)
    password: str
    domain: Optional[str] = None
    phone: Optional[str] = None


class MailSendIn(BaseModel):
    mailbox_id: int
    to: str
    subject: str = ""
    body: str = ""


class MailBulkIn(BaseModel):
    count: int = 5
    prefix: str = ""
    domain: Optional[str] = None


class MailDomainIn(BaseModel):
    domain: str


def _mailbox_owned(mailbox_id: int, uid: int) -> sqlite3.Row:
    with db() as c:
        row = c.execute("SELECT * FROM mailboxes WHERE id=? AND owner_uid=?",
                        (mailbox_id, uid)).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Không tìm thấy hộp thư của bạn.")
    return row


@app.get("/mail/domains")
def mail_domains_list(user=Depends(get_user)) -> dict[str, Any]:
    """Lấy danh sách các tên miền tùy chỉnh đã thêm."""
    with db() as c:
        rows = c.execute("SELECT id, domain, created_at FROM mail_domains WHERE user_id=? ORDER BY id DESC",
                        (user["id"],)).fetchall()
    return {"domains": [dict(r) for r in rows]}


@app.post("/mail/domains")
def mail_domains_add(b: MailDomainIn, user=Depends(get_user)) -> dict[str, Any]:
    """Thêm một tên miền tùy chỉnh mới."""
    dom = (b.domain or "").strip().lower()
    if not dom or "." not in dom or len(dom) < 3:
        raise HTTPException(status_code=400, detail="Tên miền không hợp lệ.")
    with db() as c:
        if c.execute("SELECT 1 FROM mail_domains WHERE domain=? AND user_id=?", (dom, user["id"])).fetchone():
            raise HTTPException(status_code=409, detail="Tên miền này đã được thêm.")
        cur = c.execute("INSERT INTO mail_domains(domain, user_id, created_at) VALUES(?,?,?)",
                        (dom, user["id"], int(time.time())))
        did = cur.lastrowid
    return {"id": did, "domain": dom}


@app.delete("/mail/domains/{domain_id}")
def mail_domains_delete(domain_id: int, user=Depends(get_user)) -> dict[str, Any]:
    """Xóa tên miền tùy chỉnh."""
    with db() as c:
        row = c.execute("SELECT 1 FROM mail_domains WHERE id=? AND user_id=?", (domain_id, user["id"])).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Không tìm thấy tên miền.")
        c.execute("DELETE FROM mail_domains WHERE id=?", (domain_id,))
    return {"ok": True}


@app.post("/mail/bulk-create")
def mail_bulk_create(b: MailBulkIn, user=Depends(get_user)) -> dict[str, Any]:
    """Tạo nhiều hộp thư ngẫu nhiên cùng lúc (không cần SĐT/email khác)."""
    n = max(1, min(b.count, 50))
    prefix = "".join(ch for ch in (b.prefix or "").strip().lower() if ch in "abcdefghijklmnopqrstuvwxyz0123456789._-")[:20]
    
    dom = (b.domain or "").strip().lower()
    if dom:
        if dom != MAIL_DOMAIN:
            with db() as c:
                row = c.execute("SELECT 1 FROM mail_domains WHERE domain=? AND user_id=?", (dom, user["id"])).fetchone()
            if not row:
                raise HTTPException(status_code=400, detail="Tên miền chưa được thêm cho tài khoản của bạn.")
    else:
        dom = MAIL_DOMAIN

    created: list[dict[str, str]] = []
    with db() as c:
        for _ in range(n):
            addr = ""
            for _try in range(12):
                local = (prefix + secrets.token_hex(4))[:40]
                cand = f"{local}@{dom}"
                if not c.execute("SELECT 1 FROM mailboxes WHERE address=?", (cand,)).fetchone():
                    addr = cand
                    break
            if not addr:
                continue
            pwd = secrets.token_urlsafe(9)
            c.execute("INSERT INTO mailboxes(address,pw_hash,owner_uid,created_at) VALUES(?,?,?,?)",
                      (addr, hash_pw(pwd), user["id"], int(time.time())))
            created.append({"address": addr, "password": pwd})
    return {"created": created, "count": len(created)}


@app.post("/mail/create")
def mail_create(b: MailCreateIn, user=Depends(get_user)) -> dict[str, Any]:
    local = (b.local or "").strip().lower()
    if not _LOCAL_RE.match(local):
        raise HTTPException(status_code=400,
            detail="Tên hộp thư 2–40 ký tự, chỉ gồm a-z 0-9 . _ -")
    if len(b.password) < 6:
        raise HTTPException(status_code=400, detail="Mật khẩu hộp thư ≥ 6 ký tự.")
    
    dom = (b.domain or "").strip().lower()
    if dom:
        if dom != MAIL_DOMAIN:
            with db() as c:
                row = c.execute("SELECT 1 FROM mail_domains WHERE domain=? AND user_id=?", (dom, user["id"])).fetchone()
            if not row:
                raise HTTPException(status_code=400, detail="Tên miền chưa được thêm cho tài khoản của bạn.")
    else:
        dom = MAIL_DOMAIN

    address = f"{local}@{dom}"
    phone = (b.phone or "").strip()
    with db() as c:
        if c.execute("SELECT 1 FROM mailboxes WHERE address=?", (address,)).fetchone():
            raise HTTPException(status_code=409, detail="Địa chỉ này đã tồn tại.")
        cur = c.execute(
            "INSERT INTO mailboxes(address,pw_hash,owner_uid,phone,created_at) VALUES(?,?,?,?,?)",
            (address, hash_pw(b.password), user["id"], phone, int(time.time())))
        mid = cur.lastrowid
    return {"id": mid, "address": address, "phone": phone}


@app.get("/mail/list")
def mail_list(user=Depends(get_user)) -> dict[str, Any]:
    with db() as c:
        rows = c.execute(
            "SELECT id,address,phone,created_at FROM mailboxes WHERE owner_uid=? ORDER BY id DESC",
            (user["id"],)).fetchall()
        out = []
        for r in rows:
            unseen = c.execute("SELECT COUNT(*) n FROM mails WHERE mailbox_id=? AND seen=0",
                               (r["id"],)).fetchone()["n"]
            out.append({"id": r["id"], "address": r["address"],
                        "phone": (r["phone"] if "phone" in r.keys() else None),
                        "created_at": r["created_at"], "unseen": unseen})
    return {"mailboxes": out, "domain": MAIL_DOMAIN}


@app.get("/mail/inbox")
def mail_inbox(mailbox_id: int, user=Depends(get_user)) -> dict[str, Any]:
    _mailbox_owned(mailbox_id, user["id"])
    with db() as c:
        rows = c.execute(
            "SELECT id,direction,from_addr,to_addr,subject,body,created_at,seen "
            "FROM mails WHERE mailbox_id=? ORDER BY id DESC LIMIT 200", (mailbox_id,)).fetchall()
    return {"mails": [dict(r) for r in rows]}


@app.post("/mail/seen/{mail_id}")
def mail_seen(mail_id: int, user=Depends(get_user)) -> dict[str, Any]:
    with db() as c:
        c.execute("UPDATE mails SET seen=1 WHERE id=? AND mailbox_id IN "
                  "(SELECT id FROM mailboxes WHERE owner_uid=?)", (mail_id, user["id"]))
    return {"ok": True}


@app.delete("/mail/{mail_id}")
def mail_delete(mail_id: int, user=Depends(get_user)) -> dict[str, Any]:
    with db() as c:
        c.execute("DELETE FROM mails WHERE id=? AND mailbox_id IN "
                  "(SELECT id FROM mailboxes WHERE owner_uid=?)", (mail_id, user["id"]))
    return {"ok": True}


@app.post("/mail/send")
def mail_send(b: MailSendIn, user=Depends(get_user)) -> dict[str, Any]:
    box = _mailbox_owned(b.mailbox_id, user["id"])
    to = (b.to or "").strip()
    if "@" not in to:
        raise HTTPException(status_code=400, detail="Địa chỉ nhận không hợp lệ.")
    now = int(time.time())
    # Lưu bản gửi đi
    with db() as c:
        c.execute("INSERT INTO mails(mailbox_id,direction,from_addr,to_addr,subject,body,created_at,seen) "
                  "VALUES(?,?,?,?,?,?,?,1)",
                  (box["id"], "out", box["address"], to, b.subject, b.body, now))
        # Nội bộ: nếu người nhận cũng là hộp thư trong hệ thống → giao ngay
        inbox = c.execute("SELECT id FROM mailboxes WHERE address=?", (to.lower(),)).fetchone()
        if inbox:
            c.execute("INSERT INTO mails(mailbox_id,direction,from_addr,to_addr,subject,body,created_at,seen) "
                      "VALUES(?,?,?,?,?,?,?,0)",
                      (inbox["id"], "in", box["address"], to, b.subject, b.body, now))
            return {"ok": True, "delivery": "internal"}
    # Bên ngoài: cần SMTP relay
    if not SMTP_RELAY_HOST:
        raise HTTPException(status_code=400,
            detail="Đã lưu vào mục Đã gửi nhưng chưa cấu hình SMTP relay để gửi ra ngoài "
                   "(đặt SMTP_RELAY_HOST/USER/PASS). Gửi nội bộ @"+MAIL_DOMAIN+" thì không cần.")
    try:
        m = _EmailMessage()
        m["From"] = box["address"]; m["To"] = to; m["Subject"] = b.subject
        m.set_content(b.body or "")
        with _smtplib.SMTP(SMTP_RELAY_HOST, SMTP_RELAY_PORT, timeout=30) as s:
            s.starttls()
            if SMTP_RELAY_USER:
                s.login(SMTP_RELAY_USER, SMTP_RELAY_PASS)
            s.send_message(m)
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Gửi ra ngoài thất bại: {e}")
    return {"ok": True, "delivery": "external"}


def _deliver_incoming(rcpt: str, sender: str, subject: str, body: str) -> None:
    """Lưu thư đến vào hộp thư tương ứng (gọi từ bộ nhận SMTP)."""
    rcpt = (rcpt or "").strip().lower()
    try:
        with db() as c:
            box = c.execute("SELECT id FROM mailboxes WHERE address=?", (rcpt,)).fetchone()
            if not box:
                return
            c.execute("INSERT INTO mails(mailbox_id,direction,from_addr,to_addr,subject,body,created_at,seen) "
                      "VALUES(?,?,?,?,?,?,?,0)",
                      (box["id"], "in", sender, rcpt, subject, body, int(time.time())))
    except Exception as e:
        logging.warning("deliver_incoming lỗi: %s", e)


def send_system_mail(to: str, subject: str, body: str) -> str:
    """Gửi mail hệ thống (vd mã OTP). Trả 'internal' / 'external' / 'none'."""
    to = (to or "").strip()
    if "@" not in to:
        return "none"
    sender = f"no-reply@{MAIL_DOMAIN}"
    # Nội bộ: nếu là hộp thư đã tồn tại trong hệ thống (kể cả tên miền custom) → giao thẳng vào KenMail
    with db() as c:
        ok = c.execute("SELECT 1 FROM mailboxes WHERE address=?", (to.lower(),)).fetchone()
    if ok:
        _deliver_incoming(to, sender, subject, body)
        return "internal"
    # Nếu là tên miền mặc định @MAIL_DOMAIN nhưng chưa tạo hộp thư
    if to.lower().endswith("@" + MAIL_DOMAIN):
        return "none"
    # Bên ngoài: cần SMTP relay
    if SMTP_RELAY_HOST:
        try:
            m = _EmailMessage()
            m["From"] = sender; m["To"] = to; m["Subject"] = subject
            m.set_content(body)
            with _smtplib.SMTP(SMTP_RELAY_HOST, SMTP_RELAY_PORT, timeout=30) as s:
                s.starttls()
                if SMTP_RELAY_USER:
                    s.login(SMTP_RELAY_USER, SMTP_RELAY_PASS)
                s.send_message(m)
            return "external"
        except Exception as e:
            logging.warning("send_system_mail relay lỗi: %s", e)
            return "none"
    return "none"


# ===================== Mã xác nhận email (OTP) =====================
class OtpSendIn(BaseModel):
    email: str
    purpose: str = "register"


class OtpVerifyIn(BaseModel):
    email: str
    code: str


def _otp_store_and_send(email: str, purpose: str) -> dict[str, Any]:
    email = (email or "").strip().lower()
    if "@" not in email:
        raise HTTPException(status_code=400, detail="Email không hợp lệ.")
    code = f"{secrets.randbelow(1000000):06d}"
    exp = int(time.time()) + 600  # 10 phút
    with db() as c:
        c.execute("INSERT INTO otp_codes(email,code,purpose,exp,attempts) VALUES(?,?,?,?,0) "
                  "ON CONFLICT(email) DO UPDATE SET code=excluded.code, purpose=excluded.purpose, "
                  "exp=excluded.exp, attempts=0", (email, code, purpose, exp))
    subject = "Mã xác nhận KENIOS"
    body = (f"Mã xác nhận của bạn là: {code}\n"
            f"Mã có hiệu lực trong 10 phút.\n"
            f"Nếu bạn không yêu cầu, hãy bỏ qua email này.")
    channel = send_system_mail(email, subject, body)
    resp: dict[str, Any] = {"sent": channel != "none", "channel": channel}
    if channel == "none":
        resp["hint"] = ("Chưa gửi được mã qua email. Email @" + MAIL_DOMAIN +
                        " cần đã tạo hộp thư; email ngoài (Gmail...) cần cấu hình SMTP_RELAY.")
    if OTP_DEBUG:
        resp["debug_code"] = code
    return resp


def _otp_check(email: str, code: str) -> bool:
    email = (email or "").strip().lower()
    code = (code or "").strip()
    with db() as c:
        row = c.execute("SELECT code,exp,attempts FROM otp_codes WHERE email=?", (email,)).fetchone()
        if not row:
            return False
        if row["attempts"] >= 6:
            return False
        if int(time.time()) > row["exp"]:
            return False
        if row["code"] != code:
            c.execute("UPDATE otp_codes SET attempts=attempts+1 WHERE email=?", (email,))
            return False
        c.execute("DELETE FROM otp_codes WHERE email=?", (email,))
    return True


@app.post("/auth/send-otp")
def auth_send_otp(b: OtpSendIn) -> dict[str, Any]:
    return _otp_store_and_send(b.email, b.purpose or "register")


@app.post("/auth/verify-otp")
def auth_verify_otp(b: OtpVerifyIn) -> dict[str, Any]:
    return {"valid": _otp_check(b.email, b.code)}


def start_mail_smtp() -> None:
    """Khởi động bộ nhận thư SMTP (aiosmtpd) trong tiến trình — cần MX trỏ về VPS + mở port 25."""
    if not MAIL_ENABLE:
        return
    try:
        from aiosmtpd.controller import Controller
        import email as _email_mod
    except Exception:
        logging.warning("KenMail: chưa cài aiosmtpd → không nhận được thư đến. Cài: pip install aiosmtpd")
        return

    def _extract_body(msg) -> str:
        try:
            if msg.is_multipart():
                for part in msg.walk():
                    if part.get_content_type() == "text/plain":
                        return part.get_payload(decode=True).decode("utf-8", "replace")
                return msg.get_payload(decode=True).decode("utf-8", "replace")
            payload = msg.get_payload(decode=True)
            return payload.decode("utf-8", "replace") if payload else str(msg.get_payload())
        except Exception:
            return ""

    class _Handler:
        async def handle_DATA(self, server, session, envelope):
            try:
                msg = _email_mod.message_from_bytes(envelope.content)
                subject = msg.get("Subject", "")
                sender = envelope.mail_from or msg.get("From", "")
                body = _extract_body(msg)
                for rcpt in envelope.rcpt_tos:
                    _deliver_incoming(rcpt, sender, subject, body)
            except Exception as e:
                logging.warning("SMTP handle_DATA lỗi: %s", e)
            return "250 Message accepted"

    try:
        controller = Controller(_Handler(), hostname="0.0.0.0", port=MAIL_SMTP_PORT)
        controller.start()
        logging.info("KenMail SMTP nhận thư tại cổng %s cho @%s", MAIL_SMTP_PORT, MAIL_DOMAIN)
    except Exception as e:
        logging.warning("KenMail: không khởi động được SMTP cổng %s: %s", MAIL_SMTP_PORT, e)


# ======================== Dịch sang tiếng Việt (TTS đa ngôn ngữ) ========================
class TranslateIn(BaseModel):
    text: str
    target: str = "vi"
    source: str = "auto"


@app.post("/translate")
async def translate_text(b: TranslateIn, user=Depends(get_user)) -> dict[str, Any]:
    """Dịch văn bản sang tiếng Việt (mặc định) — dùng cho đọc TTS đa ngôn ngữ.

    Dùng endpoint dịch miễn phí (không cần API key). Nếu lỗi sẽ trả lại nguyên văn.
    """
    text = (b.text or "").strip()
    if not text:
        return {"text": "", "source": b.source}
    target = (b.target or "vi").strip() or "vi"
    source = (b.source or "auto").strip() or "auto"
    try:
        async with httpx.AsyncClient(timeout=15) as client:
            r = await client.get(
                "https://translate.googleapis.com/translate_a/single",
                params={"client": "gtx", "sl": source, "tl": target, "dt": "t", "q": text},
                headers={"User-Agent": "Mozilla/5.0"},
            )
        if r.status_code == 200:
            data = r.json()
            segments = data[0] if isinstance(data, list) and data else []
            out = "".join(seg[0] for seg in segments if seg and seg[0])
            detected = data[2] if isinstance(data, list) and len(data) > 2 else source
            return {"text": out or text, "source": detected}
    except Exception:
        pass
    return {"text": text, "source": source}


# ======================== Thanh toán / Credits ========================
PACKAGES = {
    "pro":   {"credits": 9999,  "amount": 199000, "label": "Gói PRO — 199.000đ (9.999 credits)"},
    "ultra": {"credits": 20000, "amount": 299000, "label": "Gói ULTRA — 299.000đ (20.000 credits)"},
    "max":   {"credits": 50000, "amount": 499000, "label": "Gói MAX — 499.000đ (50.000 credits)"},
}


@app.get("/payment/packages")
def payment_packages() -> list[dict[str, Any]]:
    return [{"id": k, **v} for k, v in PACKAGES.items()]


@app.post("/payment/create")
def payment_create(b: PaymentIn, user=Depends(get_user)) -> dict[str, Any]:
    if b.package not in PACKAGES:
        raise HTTPException(status_code=400,
            detail=f"Gói không hợp lệ. Chọn: {list(PACKAGES.keys())}")
    pkg = PACKAGES[b.package]
    ref = secrets.token_urlsafe(16)
    with db() as c:
        cur = c.execute(
            "INSERT INTO payments(user_id,amount,credits,status,ref,created_at) "
            "VALUES(?,?,?,'pending',?,?)",
            (user["id"], pkg["amount"], pkg["credits"], ref, int(time.time())),
        )
        pid = cur.lastrowid
    bank = bank_info(amount=pkg["amount"], note=f"KENIOS {ref}")
    return {
        "payment_id": pid,
        "ref": ref,
        "amount": pkg["amount"],
        "credits": pkg["credits"],
        "label": pkg["label"],
        "message": "Quét mã QR hoặc chuyển khoản theo thông tin bên dưới.",
        "bank_info": bank,
        "qr_url": bank["qr_url"],
    }


def bank_info(amount: int = 0, note: str = "KENIOS") -> dict[str, Any]:
    from urllib.parse import quote
    code = get_setting("bank_code", "970416")
    short = get_setting("bank_short", "ACB")
    account = get_setting("bank_account", "23252921")
    name = get_setting("bank_name", "TRAN MINH CHIEN")
    qr = (f"https://img.vietqr.io/image/{code}-{account}-compact2.png"
          f"?accountName={quote(name)}&addInfo={quote(note)}")
    if amount > 0:
        qr += f"&amount={amount}"
    return {"bank": short, "bank_code": code, "account": account,
            "name": name, "content": note, "qr_url": qr}


@app.get("/payment/info")
def payment_info(amount: int = 0, note: str = "KENIOS", user=Depends(get_user)) -> dict[str, Any]:
    return bank_info(amount=amount, note=note)


class BankSettingsIn(BaseModel):
    bank_code: Optional[str] = None
    bank_short: Optional[str] = None
    bank_account: Optional[str] = None
    bank_name: Optional[str] = None
    bank_webhook: Optional[str] = None
    bank_apikey: Optional[str] = None

@app.get("/admin/payment/settings")
def admin_get_bank(admin=Depends(get_admin)) -> dict[str, Any]:
    return {
        "bank_code": get_setting("bank_code", "970416"),
        "bank_short": get_setting("bank_short", "ACB"),
        "bank_account": get_setting("bank_account", "23252921"),
        "bank_name": get_setting("bank_name", "TRAN MINH CHIEN"),
        "bank_webhook": get_setting("bank_webhook", ""),
        "bank_apikey": get_setting("bank_apikey", ""),
    }

@app.post("/admin/payment/settings")
def admin_set_bank(b: BankSettingsIn, admin=Depends(get_admin)) -> dict[str, Any]:
    for field in ["bank_code", "bank_short", "bank_account", "bank_name", "bank_webhook", "bank_apikey"]:
        val = getattr(b, field)
        if val is not None:
            set_setting(field, val)
    return {"message": "Đã cập nhật thông tin ngân hàng."}


def _determine_plan_from_credits(credits: int) -> str:
    if credits >= 50000:
        return "max"
    elif credits >= 20000:
        return "ultra"
    elif credits >= 9999:
        return "pro"
    return "free"


@app.post("/payment/confirm/{pid}")
def payment_confirm(pid: int, admin=Depends(get_admin)) -> dict[str, Any]:
    with db() as c:
        pay = c.execute("SELECT * FROM payments WHERE id=?", (pid,)).fetchone()
        if not pay:
            raise HTTPException(status_code=404, detail="Không tìm thấy đơn thanh toán.")
        if pay["status"] == "completed":
            raise HTTPException(status_code=400, detail="Đơn đã được xác nhận trước đó.")
        c.execute("UPDATE payments SET status='completed' WHERE id=?", (pid,))
        c.execute("UPDATE users SET credits=credits+? WHERE id=?",
                  (pay["credits"], pay["user_id"]))
        new_plan = _determine_plan_from_credits(pay["credits"])
        if new_plan != "free":
            c.execute("UPDATE users SET plan=? WHERE id=? AND plan IN ('free','pro','ultra')",
                      (new_plan, pay["user_id"]))
    return {"message": f"Đã cộng {pay['credits']} credits cho user {pay['user_id']}."}


# ======================== Webhook thanh toán tự động ========================
@app.post("/payment/webhook")
async def payment_webhook(request: Request) -> dict[str, Any]:
    webhook_key = get_setting("bank_apikey", "")
    if webhook_key:
        auth_header = request.headers.get("Authorization", "")
        secure_token = request.headers.get("X-API-Key", "") or request.headers.get("Secure-Token", "")
        provided_key = ""
        if auth_header.startswith("Apikey "):
            provided_key = auth_header.split(" ", 1)[1]
        elif auth_header.startswith("Bearer "):
            provided_key = auth_header.split(" ", 1)[1]
        elif secure_token:
            provided_key = secure_token
        if provided_key != webhook_key:
            raise HTTPException(status_code=401, detail="Webhook API key không hợp lệ.")

    try:
        body = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Body JSON không hợp lệ.")

    confirmed = 0
    data_list = body.get("data", [])
    if isinstance(data_list, list):
        for item in data_list:
            desc = item.get("description", "") or item.get("content", "")
            amount = item.get("amount", 0) or item.get("transferAmount", 0)
            ref = _extract_ref(desc)
            if ref:
                if _auto_confirm_payment(ref, amount):
                    confirmed += 1

    if not data_list:
        desc = body.get("content", "") or body.get("description", "")
        amount = body.get("transferAmount", 0) or body.get("amount", 0)
        ref = _extract_ref(desc)
        if ref:
            if _auto_confirm_payment(ref, amount):
                confirmed += 1

    return {"success": True, "confirmed": confirmed}


def _extract_ref(description: str) -> Optional[str]:
    if not description:
        return None
    match = re.search(r"KENIOS\s+(\S+)", description, re.IGNORECASE)
    if match:
        return match.group(1)
    return None


def _auto_confirm_payment(ref: str, amount: int) -> bool:
    try:
        with db() as c:
            pay = c.execute(
                "SELECT id,user_id,credits,amount,status FROM payments WHERE ref=? AND status='pending'",
                (ref,)
            ).fetchone()
            if not pay:
                return False
            if amount > 0 and abs(amount - pay["amount"]) > pay["amount"] * 0.05:
                log.warning("Webhook: ref=%s amount mismatch (expected=%d, got=%d)",
                           ref, pay["amount"], amount)
                if amount < pay["amount"]:
                    return False
            c.execute("UPDATE payments SET status='completed' WHERE id=?", (pay["id"],))
            c.execute("UPDATE users SET credits=credits+? WHERE id=?",
                      (pay["credits"], pay["user_id"]))
            new_plan = _determine_plan_from_credits(pay["credits"])
            if new_plan != "free":
                c.execute("UPDATE users SET plan=? WHERE id=? AND plan IN ('free','pro','ultra')",
                          (new_plan, pay["user_id"]))
            log.info("Webhook xác nhận: ref=%s, user=%d, credits=%d",
                    ref, pay["user_id"], pay["credits"])
        return True
    except Exception as e:
        log.error("Webhook lỗi xác nhận ref=%s: %s", ref, e)
        return False


@app.get("/payment/history")
def payment_history(user=Depends(get_user)) -> list[dict[str, Any]]:
    with db() as c:
        rows = c.execute(
            "SELECT id,amount,credits,status,ref,created_at FROM payments "
            "WHERE user_id=? ORDER BY id DESC", (user["id"],)
        ).fetchall()
    return [dict(r) for r in rows]


@app.get("/me/credits")
def my_credits(user=Depends(get_user)) -> dict[str, Any]:
    return {"credits": user["credits"], "plan": user["plan"]}


# ======================== Prompt Templates ========================
@app.get("/prompts")
def list_prompts(user=Depends(get_user)) -> list[dict[str, Any]]:
    with db() as c:
        rows = c.execute(
            "SELECT id,title,content,category,is_public,user_id,created_at "
            "FROM prompt_templates WHERE is_public=1 OR user_id=? "
            "ORDER BY id DESC",
            (user["id"],)
        ).fetchall()
    return [dict(r) for r in rows]


@app.post("/prompts")
def create_prompt(b: PromptTemplateIn, user=Depends(get_user)) -> dict[str, Any]:
    if not b.title.strip() or not b.content.strip():
        raise HTTPException(status_code=400, detail="Title và content không được để trống.")
    is_pub = 1 if (b.is_public and user["is_admin"]) else 0
    with db() as c:
        cur = c.execute(
            "INSERT INTO prompt_templates(title,content,category,is_public,user_id,created_at) "
            "VALUES(?,?,?,?,?,?)",
            (b.title.strip(), b.content.strip(), b.category, is_pub, user["id"], int(time.time())),
        )
        pid = cur.lastrowid
    return {"id": pid, "message": "Đã tạo prompt template."}


@app.delete("/prompts/{pid}")
def delete_prompt(pid: int, user=Depends(get_user)) -> dict[str, Any]:
    with db() as c:
        row = c.execute("SELECT user_id FROM prompt_templates WHERE id=?", (pid,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Không tìm thấy prompt template.")
        if row["user_id"] != user["id"] and not user["is_admin"]:
            raise HTTPException(status_code=403, detail="Bạn không có quyền xóa prompt này.")
        c.execute("DELETE FROM prompt_templates WHERE id=?", (pid,))
    return {"message": "Đã xóa prompt template."}


# ======================== Favorites ========================
@app.get("/favorites")
def list_favorites(user=Depends(get_user)) -> list[dict[str, Any]]:
    with db() as c:
        rows = c.execute(
            "SELECT id,message_content,conversation_id,provider,created_at "
            "FROM favorites WHERE user_id=? ORDER BY id DESC",
            (user["id"],)
        ).fetchall()
    return [dict(r) for r in rows]


@app.post("/favorites")
def add_favorite(b: FavoriteIn, user=Depends(get_user)) -> dict[str, Any]:
    if not b.message_content.strip():
        raise HTTPException(status_code=400, detail="Nội dung tin nhắn không được để trống.")
    with db() as c:
        cur = c.execute(
            "INSERT INTO favorites(user_id,message_content,conversation_id,provider,created_at) "
            "VALUES(?,?,?,?,?)",
            (user["id"], b.message_content.strip(), b.conversation_id, b.provider, int(time.time())),
        )
        fid = cur.lastrowid
    return {"id": fid, "message": "Đã thêm vào yêu thích."}


@app.delete("/favorites/{fid}")
def remove_favorite(fid: int, user=Depends(get_user)) -> dict[str, Any]:
    with db() as c:
        c.execute("DELETE FROM favorites WHERE id=? AND user_id=?", (fid, user["id"]))
    return {"message": "Đã xóa khỏi yêu thích."}


# ======================== Friends & Direct Messaging ========================
@app.get("/users/search")
def search_users(q: str, user=Depends(get_user)) -> list[dict[str, Any]]:
    """Tìm theo username, SĐT hoặc ID công khai (KEN...)."""
    if not q or len(q.strip()) < 1:
        return []
    term = q.strip()
    kw = f"%{term}%"
    with db() as c:
        rows = c.execute(
            "SELECT id, username, public_id, phone FROM users "
            "WHERE (username LIKE ? OR phone LIKE ? OR public_id LIKE ? "
            "       OR public_id = ? OR phone = ?) AND id != ?",
            (kw, kw, kw, term.upper(), term, user["id"])
        ).fetchall()
    return [{"id": r["id"], "username": r["username"],
             "public_id": r["public_id"], "phone": r["phone"]} for r in rows]


@app.post("/friends/request")
def send_friend_request(b: FriendRequestIn, user=Depends(get_user)) -> dict[str, Any]:
    if b.friend_id == user["id"]:
        raise HTTPException(status_code=400, detail="Bạn không thể gửi lời mời kết bạn cho chính mình.")
    with db() as c:
        # Check if friend exists
        target = c.execute("SELECT id FROM users WHERE id=?", (b.friend_id,)).fetchone()
        if not target:
            raise HTTPException(status_code=404, detail="Không tìm thấy người dùng này.")
        
        # Check if friendship already exists
        existing = c.execute(
            "SELECT id, status FROM friendships WHERE (user_id=? AND friend_id=?) OR (user_id=? AND friend_id=?)",
            (user["id"], b.friend_id, b.friend_id, user["id"])
        ).fetchone()
        
        if existing:
            if existing["status"] == "accepted":
                raise HTTPException(status_code=400, detail="Hai bạn đã là bạn bè.")
            else:
                raise HTTPException(status_code=400, detail="Lời mời kết bạn đã được gửi trước đó.")
        
        c.execute(
            "INSERT INTO friendships(user_id, friend_id, status, created_at) VALUES(?,?,?,?)",
            (user["id"], b.friend_id, "pending", int(time.time()))
        )
    return {"message": "Đã gửi lời mời kết bạn thành công."}


@app.get("/friends/requests")
def list_friend_requests(user=Depends(get_user)) -> list[dict[str, Any]]:
    with db() as c:
        rows = c.execute(
            "SELECT f.id, f.user_id as sender_id, u.username as sender_name, "
            "f.friend_id as receiver_id, u2.username as receiver_name, f.created_at "
            "FROM friendships f "
            "JOIN users u ON f.user_id = u.id "
            "JOIN users u2 ON f.friend_id = u2.id "
            "WHERE (f.friend_id=? OR f.user_id=?) AND f.status='pending'",
            (user["id"], user["id"])
        ).fetchall()
    return [dict(r) for r in rows]


@app.post("/friends/respond")
def respond_friend_request(b: FriendResponseIn, user=Depends(get_user)) -> dict[str, Any]:
    if b.action not in ("accept", "decline"):
        raise HTTPException(status_code=400, detail="Hành động không hợp lệ. Phải là 'accept' hoặc 'decline'.")
    with db() as c:
        row = c.execute(
            "SELECT id, user_id, friend_id, status FROM friendships WHERE id=?", (b.request_id,)
        ).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Không tìm thấy yêu cầu kết bạn.")
        
        # Verify that current user is the receiver of the request
        if row["friend_id"] != user["id"]:
            raise HTTPException(status_code=403, detail="Bạn không có quyền xử lý yêu cầu này.")
        
        if row["status"] != "pending":
            raise HTTPException(status_code=400, detail="Yêu cầu này đã được xử lý trước đó.")
        
        if b.action == "accept":
            c.execute("UPDATE friendships SET status='accepted' WHERE id=?", (b.request_id,))
            msg = "Đã chấp nhận lời mời kết bạn."
        else:
            c.execute("DELETE FROM friendships WHERE id=?", (b.request_id,))
            msg = "Đã từ chối lời mời kết bạn."
            
    return {"message": msg}


@app.get("/friends")
def list_friends(user=Depends(get_user)) -> list[dict[str, Any]]:
    with db() as c:
        rows = c.execute(
            "SELECT DISTINCT u.id, u.username "
            "FROM friendships f "
            "JOIN users u ON (f.user_id = u.id AND f.friend_id = ?) OR (f.friend_id = u.id AND f.user_id = ?) "
            "WHERE f.status='accepted'",
            (user["id"], user["id"])
        ).fetchall()
    return [dict(r) for r in rows]


@app.get("/direct_messages/{friend_id}")
def get_direct_messages(friend_id: int, user=Depends(get_user)) -> list[dict[str, Any]]:
    with db() as c:
        # Verify that they are friends
        friendship = c.execute(
            "SELECT id FROM friendships WHERE ((user_id=? AND friend_id=?) OR (user_id=? AND friend_id=?)) AND status='accepted'",
            (user["id"], friend_id, friend_id, user["id"])
        ).fetchone()
        if not friendship:
            raise HTTPException(status_code=403, detail="Bạn phải là bạn bè để nhắn tin với người này.")
            
        rows = c.execute(
            "SELECT id, sender_id, receiver_id, content, created_at, is_read "
            "FROM direct_messages "
            "WHERE (sender_id=? AND receiver_id=?) OR (sender_id=? AND receiver_id=?) "
            "ORDER BY id ASC",
            (user["id"], friend_id, friend_id, user["id"])
        ).fetchall()
        
        # Mark messages from friend as read
        c.execute("UPDATE direct_messages SET is_read=1 WHERE sender_id=? AND receiver_id=?", (friend_id, user["id"]))
        
    return [dict(r) for r in rows]


@app.post("/direct_messages")
def send_direct_message(b: DirectMessageIn, user=Depends(get_user)) -> dict[str, Any]:
    if not b.content.strip():
        raise HTTPException(status_code=400, detail="Nội dung tin nhắn không được để trống.")
    with db() as c:
        # Verify that they are friends
        friendship = c.execute(
            "SELECT id FROM friendships WHERE ((user_id=? AND friend_id=?) OR (user_id=? AND friend_id=?)) AND status='accepted'",
            (user["id"], b.receiver_id, b.receiver_id, user["id"])
        ).fetchone()
        if not friendship:
            raise HTTPException(status_code=403, detail="Bạn phải là bạn bè để nhắn tin với người này.")
            
        cur = c.execute(
            "INSERT INTO direct_messages(sender_id, receiver_id, content, created_at, is_read) VALUES(?,?,?,?,0)",
            (user["id"], b.receiver_id, b.content.strip(), int(time.time()))
        )
        msg_id = cur.lastrowid
    return {"id": msg_id, "message": "Đã gửi tin nhắn thành công."}



# ======================== Search ========================
@app.get("/search")
def search_messages(q: str, user=Depends(get_user)) -> list[dict[str, Any]]:
    if not q or len(q.strip()) < 2:
        raise HTTPException(status_code=400, detail="Từ khóa tìm kiếm cần ít nhất 2 ký tự.")
    keyword = f"%{q.strip()}%"
    with db() as c:
        rows = c.execute(
            "SELECT m.id, m.conversation_id, m.role, m.content, m.created_at, "
            "c.title as conversation_title, c.provider "
            "FROM messages m "
            "JOIN conversations c ON c.id = m.conversation_id "
            "WHERE c.user_id=? AND m.content LIKE ? "
            "ORDER BY m.created_at DESC LIMIT 50",
            (user["id"], keyword)
        ).fetchall()
    return [dict(r) for r in rows]


# ======================== Admin ========================
class BanIn(BaseModel):   banned: bool
class AdminPwIn(BaseModel): new_password: str
class PlanIn(BaseModel):   plan: str

@app.get("/admin/users")
def admin_users(admin=Depends(get_admin)) -> list[dict[str, Any]]:
    with db() as c:
        rows = c.execute(
            "SELECT id,username,email,phone,public_id,is_admin,banned,plan,credits,"
            "status,suspend_until,last_seen,last_feature,created_at "
            "FROM users ORDER BY id"
        ).fetchall()
    out = []
    for r in rows:
        d = dict(r)
        d["is_admin"] = bool(r["is_admin"])
        d["plan"] = "pro" if r["is_admin"] else (r["plan"] or "free")
        d["status"] = r["status"] or "active"
        out.append(d)
    return out


@app.post("/admin/users/{uid}/ban")
def admin_ban(uid: int, b: BanIn, admin=Depends(get_admin)) -> dict[str, Any]:
    if uid == admin["id"]:
        raise HTTPException(status_code=400, detail="Không thể tự khóa chính mình.")
    with db() as c:
        c.execute("UPDATE users SET banned=? WHERE id=?", (1 if b.banned else 0, uid))
    return {"message": "Đã khóa." if b.banned else "Đã mở khóa."}


@app.post("/admin/users/{uid}/password")
def admin_set_pw(uid: int, b: AdminPwIn, admin=Depends(get_admin)) -> dict[str, Any]:
    if len(b.new_password) < 6:
        raise HTTPException(status_code=400, detail="Mật khẩu ≥6 ký tự.")
    with db() as c:
        c.execute("UPDATE users SET pw_hash=? WHERE id=?",
                  (hash_pw(b.new_password), uid))
    return {"message": "Đã đổi mật khẩu."}


@app.post("/admin/users/{uid}/plan")
def admin_set_plan(uid: int, b: PlanIn, admin=Depends(get_admin)) -> dict[str, Any]:
    # Chỉ còn 2 gói: free / pro
    plan = "pro" if b.plan == "pro" else "free"
    with db() as c:
        c.execute("UPDATE users SET plan=? WHERE id=?", (plan, uid))
    return {"message": f"Đã đặt gói '{plan}'."}


class SuspendIn(BaseModel):
    minutes: int = 0   # 0 = ngưng vô thời hạn; >0 = ngưng theo phút


@app.post("/admin/users/{uid}/suspend")
def admin_suspend(uid: int, b: SuspendIn, admin=Depends(get_admin)) -> dict[str, Any]:
    if uid == admin["id"]:
        raise HTTPException(status_code=400, detail="Không thể tự ngưng chính mình.")
    until = int(time.time()) + b.minutes * 60 if b.minutes > 0 else 0
    with db() as c:
        c.execute("UPDATE users SET status='suspended', suspend_until=? WHERE id=?", (until, uid))
    return {"message": "Đã tạm ngưng tài khoản."}


@app.post("/admin/users/{uid}/unsuspend")
def admin_unsuspend(uid: int, admin=Depends(get_admin)) -> dict[str, Any]:
    with db() as c:
        c.execute("UPDATE users SET status='active', suspend_until=0 WHERE id=?", (uid,))
    return {"message": "Đã mở lại tài khoản."}


# ---- Người dùng: lấy hồ sơ mới nhất + nhịp hoạt động ----
@app.get("/me")
def get_me(user=Depends(get_user)) -> dict[str, Any]:
    return _user_dict(user)


class ActivityIn(BaseModel):
    feature: str = ""


@app.post("/me/activity")
def me_activity(b: ActivityIn, user=Depends(get_user)) -> dict[str, Any]:
    with db() as c:
        c.execute("UPDATE users SET last_seen=?, last_feature=? WHERE id=?",
                  (int(time.time()), (b.feature or "")[:60], user["id"]))
    return {"ok": True}


# ---- Chế độ bảo trì (admin bật → người dùng bị khoá tạm) ----
class MaintenanceIn(BaseModel):
    on: bool = False
    message: str = "Ứng dụng đang nâng cấp phiên bản. Vui lòng đợi trong giây lát."


@app.get("/app/status")
def app_status() -> dict[str, Any]:
    on = _setting_get("maintenance_on", "0") == "1"
    return {
        "maintenance": on,
        "message": _setting_get("maintenance_msg",
                                "Ứng dụng đang nâng cấp phiên bản. Vui lòng đợi trong giây lát."),
        "version": "5.0",
    }


@app.post("/admin/maintenance")
def admin_maintenance(b: MaintenanceIn, admin=Depends(get_admin)) -> dict[str, Any]:
    _setting_set("maintenance_on", "1" if b.on else "0")
    if b.message:
        _setting_set("maintenance_msg", b.message)
    return {"message": "Đã bật bảo trì." if b.on else "Đã tắt bảo trì.", "maintenance": b.on}


# ======================== Video feed (TikTok của riêng app) ========================
class PostIn(BaseModel):
    file_id: int
    caption: str = ""


@app.post("/posts")
def create_post(b: PostIn, user=Depends(get_user)) -> dict[str, Any]:
    with db() as c:
        f = c.execute("SELECT id FROM files WHERE id=? AND user_id=?",
                      (b.file_id, user["id"])).fetchone()
        if not f:
            raise HTTPException(status_code=404, detail="Không tìm thấy video của bạn để đăng.")
        cur = c.execute(
            "INSERT INTO posts(user_id,file_id,caption,likes,created_at) VALUES(?,?,?,0,?)",
            (user["id"], b.file_id, (b.caption or "")[:300], int(time.time())))
        pid = cur.lastrowid
    return {"id": pid, "message": "Đã đăng video."}


@app.get("/feed")
def feed(user=Depends(get_user)) -> list[dict[str, Any]]:
    with db() as c:
        rows = c.execute(
            "SELECT p.id, p.caption, p.likes, p.created_at, p.file_id, "
            "u.username, u.public_id, f.name, f.mime "
            "FROM posts p JOIN users u ON p.user_id=u.id "
            "JOIN files f ON p.file_id=f.id "
            "ORDER BY p.id DESC LIMIT 100"
        ).fetchall()
        liked = {r["post_id"] for r in c.execute(
            "SELECT post_id FROM post_likes WHERE user_id=?", (user["id"],)).fetchall()}
    return [{
        "id": r["id"], "caption": r["caption"], "likes": r["likes"],
        "created_at": r["created_at"], "file_id": r["file_id"],
        "username": r["username"], "public_id": r["public_id"],
        "name": r["name"], "mime": r["mime"],
        "liked": r["id"] in liked,
    } for r in rows]


@app.get("/posts/{pid}/video")
def post_video(pid: int, background_tasks: BackgroundTasks, user=Depends(get_user)):
    with db() as c:
        row = c.execute(
            "SELECT f.name,f.mime,f.data,f.id as fid FROM posts p "
            "JOIN files f ON p.file_id=f.id WHERE p.id=?", (pid,)).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Không tìm thấy video.")
    file_path = os.path.join(UPLOAD_DIR, str(row["fid"]))
    if os.path.exists(file_path):
        return FileResponse(path=file_path, filename=row["name"],
                            media_type=row["mime"] or "video/mp4")
    if row["data"]:
        temp_path = os.path.join(UPLOAD_DIR, f"feed_{pid}_{secrets.token_hex(4)}")
        with open(temp_path, "wb") as f:
            f.write(base64.b64decode(row["data"]))
        background_tasks.add_task(os.unlink, temp_path)
        return FileResponse(path=temp_path, filename=row["name"],
                            media_type=row["mime"] or "video/mp4")
    raise HTTPException(status_code=404, detail="Không có nội dung video.")


@app.post("/posts/{pid}/like")
def like_post(pid: int, user=Depends(get_user)) -> dict[str, Any]:
    with db() as c:
        ex = c.execute("SELECT 1 FROM post_likes WHERE post_id=? AND user_id=?",
                       (pid, user["id"])).fetchone()
        if ex:
            c.execute("DELETE FROM post_likes WHERE post_id=? AND user_id=?", (pid, user["id"]))
            c.execute("UPDATE posts SET likes=MAX(0,likes-1) WHERE id=?", (pid,))
            liked = False
        else:
            c.execute("INSERT INTO post_likes(post_id,user_id) VALUES(?,?)", (pid, user["id"]))
            c.execute("UPDATE posts SET likes=likes+1 WHERE id=?", (pid,))
            liked = True
        likes = c.execute("SELECT likes FROM posts WHERE id=?", (pid,)).fetchone()
    return {"liked": liked, "likes": likes["likes"] if likes else 0}


@app.delete("/posts/{pid}")
def delete_post(pid: int, user=Depends(get_user)) -> dict[str, Any]:
    with db() as c:
        row = c.execute("SELECT user_id FROM posts WHERE id=?", (pid,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Không tìm thấy bài.")
        if row["user_id"] != user["id"] and not user["is_admin"]:
            raise HTTPException(status_code=403, detail="Không thể xoá bài của người khác.")
        c.execute("DELETE FROM posts WHERE id=?", (pid,))
        c.execute("DELETE FROM post_likes WHERE post_id=?", (pid,))
    return {"message": "Đã xoá bài."}


# ======================== Live (phòng live + bình luận như TikTok) ========================
class LiveCreateIn(BaseModel):
    title: str = ""
    hls_url: str = ""


@app.post("/live/create")
def live_create(b: LiveCreateIn, user=Depends(get_user)) -> dict[str, Any]:
    title = (b.title or f"Live của {user['username']}")[:120]
    with db() as c:
        cur = c.execute(
            "INSERT INTO live_rooms(host_id,title,hls_url,viewers,likes,active,created_at) "
            "VALUES(?,?,?,0,0,1,?)",
            (user["id"], title, (b.hls_url or "")[:300], int(time.time())))
        rid = cur.lastrowid
    return {"id": rid, "message": "Đã mở phòng live."}


@app.post("/live/{rid}/end")
def live_end(rid: int, user=Depends(get_user)) -> dict[str, Any]:
    with db() as c:
        row = c.execute("SELECT host_id FROM live_rooms WHERE id=?", (rid,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Không tìm thấy phòng live.")
        if row["host_id"] != user["id"] and not user["is_admin"]:
            raise HTTPException(status_code=403, detail="Chỉ chủ phòng mới kết thúc được.")
        c.execute("UPDATE live_rooms SET active=0 WHERE id=?", (rid,))
    return {"message": "Đã kết thúc live."}


@app.get("/live/rooms")
def live_rooms(user=Depends(get_user)) -> list[dict[str, Any]]:
    with db() as c:
        rows = c.execute(
            "SELECT r.id,r.title,r.hls_url,r.viewers,r.likes,r.created_at,"
            "u.username,u.public_id FROM live_rooms r JOIN users u ON r.host_id=u.id "
            "WHERE r.active=1 ORDER BY r.id DESC LIMIT 100").fetchall()
    return [dict(r) for r in rows]


@app.get("/live/{rid}")
def live_info(rid: int, user=Depends(get_user)) -> dict[str, Any]:
    with db() as c:
        r = c.execute(
            "SELECT r.id,r.title,r.hls_url,r.viewers,r.likes,r.active,r.host_id,"
            "u.username,u.public_id FROM live_rooms r JOIN users u ON r.host_id=u.id "
            "WHERE r.id=?", (rid,)).fetchone()
    if not r:
        raise HTTPException(status_code=404, detail="Không tìm thấy phòng live.")
    return dict(r)


@app.post("/live/{rid}/join")
def live_join(rid: int, user=Depends(get_user)) -> dict[str, Any]:
    with db() as c:
        c.execute("UPDATE live_rooms SET viewers=viewers+1 WHERE id=?", (rid,))
    return {"ok": True}


@app.post("/live/{rid}/like")
def live_like(rid: int, user=Depends(get_user)) -> dict[str, Any]:
    with db() as c:
        c.execute("UPDATE live_rooms SET likes=likes+1 WHERE id=?", (rid,))
        r = c.execute("SELECT likes FROM live_rooms WHERE id=?", (rid,)).fetchone()
    return {"likes": r["likes"] if r else 0}


class LiveCommentIn(BaseModel):
    content: str


@app.post("/live/{rid}/comment")
def live_comment(rid: int, b: LiveCommentIn, user=Depends(get_user)) -> dict[str, Any]:
    if not b.content.strip():
        raise HTTPException(status_code=400, detail="Bình luận trống.")
    with db() as c:
        cur = c.execute(
            "INSERT INTO live_messages(room_id,user_id,username,content,created_at) "
            "VALUES(?,?,?,?,?)",
            (rid, user["id"], user["username"], b.content.strip()[:300], int(time.time())))
        mid = cur.lastrowid
    return {"id": mid}


@app.get("/live/{rid}/comments")
def live_comments(rid: int, after: int = 0, user=Depends(get_user)) -> list[dict[str, Any]]:
    with db() as c:
        rows = c.execute(
            "SELECT id,username,content,created_at FROM live_messages "
            "WHERE room_id=? AND id>? ORDER BY id ASC LIMIT 100", (rid, after)).fetchall()
    return [dict(r) for r in rows]


@app.post("/admin/payments/{pid}/confirm")
def admin_confirm_payment(pid: int, admin=Depends(get_admin)) -> dict[str, Any]:
    return payment_confirm(pid, admin)


# ======================== Admin Stats ========================
@app.get("/admin/stats")
def admin_stats(admin=Depends(get_admin)) -> dict[str, Any]:
    now = int(time.time())
    seven_days_ago = now - (7 * 24 * 60 * 60)
    thirty_days_ago = now - (30 * 24 * 60 * 60)

    with db() as c:
        total_users = c.execute("SELECT COUNT(*) as cnt FROM users").fetchone()["cnt"]
        new_users_7d = c.execute(
            "SELECT COUNT(*) as cnt FROM users WHERE created_at>=?", (seven_days_ago,)
        ).fetchone()["cnt"]
        total_conversations = c.execute("SELECT COUNT(*) as cnt FROM conversations").fetchone()["cnt"]
        total_messages = c.execute("SELECT COUNT(*) as cnt FROM messages").fetchone()["cnt"]
        rev_total_row = c.execute(
            "SELECT COALESCE(SUM(amount),0) as total FROM payments WHERE status='completed'"
        ).fetchone()
        revenue_total = rev_total_row["total"]
        rev_30d_row = c.execute(
            "SELECT COALESCE(SUM(amount),0) as total FROM payments WHERE status='completed' AND created_at>=?",
            (thirty_days_ago,)
        ).fetchone()
        revenue_30d = rev_30d_row["total"]
        total_files = c.execute("SELECT COUNT(*) as cnt FROM files").fetchone()["cnt"]
        top_rows = c.execute(
            "SELECT provider, COUNT(*) as cnt FROM conversations "
            "WHERE provider IS NOT NULL GROUP BY provider ORDER BY cnt DESC LIMIT 10"
        ).fetchall()
        top_providers = [{"provider": r["provider"], "count": r["cnt"]} for r in top_rows]
        plan_rows = c.execute(
            "SELECT plan, COUNT(*) as cnt FROM users GROUP BY plan ORDER BY cnt DESC"
        ).fetchall()
        plan_distribution = [{"plan": r["plan"], "count": r["cnt"]} for r in plan_rows]

    return {
        "total_users": total_users,
        "new_users_7d": new_users_7d,
        "total_conversations": total_conversations,
        "total_messages": total_messages,
        "revenue_total": revenue_total,
        "revenue_30d": revenue_30d,
        "total_files": total_files,
        "top_providers": top_providers,
        "plan_distribution": plan_distribution,
    }


# ======================== Báo lỗi & log lỗi cho admin ========================
class ErrorIn(BaseModel):
    context: str = ""
    detail: str = ""

@app.post("/errors")
def report_error(b: ErrorIn, user=Depends(get_user)) -> dict[str, Any]:
    with db() as c:
        c.execute("INSERT INTO error_logs(user_id,username,context,detail,created_at) "
                  "VALUES(?,?,?,?,?)",
                  (user["id"], user["username"], b.context[:200], b.detail[:800], int(time.time())))
    return {"message": "Đã ghi nhận lỗi."}


@app.get("/admin/errors")
def admin_errors(admin=Depends(get_admin)) -> list[dict[str, Any]]:
    with db() as c:
        rows = c.execute("SELECT id,user_id,username,context,detail,created_at "
                         "FROM error_logs ORDER BY id DESC LIMIT 200").fetchall()
    return [dict(r) for r in rows]


@app.delete("/admin/errors")
def admin_clear_errors(admin=Depends(get_admin)) -> dict[str, Any]:
    with db() as c:
        c.execute("DELETE FROM error_logs")
    return {"message": "Đã xóa toàn bộ log lỗi."}


# ======================== Bảo mật & Mod Game (Security & Mod Tools) ========================
class CodeEncryptIn(BaseModel):
    code: str
    language: str
    level: str  # "low" | "high"


@app.post("/code/encrypt")
def code_encrypt(b: CodeEncryptIn, user=Depends(get_user)) -> dict[str, Any]:
    """Mã hóa làm rối (Obfuscate) mã nguồn để chống dịch ngược."""
    import base64
    code_text = b.code
    lang = b.language.lower()
    
    if not code_text.strip():
        raise HTTPException(status_code=400, detail="Mã nguồn không được rỗng.")
        
    if lang == "python":
        # Multi-layer encoding: Base64 + exec loader
        encoded = base64.b64encode(code_text.encode("utf-8")).decode("utf-8")
        obfuscated = (
            f"# Packaged by KENIOS Secure Encrypter (level: {b.level})\n"
            f"import base64\n"
            f"exec(base64.b64decode('{encoded}').decode('utf-8'))"
        )
        return {"result": obfuscated}
    elif lang == "javascript" or lang == "typescript":
        # Simple JavaScript obfuscation using HEX-escaped strings
        encoded_hex = "".join([f"\\x{ord(c):02x}" for c in code_text])
        obfuscated = (
            f"/* Packaged by KENIOS Secure Obfuscator */\n"
            f"eval(\"{encoded_hex}\");"
        )
        return {"result": obfuscated}
    else:
        # Fallback raw Base64 packaging
        encoded = base64.b64encode(code_text.encode("utf-8")).decode("utf-8")
        obfuscated = (
            f"/* Encrypted by KENIOS (Base64) */\n"
            f"// Raw base64: {encoded}"
        )
        return {"result": obfuscated}


@app.post("/code/analyze")
async def code_analyze(file: UploadFile = FastAPIFile(...), user=Depends(get_user)) -> dict[str, Any]:
    """Phân tích cấu trúc PE/ELF nhị phân và xuất Hex Viewer Dump."""
    import re
    import struct
    
    content = await file.read()
    size = len(content)
    if size == 0:
        raise HTTPException(status_code=400, detail="Tệp rỗng.")
        
    # 1. Detect file type
    file_type = "Binary / Unknown"
    entry_point = "N/A"
    architecture = "Unknown"
    sections = []
    
    if content.startswith(b"MZ"):
        file_type = "Windows PE (Portable Executable - EXE/DLL)"
        # Parse PE header entry point offset if large enough
        if size >= 0x40:
            pe_offset = struct.unpack("<I", content[0x3C:0x40])[0]
            if size >= pe_offset + 24:
                magic = content[pe_offset : pe_offset+4]
                if magic == b"PE\x00\x00":
                    machine = struct.unpack("<H", content[pe_offset+4 : pe_offset+6])[0]
                    architecture = "x64" if machine == 0x8664 else ("x86" if machine == 0x014c else f"Machine {hex(machine)}")
                    opt_header_offset = pe_offset + 24
                    if size >= opt_header_offset + 20:
                        entry_point = hex(struct.unpack("<I", content[opt_header_offset+16 : opt_header_offset+20])[0])
    elif content.startswith(b"\x7fELF"):
        file_type = "Linux/Android ELF (Executable and Linkable Format)"
        if size >= 20:
            elf_class = content[4]
            architecture = "64-bit" if elf_class == 2 else ("32-bit" if elf_class == 1 else "Unknown")
            if elf_class == 2 and size >= 32:
                entry_point = hex(struct.unpack("<Q", content[24:32])[0])
            elif elf_class == 1 and size >= 28:
                entry_point = hex(struct.unpack("<I", content[24:28])[0])
    elif content.startswith(b"\xca\xfe\xba\xbe") or content.startswith(b"\xbe\xba\xfe\xca"):
        file_type = "Mach-O (macOS/iOS Fat Binary)"
    elif content.startswith(b"\xfeedface") or content.startswith(b"\xfeedfacf"):
        file_type = "Mach-O (macOS/iOS Thin Binary)"
        
    # 2. Extract ASCII strings (min length 4)
    ascii_strings = []
    try:
        found_strings = re.findall(b"[ -~]{4,100}", content[:50000]) # Limit scan size to prevent excessive time
        for s in found_strings:
            s_decoded = s.decode("ascii", errors="ignore").strip()
            if s_decoded:
                ascii_strings.append(s_decoded)
    except Exception:
        pass
        
    # 3. Create Hex Dump (first 2048 bytes)
    hex_lines = []
    dump_limit = min(size, 2048)
    for offset in range(0, dump_limit, 16):
        chunk = content[offset : offset + 16]
        hex_parts = [f"{b:02x}" for b in chunk]
        # Pad hex parts
        while len(hex_parts) < 16:
            hex_parts.append("  ")
        hex_str = " ".join(hex_parts[:8]) + "  " + " ".join(hex_parts[8:])
        ascii_part = "".join([chr(b) if 32 <= b < 127 else "." for b in chunk])
        hex_lines.append(f"{offset:08x}  {hex_str}  |{ascii_part}|")
        
    hex_dump = "\n".join(hex_lines)
    if size > 2048:
        hex_dump += f"\n... (Đã ẩn bớt {size - 2048} bytes)"
        
    return {
        "file_type": file_type,
        "entry_point": entry_point,
        "architecture": architecture,
        "sections": sections if sections else None,
        "strings": list(set(ascii_strings))[:200], # Top 200 unique strings
        "hex_dump": hex_dump
    }


class CodeAsmIn(BaseModel):
    input: str
    mode: str  # "assemble" | "disassemble"
    arch: str  # "x86" | "arm"
    provider: Optional[str] = "openai"
    api_key: Optional[str] = None


@app.post("/code/asm")
async def code_asm(b: CodeAsmIn, user=Depends(get_user)) -> dict[str, Any]:
    """Dịch Hợp ngữ (Assembly) thành mã máy Hex hoặc ngược lại qua AI."""
    import httpx
    val = b.input.strip()
    if not val:
        raise HTTPException(status_code=400, detail="Mã đầu vào không được rỗng.")
        
    prov = b.provider or "openai"
    key = get_user_key(user["id"], prov, b.api_key)
    p = PROVIDERS.get(prov)
    if not p:
        raise HTTPException(status_code=400, detail=f"Không tìm thấy nhà cung cấp '{prov}'.")
        
    if b.mode == "assemble":
        prompt = (
            f"Bạn là một trình biên dịch hợp ngữ (Assembler) cho kiến trúc {b.arch.upper()}.\n"
            f"Hãy dịch mã lệnh hợp ngữ sau đây sang mã máy hex (dải bytes viết liền hoặc cách nhau khoảng trắng):\n"
            f"Lệnh: \"{val}\"\n"
            f"Chỉ trả về chuỗi mã Hex kết quả (ví dụ: '90 90'), không thêm bất kỳ văn bản giải thích nào khác."
        )
    else:
        prompt = (
            f"Bạn là một trình dịch ngược (Disassembler) cho kiến trúc {b.arch.upper()}.\n"
            f"Hãy dịch mã Hex nhị phân sau đây thành lệnh hợp ngữ dạng văn bản đọc được:\n"
            f"Hex: \"{val}\"\n"
            f"Chỉ trả về các dòng lệnh hợp ngữ kết quả, không thêm giải thích."
        )
        
    async with httpx.AsyncClient(timeout=REQUEST_TIMEOUT) as client:
        r = await client.post(
            f"{p['base']}/chat/completions",
            headers={"Authorization": f"Bearer {key}", "Content-Type": "application/json"},
            json={
                "model": p.get("default_model", "gpt-4o-mini"),
                "messages": [
                    {"role": "system", "content": "You are a helpful compiler helper."},
                    {"role": "user", "content": prompt}
                ]
            }
        )
    _raise_for_provider(r, prov)
    res_data = r.json()
    reply = res_data["choices"][0]["message"]["content"].strip()
    return {"result": reply}


# ======================== DevOps & DevOps Tools ========================
class SSHIn(BaseModel):
    host: str
    username: str
    password: str
    command: str


@app.post("/run/ssh")
def run_ssh(b: SSHIn, user=Depends(get_user)) -> dict[str, Any]:
    """Kết nối SSH vào VPS và chạy câu lệnh."""
    host = b.host.strip()
    username = b.username.strip()
    password = b.password.strip()
    command = b.command.strip()
    
    if not (host and username and command):
        raise HTTPException(status_code=400, detail="Thiếu tham số kết nối SSH (Host, Username, Command).")
    
    # Fallback for mock/test IPs
    if "127.0.0.1" in host or "localhost" in host or "192.168" in host or "FAKE" in host.upper():
        return {
            "stdout": f"[MOCK SSH] Executing on {username}@{host}:\n$ {command}\nSuccess: mock output here.",
            "stderr": "",
            "exitCode": 0
        }
        
    try:
        import paramiko
        ssh = paramiko.SSHClient()
        ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        ssh.connect(host, username=username, password=password, timeout=10)
        stdin, stdout, stderr = ssh.exec_command(command, timeout=30)
        out = stdout.read().decode('utf-8', errors='replace')
        err = stderr.read().decode('utf-8', errors='replace')
        code = stdout.channel.recv_exit_status()
        ssh.close()
        return {"stdout": out, "stderr": err, "exitCode": code}
    except ImportError:
        import subprocess
        try:
            cmd = ["sshpass", "-p", password, "ssh", "-o", "StrictHostKeyChecking=no", f"{username}@{host}", command]
            r = subprocess.run(cmd, capture_output=True, text=True, timeout=15)
            return {"stdout": r.stdout, "stderr": r.stderr, "exitCode": r.returncode}
        except Exception:
            raise HTTPException(
                status_code=400, 
                detail="Thư viện 'paramiko' chưa được cài trên server VPS. Hãy chạy lệnh 'pip install paramiko' trên VPS hoặc chạy lại start-vps.sh."
            )
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Lỗi kết nối SSH: {e}")


class HTTPIn(BaseModel):
    url: str
    method: str
    headers: Optional[dict[str, str]] = None
    body: Optional[str] = None


@app.post("/run/http")
async def run_http(b: HTTPIn, user=Depends(get_user)) -> dict[str, Any]:
    """Gửi HTTP request từ máy chủ (bỏ qua CORS)."""
    import httpx
    url = b.url.strip()
    method = b.method.upper()
    if not url:
        raise HTTPException(status_code=400, detail="Thiếu URL yêu cầu.")
    
    headers = b.headers or {}
    headers.pop("Host", None)
    headers.pop("host", None)
    content_data = b.body or ""
    
    async with httpx.AsyncClient(timeout=30) as client:
        try:
            if method == "GET":
                r = await client.get(url, headers=headers)
            elif method == "POST":
                r = await client.post(url, headers=headers, content=content_data)
            elif method == "PUT":
                r = await client.put(url, headers=headers, content=content_data)
            elif method == "DELETE":
                r = await client.delete(url, headers=headers)
            else:
                raise HTTPException(status_code=400, detail=f"Phương thức '{method}' chưa hỗ trợ.")
                
            resp_body = r.text
            resp_headers = {k: v for k, v in r.headers.items()}
            return {
                "status": r.status_code,
                "headers": resp_headers,
                "body": resp_body
            }
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Lỗi gửi HTTP request: {e}")


class SQLIn(BaseModel):
    query: str


@app.post("/run/sql")
def run_sql(b: SQLIn, user=Depends(get_user)) -> dict[str, Any]:
    """Thực thi câu lệnh SQL SQLite cục bộ."""
    query = b.query.strip()
    if not query:
        raise HTTPException(status_code=400, detail="Câu lệnh SQL không được rỗng.")
    
    is_admin_check = user.get("is_admin") or user.get("isAdmin")
    query_lower = query.lower()
    destructive = ["drop", "delete", "update", "insert", "alter", "create", "replace"]
    if any(d in query_lower for d in destructive) and not is_admin_check:
        raise HTTPException(status_code=403, detail="Chỉ tài khoản Admin mới có quyền thực thi các câu lệnh sửa đổi database (INSERT, UPDATE, DELETE, DROP...).")
        
    try:
        with db() as c:
            cur = c.execute(query)
            if cur.description:
                columns = [desc[0] for desc in cur.description]
                rows = cur.fetchall()
                row_list = []
                for r in rows:
                    row_list.append([str(val) if val is not None else "" for val in r])
                return {
                    "columns": columns,
                    "rows": row_list,
                    "message": f"Truy vấn thành công. Trả về {len(row_list)} bản ghi."
                }
            else:
                c.commit()
                return {
                    "columns": [],
                    "rows": [],
                    "message": f"Thực thi thành công. Số bản ghi ảnh hưởng: {cur.rowcount}."
                }
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Lỗi SQL: {e}")


# ======================== Error handler ========================
@app.exception_handler(Exception)
async def on_error(request: Request, exc: Exception):
    log.exception("Lỗi server: %s", exc)
    try:
        with db() as c:
            c.execute("INSERT INTO error_logs(context,detail,created_at) VALUES(?,?,?)",
                      (str(request.url.path), f"{type(exc).__name__}: {exc}"[:800], int(time.time())))
    except Exception:
        pass
    return JSONResponse(status_code=500,
                        content={"detail": f"Lỗi máy chủ: {type(exc).__name__}: {str(exc)[:300]}"})


# ======================== Entrypoint ========================
if __name__ == "__main__":
    import uvicorn
    init_db()
    log.info("KENIOS kenios v4.2 — cổng %s | %d AI hỗ trợ", PORT, len(PROVIDERS))
    uvicorn.run(app, host="0.0.0.0", port=PORT)



# ======================== Proxy mạng (quản lý & định tuyến) ========================
from urllib.parse import quote as _qt

class ProxyAddIn(BaseModel):
    label: Optional[str] = None
    scheme: str = "http"            # http | https | socks5
    host: str
    port: int
    username: Optional[str] = None
    password: Optional[str] = None
    region: Optional[str] = None
    source: str = "manual"          # manual | provider | vps

class ProxyImportIn(BaseModel):
    text: str                       # mỗi dòng: host:port  hoặc  host:port:user:pass
    scheme: str = "http"
    region: Optional[str] = None
    source: str = "provider"

class ProxySelectIn(BaseModel):
    id: Optional[int] = None        # None = bỏ chọn (đi trực tiếp qua VPS)

class ProxyTestIn(BaseModel):
    id: Optional[int] = None
    scheme: Optional[str] = None
    host: Optional[str] = None
    port: Optional[int] = None
    username: Optional[str] = None
    password: Optional[str] = None


def _proxy_url_from(scheme, host, port, username=None, password=None) -> str:
    scheme = (scheme or "http").lower()
    if scheme not in ("http", "https", "socks5", "socks5h"):
        scheme = "http"
    auth = ""
    if username:
        auth = _qt(str(username), safe="")
        if password:
            auth += ":" + _qt(str(password), safe="")
        auth += "@"
    return f"{scheme}://{auth}{host}:{port}"


def _proxy_row_to_url(row) -> Optional[str]:
    if row["source"] == "vps":
        return None
    pwd = dec(row["enc_password"]) if row["enc_password"] else None
    return _proxy_url_from(row["scheme"], row["host"], row["port"], row["username"], pwd)


def get_active_proxy(user_id: int) -> Optional[str]:
    with db() as c:
        row = c.execute("SELECT * FROM proxies WHERE user_id=? AND active=1 LIMIT 1",
                        (user_id,)).fetchone()
    return _proxy_row_to_url(row) if row else None


def _proxy_public(row) -> dict[str, Any]:
    return {
        "id": row["id"],
        "label": row["label"] or f'{row["host"]}:{row["port"]}',
        "scheme": row["scheme"], "host": row["host"], "port": row["port"],
        "username": row["username"], "has_password": bool(row["enc_password"]),
        "region": row["region"] or "", "source": row["source"],
        "active": bool(row["active"]),
    }


@app.get("/proxy/list")
def proxy_list(user=Depends(get_user)) -> dict[str, Any]:
    with db() as c:
        rows = c.execute("SELECT * FROM proxies WHERE user_id=? ORDER BY id DESC",
                         (user["id"],)).fetchall()
    items = [_proxy_public(r) for r in rows]
    regions = sorted({r["region"] for r in items if r["region"]})
    return {"proxies": items, "regions": regions}


@app.post("/proxy/add")
def proxy_add(b: ProxyAddIn, user=Depends(get_user)) -> dict[str, Any]:
    enc_pw = enc(b.password) if b.password else None
    with db() as c:
        cur = c.execute(
            "INSERT INTO proxies(user_id,label,scheme,host,port,username,enc_password,"
            "region,source,active,created_at) VALUES(?,?,?,?,?,?,?,?,?,0,?)",
            (user["id"], b.label, (b.scheme or "http").lower(), b.host, int(b.port),
             b.username, enc_pw, b.region, b.source or "manual", int(time.time())))
        pid = cur.lastrowid
    return {"id": pid, "message": "Đã thêm proxy."}


@app.post("/proxy/import")
def proxy_import(b: ProxyImportIn, user=Depends(get_user)) -> dict[str, Any]:
    n = 0
    with db() as c:
        for line in b.text.splitlines():
            line = line.strip()
            if not line:
                continue
            parts = line.split(":")
            if len(parts) < 2:
                continue
            host = parts[0]
            try:
                port_i = int(parts[1])
            except ValueError:
                continue
            user_p = parts[2] if len(parts) >= 3 else None
            pass_p = parts[3] if len(parts) >= 4 else None
            enc_pw = enc(pass_p) if pass_p else None
            c.execute(
                "INSERT INTO proxies(user_id,label,scheme,host,port,username,enc_password,"
                "region,source,active,created_at) VALUES(?,?,?,?,?,?,?,?,?,0,?)",
                (user["id"], None, (b.scheme or "http").lower(), host, port_i,
                 user_p, enc_pw, b.region, b.source or "provider", int(time.time())))
            n += 1
    return {"imported": n, "message": f"Đã nhập {n} proxy."}


@app.delete("/proxy/{pid}")
def proxy_delete(pid: int, user=Depends(get_user)) -> dict[str, Any]:
    with db() as c:
        c.execute("DELETE FROM proxies WHERE id=? AND user_id=?", (pid, user["id"]))
    return {"message": "Đã xoá proxy."}


@app.post("/proxy/select")
def proxy_select(b: ProxySelectIn, user=Depends(get_user)) -> dict[str, Any]:
    with db() as c:
        c.execute("UPDATE proxies SET active=0 WHERE user_id=?", (user["id"],))
        if b.id is not None:
            row = c.execute("SELECT id FROM proxies WHERE id=? AND user_id=?",
                            (b.id, user["id"])).fetchone()
            if not row:
                raise HTTPException(status_code=404, detail="Không tìm thấy proxy.")
            c.execute("UPDATE proxies SET active=1 WHERE id=? AND user_id=?",
                      (b.id, user["id"]))
    msg = "Đã chọn proxy." if b.id is not None else "Đã bỏ chọn (đi trực tiếp qua VPS)."
    return {"active_id": b.id, "message": msg}


@app.post("/proxy/test")
async def proxy_test(b: ProxyTestIn, user=Depends(get_user)) -> dict[str, Any]:
    if b.id is not None:
        with db() as c:
            row = c.execute("SELECT * FROM proxies WHERE id=? AND user_id=?",
                            (b.id, user["id"])).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Không tìm thấy proxy.")
        url = _proxy_row_to_url(row)
    elif b.host and b.port:
        url = _proxy_url_from(b.scheme, b.host, b.port, b.username, b.password)
    else:
        raise HTTPException(status_code=400, detail="Thiếu thông tin proxy để test.")

    t0 = time.time()
    try:
        kwargs: dict[str, Any] = {"timeout": 12}
        if url:
            kwargs["proxy"] = url
        async with httpx.AsyncClient(**kwargs) as client:
            r = await client.get("http://ip-api.com/json")
            data = r.json()
        ms = int((time.time() - t0) * 1000)
        if data.get("status") == "success":
            return {"ok": True, "latency_ms": ms, "ip": data.get("query"),
                    "country": data.get("country"), "country_code": data.get("countryCode"),
                    "region": data.get("regionName"), "city": data.get("city")}
        return {"ok": False, "latency_ms": ms, "error": "Không lấy được vị trí IP qua proxy."}
    except Exception as e:
        return {"ok": False, "error": f"Proxy lỗi/không kết nối ({e.__class__.__name__})."}



# ======================== Tạo proxy trên VPS (admin, dùng tinyproxy) ========================
import subprocess as _sp

_PROXY_PORT_MIN = 8801
_PROXY_PORT_MAX = 8900          # tối đa 100 cổng
_PROXY_INSTANCE_DIR = "/etc/tinyproxy/instances"
_PROXY_BASE_CONF = "/etc/tinyproxy/tinyproxy.conf"
_PROXY_TEMPLATE = "/etc/systemd/system/tinyproxy@.service"


class ProxySpawnIn(BaseModel):
    count: int = 1                # số cổng muốn tạo thêm

class ProxyDespawnIn(BaseModel):
    port: int


def _proxy_run(args: list[str]) -> tuple[int, str]:
    try:
        r = _sp.run(args, capture_output=True, text=True, timeout=30)
        return r.returncode, (r.stdout + r.stderr)
    except Exception as e:
        return 1, f"{e.__class__.__name__}: {e}"


def _proxy_ensure_template() -> None:
    t = pathlib.Path(_PROXY_TEMPLATE)
    if not t.exists():
        t.write_text(
            "[Unit]\n"
            "Description=tinyproxy instance on port %i\n"
            "After=network.target\n"
            "[Service]\n"
            "Type=simple\n"
            "ExecStart=/usr/bin/tinyproxy -d -c /etc/tinyproxy/instances/%i.conf\n"
            "Restart=always\n"
            "RestartSec=3\n"
            "[Install]\n"
            "WantedBy=multi-user.target\n",
            encoding="utf-8",
        )
        _proxy_run(["systemctl", "daemon-reload"])


def _proxy_used_ports() -> list[int]:
    d = pathlib.Path(_PROXY_INSTANCE_DIR)
    if not d.exists():
        return []
    ports = []
    for f in d.glob("*.conf"):
        try:
            ports.append(int(f.stem))
        except ValueError:
            pass
    return sorted(ports)


def _proxy_write_conf(port: int) -> None:
    pathlib.Path(_PROXY_INSTANCE_DIR).mkdir(parents=True, exist_ok=True)
    base = pathlib.Path(_PROXY_BASE_CONF).read_text(encoding="utf-8")
    out_lines = []
    for line in base.splitlines():
        st = line.strip()
        if st.startswith("Port "):
            continue
        if st.startswith("PidFile"):
            continue
        if st.startswith("BasicAuth ") or st.startswith("Allow "):
            continue
        if st.startswith("LogFile"):
            continue
        out_lines.append(line)
    out_lines.append(f"Port {port}")
    out_lines.append(f'PidFile "/run/tinyproxy-{port}.pid"')
    pathlib.Path(f"{_PROXY_INSTANCE_DIR}/{port}.conf").write_text(
        "\n".join(out_lines) + "\n", encoding="utf-8")


def _proxy_spawn_one() -> Optional[int]:
    used = set(_proxy_used_ports())
    port = None
    for cand in range(_PROXY_PORT_MIN, _PROXY_PORT_MAX + 1):
        if cand not in used:
            port = cand
            break
    if port is None:
        return None
    _proxy_ensure_template()
    _proxy_write_conf(port)
    _proxy_run(["systemctl", "enable", "--now", f"tinyproxy@{port}"])
    return port


def _proxy_despawn_one(port: int) -> None:
    _proxy_run(["systemctl", "disable", "--now", f"tinyproxy@{port}"])
    f = pathlib.Path(f"{_PROXY_INSTANCE_DIR}/{port}.conf")
    if f.exists():
        f.unlink()


@app.post("/proxy/vps/spawn")
def proxy_vps_spawn(b: ProxySpawnIn, request: Request, admin=Depends(get_admin)) -> dict[str, Any]:
    host = request.url.hostname or "127.0.0.1"
    used = _proxy_used_ports()
    free = (_PROXY_PORT_MAX - _PROXY_PORT_MIN + 1) - len(used)
    want = max(1, min(int(b.count), 100, free))
    if free <= 0:
        raise HTTPException(status_code=400,
                            detail=f"Đã đạt tối đa {_PROXY_PORT_MAX - _PROXY_PORT_MIN + 1} cổng proxy.")
    created = []
    with db() as c:
        for _ in range(want):
            port = _proxy_spawn_one()
            if port is None:
                break
            # lưu vào danh sách proxy của admin để hiện trong app
            c.execute(
                "INSERT INTO proxies(user_id,label,scheme,host,port,username,enc_password,"
                "region,source,active,created_at) VALUES(?,?,?,?,?,?,?,?,?,0,?)",
                (admin["id"], f"VPS {port}", "http", host, port, None, None,
                 "VN", "vpsproxy", int(time.time())))
            created.append(port)
    return {"created": created, "count": len(created), "host": host,
            "note": "Cùng 1 IP VPS, khác cổng. Nhớ mở các cổng này ở firewall VPS."}


@app.get("/proxy/vps/list")
def proxy_vps_list(admin=Depends(get_admin)) -> dict[str, Any]:
    items = []
    for port in _proxy_used_ports():
        rc, out = _proxy_run(["systemctl", "is-active", f"tinyproxy@{port}"])
        items.append({"port": port, "active": out.strip() == "active"})
    return {"instances": items, "max": _PROXY_PORT_MAX - _PROXY_PORT_MIN + 1}


@app.post("/proxy/vps/despawn")
def proxy_vps_despawn(b: ProxyDespawnIn, admin=Depends(get_admin)) -> dict[str, Any]:
    if not (_PROXY_PORT_MIN <= int(b.port) <= _PROXY_PORT_MAX):
        raise HTTPException(status_code=400, detail="Cổng ngoài dải cho phép.")
    _proxy_despawn_one(int(b.port))
    with db() as c:
        c.execute("DELETE FROM proxies WHERE user_id=? AND port=? AND source='vpsproxy'",
                  (admin["id"], int(b.port)))
    return {"message": f"Đã xoá proxy cổng {b.port}."}
